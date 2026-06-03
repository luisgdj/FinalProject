import sys
from datetime import datetime

from flask import Flask
import os
import torch
from transformers import AutoModelForCausalLM, AutoTokenizer
import json
import csv
import re

app = Flask(__name__)

MODEL = None
TOKENIZER = None
DATOS_TRAIN = None
STATS = None

def extraer_caracteristicas(smiles):
    return {
        "length": len(smiles),
        "num_rings": sum(c.isdigit() for c in smiles),
        "num_branches": smiles.count("("),
        "double_bonds": smiles.count("="),
        "triple_bonds": smiles.count("#"),
        "aromatic_atoms": sum(smiles.count(c) for c in "cnosp"),
        "num_N": smiles.count("N") + smiles.count("n"),
        "num_O": smiles.count("O") + smiles.count("o"),
        "num_S": smiles.count("S") + smiles.count("s"),
        "num_P": smiles.count("P"),
        "num_F": smiles.count('F'),
        "num_Cl": smiles.count('Cl'),
        "num_Br": smiles.count('Br'),
        "stereocenters": smiles.count("@"),
        "net_charge": smiles.count("+") - smiles.count("-"),
    }


def leer_csv(filepath):
    datos = []
    with open(filepath, 'r', encoding='utf-8') as f:
        reader = csv.DictReader(f)
        for row in reader:
            try:
                datos.append({
                    'smiles': row['smiles'].strip(),
                    'adduct': row['Adduct'].strip(),
                    'ccs': float(row['CCS_AVG'])
                })
            except (ValueError, KeyError):
                continue
    return datos


def analizar_datos(datos):
    ccs_values = [d['ccs'] for d in datos]
    stats = {
        'ccs_min': min(ccs_values),
        'ccs_max': max(ccs_values),
        'ccs_avg': sum(ccs_values) / len(ccs_values)
    }
    return stats


def normalizar_aducto(adduct):
    # Elimina la carga final del aducto para unificar formatos.
    return adduct.rstrip('+-').strip()

# Selecciona N moléculas del dataset estructuralmente similares al input.
def seleccionar_ejemplos(datos, smiles_input, adduct_input, n):

    # Paso 1: Calcula los descriptores del compuesto de entrada.
    caract_input = extraer_caracteristicas(smiles_input)
    adduct_norm = normalizar_aducto(adduct_input)

    # Paso 2: Filtra el training set por aducto.
    datos_filtrados = [d for d in datos if normalizar_aducto(d['adduct']) == adduct_norm]
    # Excepción: Si no tienes suficientes ejemplos del mismo aducto, se usa el dataset completo.
    if len(datos_filtrados) < n:
        print(f"Aviso: solo {len(datos_filtrados)} ejemplos para aducto {adduct_norm}, usando dataset completo")
        datos_filtrados = datos

    similitudes = []
    for d in datos_filtrados:
        caract = extraer_caracteristicas(d['smiles'])

        # Paso 3: Calcula una puntuación de similitud para cada descritor estructural.
        #  - Para cada molécula del conjunto filtrado
        #  - Compara los descriptores de la molecula con los del compuesto de entrada
        #  - Calcula una puntuación de 0 a 1
        sim_longitud   = 1 / (1 + abs(caract['length']    - caract_input['length']) / 10)
        sim_anillos    = 1 / (1 + abs(caract['num_rings'] - caract_input['num_rings']))
        sim_aromatico  = 1 / (1 + abs(caract['aromatic_atoms'] - caract_input['aromatic_atoms']) / 3)

        # Paso 4: Combina las tres puntuaciones con pesos (estructura > longitud > aromaticidad).
        similitud = 0.50 * sim_anillos + 0.30 * sim_longitud + 0.20 * sim_aromatico
        similitudes.append((similitud, d))

    # Paso 5: Ordena por puntuación y devuelve los 'n' mejores.
    similitudes.sort(reverse=True, key=lambda x: x[0])
    return [d for _, d in similitudes[:n]]


def buscar_en_dataset(smiles, adduct, dataset):
    adduct_norm = normalizar_aducto(adduct)
    for row in dataset:
        if row["smiles"] == smiles and normalizar_aducto(row["Adduct"]) == adduct_norm:
            return row["ccs"]
    return None


ADDUCT_INFO = {
    '[M+H]':     {'charge': 1,  'mass_add': 1.007,  'effect': 'protonated, baseline reference'},
    '[M+Na]':    {'charge': 1,  'mass_add': 22.989, 'effect': 'sodium adduct, ~3-5 Å² larger than [M+H]+'},
    '[M-H]':     {'charge': -1, 'mass_add': -1.007, 'effect': 'deprotonated, ~2-5 Å² smaller than [M+H]+'},
}

# Iteración 1: Monta un prompt simple solo usando la información del compuesto a analizar
def construir_prompt_simple(smiles, adduct):
    adduct_norm = normalizar_aducto(adduct)
    info = ADDUCT_INFO.get(adduct_norm, {
        'charge': 1, 'mass_add': 0, 'effect': 'unknown adduct type'
    })

    prompt = f"""You are an expert in ion mobility mass spectrometry.
Your task is to predict the Collision Cross Section (CCS, in Å²) of the molecule from its SMILES structure and adduct.

To estimate CCS, follow these steps:
  1. Read the SMILES and identify rings, chains, branches, and functional groups.
  2. Decide if the structure is compact (mostly fused rings, rigid) or extended (chains, flexible).
  3. Apply the adduct correction described in the molecule line below.
  4. Output a single CCS value, typically in the range 100-300 Å² for small molecules.

Molecule:
   SMILES: {smiles}
   Adduct: {adduct_norm} ({info['effect']})

You MUST end your answer with the line "Final CCS: <number>", do not leave the number blank.
<think>
Let me apply the 4 steps to estimate the CCS.

Step 1 (size): """
    return prompt

# Iteración 2: Monta un prompt usando conocimiento experto de metabolómica y la información del compuesto a analizar
def construir_prompt_ce(smiles, adduct):
    adduct_norm = normalizar_aducto(adduct)
    info = ADDUCT_INFO.get(adduct_norm, {
        'charge': 1, 'mass_add': 0, 'effect': 'unknown adduct type'
    })

    prompt = f"""You are an expert in ion mobility mass spectrometry.
Your task is to predict the Collision Cross Section (CCS, in Å²) of the molecule from its SMILES structure and adduct.

CCS reflects the rotationally-averaged area a molecule presents while colliding with buffer gas in an ion mobility cell. It is determined primarily by the three-dimensional shape of the ion, not by its mass.

Expert knowledge - structural factors that determine CCS:

1. Molecular size and rigidity:
   - Compact, rigid structures (fused aromatic systems, polycyclic cores) yield LOWER CCS for their mass.
   - Extended, flexible structures (long aliphatic chains, single-bond rotations) yield HIGHER CCS for their mass.

2. Branching and substitution:
   - Branched molecules pack more compactly than linear isomers of the same mass → lower CCS.
   - Bulky substituents (long side chains, multiple rings off a central atom) increase CCS.

3. Heteroatoms and functional groups:
   - Heteroatoms (N, O, S) and halogens add mass but contribute relatively little to molecular volume.
   - Polar groups (-OH, -NH2, -COOH) can promote intramolecular interactions, slightly reducing CCS.

4. Adduct effect:
   - The adduct alters charge state and ion geometry.
   - Use the adduct effect described in the molecule line below as a small final adjustment.

To estimate CCS, follow these steps:
  1. Read the SMILES and identify rings, chains, branches, and functional groups.
  2. Decide if the structure is compact (mostly fused rings, rigid) or extended (chains, flexible).
  3. Apply the adduct correction described in the molecule line below.
  4. Output a single CCS value, typically in the range 100-300 Å² for small molecules.

Molecule:
  SMILES: {smiles}
  Adduct: {adduct_norm} ({info['effect']})

You MUST end your answer with the line "Final CCS: <number>", do not leave the number blank.
<think>
Let me apply the 4 steps to estimate the CCS.

Step 1 (structure): """
    return prompt

# Iteración 3: Monta un prompt usando ejemplos de una base de datos y la información del compuesto a analizar
def construir_prompt_ejemplos(smiles, adduct, DATOS_TRAIN, n):
    adduct_norm = normalizar_aducto(adduct)
    info = ADDUCT_INFO.get(adduct_norm, {
        'charge': 1, 'mass_add': 0, 'effect': 'unknown adduct type'
    })

    # Seleccionar N moléculas estructuralmente similares del dataset
    ejemplos = seleccionar_ejemplos(DATOS_TRAIN, smiles, adduct, n)

    # Construir bloque de ejemplos (mismo formato que el target: SMILES + Adduct + CCS)
    ejemplos_texto = ""
    for i, ej in enumerate(ejemplos, 1):
        ej_adduct_norm = normalizar_aducto(ej['adduct'])
        ejemplos_texto += (
            f"  Example {i}:\n"
            f"    SMILES: {ej['smiles']}\n"
            f"    Adduct: {ej_adduct_norm}\n"
            f"    CCS: {ej['ccs']:.2f} Å²\n\n"
        )

    prompt = f"""You are an expert in ion mobility mass spectrometry.
Your task is to predict the Collision Cross Section (CCS, in Å²) of the molecule from its SMILES structure and adduct.

Below are {n} reference molecules with experimentally measured CCS values, selected for their structural similarity to the target. Use them to identify patterns and infer the CCS of the target molecule.

Reference molecules:

{ejemplos_texto}To estimate CCS, follow these steps:
  1. Read the SMILES and identify rings, chains, branches, and functional groups.
  2. Decide if the structure is compact (mostly fused rings, rigid) or extended (chains, flexible).
  3. Compare with the reference molecules above. Identify the most similar ones and use their CCS values as anchors.
  4. Apply the adduct correction described in the molecule line below.
  5. Output a single CCS value, typically in the range 100-300 Å² for small molecules.

Molecule:
   SMILES: {smiles}
   Adduct: {adduct_norm} ({info['effect']})

You MUST end your answer with the line "Final CCS: <number>", do not leave the number blank.
<think>
Let me apply the 5 steps to estimate the CCS.

Step 1 (size): """
    return prompt

# Iteración 4: Monta un prompt usando conocimiento experto, ejemplos de una base de datos y la información del compuesto a analizar
def construir_prompt_completo(smiles, adduct, DATOS_TRAIN, n):
    adduct_norm = normalizar_aducto(adduct)
    info = ADDUCT_INFO.get(adduct_norm, {
        'charge': 1, 'mass_add': 0, 'effect': 'unknown adduct type'
    })

    # Seleccionar N moléculas estructuralmente similares del dataset
    ejemplos = seleccionar_ejemplos(DATOS_TRAIN, smiles, adduct, n)

    # Construir bloque de ejemplos (mismo formato que el target: SMILES + Adduct + CCS)
    ejemplos_texto = ""
    for i, ej in enumerate(ejemplos, 1):
        ej_adduct_norm = normalizar_aducto(ej['adduct'])
        ejemplos_texto += (
            f"  Example {i}:\n"
            f"    SMILES: {ej['smiles']}\n"
            f"    Adduct: {ej_adduct_norm}\n"
            f"    CCS: {ej['ccs']:.2f} Å²\n\n"
        )

    prompt = f"""You are an expert in ion mobility mass spectrometry.
Your task is to predict the Collision Cross Section (CCS, in Å²) of the molecule from its SMILES structure and adduct.

CCS reflects the rotationally-averaged area a molecule presents while colliding with buffer gas in an ion mobility cell. It is determined primarily by the three-dimensional shape of the ion, not by its mass.

Expert knowledge - structural factors that determine CCS:

1. Molecular size and rigidity:
   - Compact, rigid structures (fused aromatic systems, polycyclic cores) yield LOWER CCS for their mass.
   - Extended, flexible structures (long aliphatic chains, single-bond rotations) yield HIGHER CCS for their mass.

2. Branching and substitution:
   - Branched molecules pack more compactly than linear isomers of the same mass → lower CCS.
   - Bulky substituents (long side chains, multiple rings off a central atom) increase CCS.

3. Heteroatoms and functional groups:
   - Heteroatoms (N, O, S) and halogens add mass but contribute relatively little to molecular volume.
   - Polar groups (-OH, -NH2, -COOH) can promote intramolecular interactions, slightly reducing CCS.

4. Adduct effect:
   - The adduct alters charge state and ion geometry.
   - Use the adduct effect described in the molecule line below as a small final adjustment.

To estimate CCS, follow these steps:
  1. Read the SMILES and identify rings, chains, branches, and functional groups.
  2. Decide if the structure is compact (mostly fused rings, rigid) or extended (chains, flexible).
  3. Apply the adduct correction described in the molecule line below.
  4. Output a single CCS value, typically in the range 100-300 Å² for small molecules.

Below are {n} reference molecules with experimentally measured CCS values, selected for their structural similarity to the target. Use them to identify patterns and infer the CCS of the target molecule.

Reference molecules:

{ejemplos_texto}To estimate CCS, follow these steps:
  1. Read the SMILES and identify rings, chains, branches, and functional groups.
  2. Decide if the structure is compact (mostly fused rings, rigid) or extended (chains, flexible).
  3. Compare with the reference molecules above. Identify the most similar ones and use their CCS values as anchors.
  4. Apply the adduct correction described in the molecule line below.
  5. Output a single CCS value, typically in the range 100-300 Å² for small molecules.

Molecule:
   SMILES: {smiles}
   Adduct: {adduct_norm} ({info['effect']})

You MUST end your answer with the line "Final CCS: <number>", do not leave the number blank.
<think>
Let me apply the 5 steps to estimate the CCS.

Step 1 (size): """
    return prompt


# Extrae el CCS predicho de la respuesta del modelo
def parsear_respuesta(respuesta_raw):

    # Limpieza inicial
    # Quitamos markdown bold/italic que ensucia las regex
    texto = respuesta_raw.replace('**', '').replace('*', '')
    # Quitamos artefactos LaTeX comunes: $...$, $$...$$, \[ \], \( \)
    texto = re.sub(r'\$+', '', texto)
    texto = re.sub(r'\\[\[\]\(\)]', '', texto)
    texto = texto.strip() # Normalizamos espacios

    # Comprobación del bloque <think>
    if '<think>' in texto and '</think>' not in texto:
        return {'predicted_ccs': None, 'fallback': True, 'source': 'think_unclosed'}
    if '</think>' not in texto:
        return {'predicted_ccs': None, 'fallback': True, 'source': 'no_think_tag'}

    # A partir de aquí trabajamos SOLO con la zona post-think
    texto_post = texto.split('</think>')[-1].strip()

    def aceptar(valor, source):
        if 50 <= valor <= 500: # rango razonable para CCS
            return {'predicted_ccs': round(valor, 2), 'fallback': False, 'source': source}
        return None

    NUM = r'([0-9]+(?:\.[0-9]+)?)'

    # ESTRATEGIA 1: "Final CCS: <número>" en variantes (con/sin \boxed{})
    patrones_final = [
        rf'Final\s+CCS\s*[:=]?\s*\\?boxed\{{?\s*{NUM}\s*\}}?', # "Final CCS: \boxed{200}" o "Final CCS: 200"
        rf'Final\s+CCS\s*[:=]\s*{NUM}', # "Final CCS: 200" o "Final CCS = 200"
        rf'Final\s+CCS\s+{NUM}' # "Final CCS 200" (sin separador)
    ]
    for patron in patrones_final:
        match = re.search(patron, texto_post, re.IGNORECASE)
        if match:
            valor = float(match.group(1))
            result = aceptar(valor, 'final_ccs_tag')
            if result:
                return result

    # ESTRATEGIA 2: \boxed{<número>} sin "Final CCS:" delante
    match = re.search(rf'\\?boxed\{{?\s*{NUM}\s*\}}?', texto_post)
    if match:
        valor = float(match.group(1))
        result = aceptar(valor, 'boxed_tag')
        if result:
            return result

    # ESTRATEGIA 3: último número plausible en el rango CCS
    matches = re.findall(r'\b([0-9]{2,3}(?:\.[0-9]+)?)\b', texto_post)
    plausibles = [float(m) for m in matches if 50 <= float(m) <= 500]
    if plausibles:
        return {
            'predicted_ccs': round(plausibles[-1], 2),
            'fallback': False,
            'source': 'last_plausible',
        }

    return {'predicted_ccs': None, 'fallback': True, 'source': 'no_number_found'}

# Clasifica si la predicción es válida o es un valor degenerado.
def clasificar_prediccion(ccs_pred, ejemplos, stats):

    ccs_referencias = {round(ej['ccs'], 2) for ej in ejemplos}
    avg = round(stats['ccs_avg'])

    if abs(ccs_pred - avg) <= 0.5:
        return 'dataset_avg', False

    if any(abs(ccs_pred - ref) <= 0.02 for ref in ccs_referencias):
        return 'exact_copy', False

    return 'interpolated', True


def predecir_ccs(model, tokenizer, prompt, stats, ejemplos):

    inputs = tokenizer(prompt, return_tensors="pt", truncation=True, max_length=5000)

    with torch.inference_mode():
        outputs = model.generate(
            **inputs,
            do_sample = False,           # 1.5B -> False ; 7B/14B -> True # True para el 7B
            temperature = None,           # 1.5B -> None  ; 7B/14B -> 0.3 # Bajo para mantener precisión
            top_p = None,                 # 1.5B -> None  ; 7B/14B -> 0.9
            repetition_penalty = 1.15,   # 1.5B -> None  ; 7B/14B -> 1.15 # Penaliza repeticiones
            max_new_tokens = 10000, # más margen para el bloque <reasoning>
            pad_token_id = tokenizer.eos_token_id
        )

    n_tokens_prompt = inputs['input_ids'].shape[1]
    n_tokens_total = outputs.shape[1]
    n_tokens_generados = n_tokens_total - n_tokens_prompt

    print(f" - Tokens del prompt: {n_tokens_prompt}")
    print(f" - Tokens del output total: {n_tokens_total}")
    print(f" - Tokens generados: {n_tokens_generados}")

    respuesta_completa = tokenizer.decode(outputs[0], skip_special_tokens=True)
    prompt_texto = tokenizer.decode(inputs['input_ids'][0], skip_special_tokens=True)
    respuesta = respuesta_completa[len(prompt_texto):].strip()

    print(" RESPUESTA COMPLETA: " + json.dumps(respuesta_completa))
    print(" RESPUESTA CRUDA: " + json.dumps(respuesta))

    resultado = parsear_respuesta(respuesta)

    # Fallback por fallo de parseo
    if resultado['fallback']:
        resultado['predicted_ccs'] = 0.0
        resultado['pred_type'] = 'heuristic_fallback'
        resultado['reasoning'] = "Heuristic fallback: parser found no valid number"
        return resultado

    # Clasificar la predicción
    tipo, es_valida = clasificar_prediccion(resultado['predicted_ccs'], ejemplos, stats)
    resultado['predicted_ccs_raw'] = resultado['predicted_ccs']
    resultado['pred_type'] = tipo

    if es_valida:
        resultado['reasoning'] = f"Model interpolation ({resultado['predicted_ccs']})"
    else:
        resultado['pred_type'] = tipo  # 'exact_copy' o 'dataset_avg', sin →heuristic

        if tipo == 'exact_copy':
            resultado['reasoning'] = f"Model output accepted (reference copy: {resultado['predicted_ccs_raw']})"
        else:
            resultado['reasoning'] = f"Model output accepted (dataset avg: {resultado['predicted_ccs_raw']})"
        # predicted_ccs queda tal cual, sin reemplazar

    return resultado


def cargar_modelo():

    os.environ['CUDA_VISIBLE_DEVICES'] = ''
    model_path = r"D:\Modelos TFG\DeepSeek-R1-Distill-Qwen-1.5B"  # Ruta local
    print(f" - Ruta: {model_path}")

    tokenizer = AutoTokenizer.from_pretrained(model_path, trust_remote_code=True)
    if tokenizer.pad_token is None:
        tokenizer.pad_token = tokenizer.eos_token

    model = AutoModelForCausalLM.from_pretrained(
        model_path,
        device_map = "cpu",
        trust_remote_code = True,
        dtype = torch.float16,
        low_cpu_mem_usage = True
    )

    model.eval()  # Modo inferencia: desactiva dropout y gradientes

    if hasattr(torch, 'compile'):
        print(" - Compilando modelo con torch.compile...")
        model = torch.compile(model, mode="reduce-overhead")
        print(" - Compilación completada")

    return model, tokenizer


def inicializar_app():
    global MODEL, TOKENIZER, DATOS_TRAIN, STATS

    print("=" * 70)
    print("Inicializando aplicación de predicción CCS.")
    print("=" * 70)

    # Cargar datos
    csv_path = r"../data/processed/other/train.csv"
    if not os.path.exists(csv_path):
        print(f"AVISO: Archivo {csv_path} no encontrado")
        print("Por favor, asegúrate de que train.csv está en la ruta correcta")
        return False

    print(f"Cargando dataset: {csv_path}")
    DATOS_TRAIN = leer_csv(csv_path)
    print(f" - Dataset cargado: {len(DATOS_TRAIN)} compuestos")

    # Analizar estadísticas
    print("Analizando estadísticas del dataset...")
    STATS = analizar_datos(DATOS_TRAIN)
    print(f" - CCS range: {STATS['ccs_min']:.1f} - {STATS['ccs_max']:.1f} Å²")

    # Cargar modelo
    print("Cargando modelo DeepSeek...")
    MODEL, TOKENIZER = cargar_modelo()
    print(" - Modelo cargado y listo")
    print(" - Aplicación lista para recibir peticiones")
    print("=" * 70)
    return True



from openpyxl import Workbook, load_workbook

def inicializar_excel(filepath, variantes):
    """Crea el Excel con cabeceras si no existe."""
    if os.path.exists(filepath):
        return
    wb = Workbook()
    ws = wb.active
    ws.title = "Predicciones CCS"
    cabeceras = ["", "SMILES", "Aducto"]
    cabeceras.extend(variantes)
    ws.append(cabeceras)
    wb.save(filepath)
    print(f" - Excel creado: {filepath}")

def guardar_fila_excel(filepath, fila):
    """Añade una fila al Excel y guarda inmediatamente."""
    wb = load_workbook(filepath)
    ws = wb.active
    ws.append(fila)
    wb.save(filepath)

def actualizar_celda_excel(filepath, fila_idx, col_idx, valor):
    """Actualiza una celda concreta y guarda inmediatamente."""
    wb = load_workbook(filepath)
    ws = wb.active
    ws.cell(row=fila_idx, column=col_idx, value=valor)
    wb.save(filepath)


def test_prompts():

    global MODEL, TOKENIZER, DATOS_TRAIN, STATS

    # 210 test cases extracted from train.csv
    # Total: 210 compounds ([M+H]: 3+67=70, [M-H]: 4+66=70, [M+Na]: 3+67=70)
    test_cases = [
        # Original test cases (+10)
        {"smiles": "O=C([C@@H](NS(=O)(=O)c1ccc(cc1)Cl)Cc1c[nH]c2c1cccc2)NC1CCCC1", "adduct": "[M+H]+"},
        {"smiles": "COc1cccc(c1)[C@@H]1N(Cc2ccc(cc2)F)C(=O)c2c([C@@H]1C(=O)O)cccc2", "adduct": "[M+H]+"},
        {"smiles": "OC(=O)/C=C/c1ccc(cc1)OC(F)(F)F", "adduct": "[M-H]-"},
        {"smiles": "O=C1N[C@@H]2[C@H](N1)[C@@H](SC2)CCCCC(=O)N1CCC(CC1)C(=O)Nc1ccc2c(c1)OCO2", "adduct": "[M-H]-"},
        {"smiles": "Clc1ccc(cc1)c1occ(n1)CSc1nnnn1CCc1cccs1", "adduct": "[M+Na]+"},
        {"smiles": "N#CC1(CCCC1)NC(=O)CSc1ccc(cn1)S(=O)(=O)N1CCCCC1", "adduct": "[M+Na]+"},
        {"smiles": "Oc1cccc(c1)I", "adduct": "[M-H]-"},
        {"smiles": "CCOc1cc2CC(Oc2cc1NC(=O)CN1C(=O)CCOc2c1cccc2)C", "adduct": "[M+H]+"},
        {"smiles": "Cc1ccc(cc1)n1nnnc1SCC(=O)N1CCCC1", "adduct": "[M+Na]+"},
        {"smiles": "COc1ccc2c(c1)CCCC12NC(=O)N(C1=O)CN1CCN(CC1)c1ccccc1", "adduct": "[M-H]-"},

        # New test cases (+200)
        {"smiles": "CCn1c(N)c(Br)c(=O)[nH]c1=O", "adduct": "[M+H]+"},
        {"smiles": "CC(=O)c1sc(nc1C)NCc1ccccc1", "adduct": "[M+H]+"},
        {"smiles": "Cc1ccc(cc1)n1sc(=N)nc1c1ccccc1", "adduct": "[M+H]+"},
        {"smiles": "C=CCNCCCOc1c(C)cc(cc1Cl)Cl", "adduct": "[M+H]+"},
        {"smiles": "O=C(NC1(CCCC1)C(=O)O)Cc1c[nH]c2c1cccc2", "adduct": "[M+H]+"},
        {"smiles": "ON=Cc1cn(nc1c1ccncc1)c1ccccc1", "adduct": "[M+H]+"},
        {"smiles": "Cc1sc2c(c1C)c(ncn2)Sc1nn2c(n1)nccc2", "adduct": "[M+H]+"},
        {"smiles": "Cc1ccc2c(c1)oc(=O)cc2COC(=O)c1cccs1", "adduct": "[M+H]+"},
        {"smiles": "O=C1NC(C(=O)N1CC(=O)N1CCCC1)Cc1c[nH]c2c1cccc2", "adduct": "[M+H]+"},
        {"smiles": "NC(=O)C1Cc2ccccc2CN1S(=O)(=O)c1c(C)noc1C", "adduct": "[M+H]+"},
        {"smiles": "CCN1CCN(CC1)C(=O)CCC1CCCCC1", "adduct": "[M+H]+"},
        {"smiles": "CCOc1ccccc1NC(=O)CNc1ccccc1C", "adduct": "[M+H]+"},
        {"smiles": "CC1CCN(CC1)CN1C(=O)NC2(C1=O)CCCc1c2cccc1", "adduct": "[M+H]+"},
        {"smiles": "O=C(C1CCN(CC1)c1nc2c(o1)cccc2)NCCc1ccccn1", "adduct": "[M+H]+"},
        {"smiles": "CN(C(=O)C(c1ccccc1)NCCc1c[nH]c2c1cccc2)C", "adduct": "[M+H]+"},
        {"smiles": "CCN(S(=O)(=O)c1ccccc1C#N)Cc1ccc2c(c1)OCO2", "adduct": "[M+H]+"},
        {"smiles": "O=C(N1CCc2c1cccc2)Cn1c(=O)cnc2c1cccc2", "adduct": "[M+H]+"},
        {"smiles": "CC1CC1C(=O)NCC(=O)NCc1ccc(cc1)C(F)(F)F", "adduct": "[M+H]+"},
        {"smiles": "Cc1ccc(cc1)C1(C)NC(=O)N(C1=O)Cc1ccc2c(c1)OCO2", "adduct": "[M+H]+"},
        {"smiles": "Cc1ccc(c(c1)C)C(=O)COc1ccc(cc1)c1nnco1", "adduct": "[M+H]+"},
        {"smiles": "CC(CC(C(=O)O)NC(=O)[C@H](C(C)C)NC(=O)c1ccco1)C", "adduct": "[M+H]+"},
        {"smiles": "CC(=O)N1N=C(CC1c1ccccc1Cl)c1ccccc1NS(=O)(=O)C", "adduct": "[M+H]+"},
        {"smiles": "O=C(N(CC1COc2c(O1)cccc2)C)Cc1ccc2c(c1)CCC2", "adduct": "[M+H]+"},
        {"smiles": "CC(CNC(=O)NC(=O)CN1C(=O)NC2(C1=O)CCC(CC2)C)C", "adduct": "[M+H]+"},
        {"smiles": "CCCN(C(=O)c1ccc(cc1)n1cccn1)Cc1nnc(o1)c1ccco1", "adduct": "[M+H]+"},
        {"smiles": "CCc1nc2c(n1CC(=O)Nc1ccccc1c1ccccc1)cccc2", "adduct": "[M+H]+"},
        {"smiles": "Fc1ccc2c(c1)c(CCC(=O)N1CCOCC1)c([nH]2)c1ccc(cc1)F", "adduct": "[M+H]+"},
        {"smiles": "CCc1ccccc1NC(=O)c1ccccc1c1ccc(cc1)Cn1cccn1", "adduct": "[M+H]+"},
        {"smiles": "Brc1ccc(cc1)S(=O)(=O)NCCC(=O)N(Cc1cccs1)C", "adduct": "[M+H]+"},
        {"smiles": "Clc1ccc(cc1)c1cnc(o1)CN1CCN(CC1)c1ccccn1", "adduct": "[M+H]+"},
        {"smiles": "CC(CC(C(=O)NCCNC(=O)C)NC(=O)c1ccc(cc1)Cl)C", "adduct": "[M+H]+"},
        {"smiles": "CCOC(=O)c1ccc(cc1)OCC(CN1CC(C)OC(C1)C)O", "adduct": "[M+H]+"},
        {"smiles": "O=C(Nc1c(C)cccc1C)CN1CCN(CC1)C(=O)Cc1ccccc1", "adduct": "[M+H]+"},
        {"smiles": "Cn1ccnc1Sc1ccc(cc1)NS(=O)(=O)c1ccccc1C(F)(F)F", "adduct": "[M+H]+"},
        {"smiles": "O=C(c1cncn1c1ccccc1)N1CCCC(C1)c1nc2c(s1)cccc2", "adduct": "[M+H]+"},
        {"smiles": "CCCn1c(NC(=O)c2ccccc2SCc2c(C)onc2C)nc2c1cccc2", "adduct": "[M+H]+"},
        {"smiles": "COc1cccc(c1)C(N1CCCC1)CNC(=O)c1ccc(cc1)Br", "adduct": "[M+H]+"},
        {"smiles": "CCn1c(SCC(=O)N2CCN(CC2)c2ncccn2)nnc1C1CC1", "adduct": "[M+H]+"},
        {"smiles": "Clc1ccccc1c1nc2ccccc2c(c1)C(=O)OCc1c(C)noc1C", "adduct": "[M+H]+"},
        {"smiles": "O=C(CSc1nnc(n1Cc1ccccc1)C1CC1)NCC1CCCO1", "adduct": "[M+H]+"},
        {"smiles": "COc1cc(OC)ccc1C1CCCN1CC(=O)Nc1ccnn1C1CCCC1", "adduct": "[M+H]+"},
        {"smiles": "CC1CN(CCCNC(=O)COc2ccccc2c2ccccc2)CC(C1)C", "adduct": "[M+H]+"},
        {"smiles": "O=C(c1cccnc1Sc1ccc2c(c1)OCCO2)N1CCN(CC1)c1ccccn1", "adduct": "[M+H]+"},
        {"smiles": "O=C(N1CCC(CC1)c1nc2c(o1)cccc2)C=Cc1ccc2c(c1)OCO2", "adduct": "[M+H]+"},
        {"smiles": "Cc1cccc(c1)n1c(SCc2nnc(o2)c2ccco2)nnc1N1CCOCC1", "adduct": "[M+H]+"},
        {"smiles": "COc1ccccc1c1nnc(n1CC)SCC(=O)N1CC(C)CC(C1)C", "adduct": "[M+H]+"},
        {"smiles": "COc1ccccc1c1nn(cc1C(=O)OC1CC(OC1=O)C)Cc1ccccc1", "adduct": "[M+H]+"},
        {"smiles": "COc1cccc(c1C(=O)N(c1ccccc1)C1CCN(CC1)Cc1ccccc1)OC", "adduct": "[M+H]+"},
        {"smiles": "CCOc1ccc(cc1)c1n[nH]c(=S)n1CC(=O)NCC1(CC1)c1ccccc1", "adduct": "[M+H]+"},
        {"smiles": "O=C(CSc1nnc(n1Cc1ccco1)c1cccc(c1)C)NCC1COc2c(O1)cccc2", "adduct": "[M+H]+"},
        {"smiles": "O=C(N(c1sc2c(n1)CCCC2)Cc1ccccc1)Cn1nnc(n1)c1ccccc1", "adduct": "[M+H]+"},
        {"smiles": "O=C(N1CCC(CC1)c1ncc([nH]1)c1ccc(c(c1)F)F)NCC(c1ccco1)N1CCCC1", "adduct": "[M+H]+"},
        {"smiles": "O=C(Nc1ccc(cc1)N1CCOCC1)CN1CCCC1c1ccc(cc1)Cl", "adduct": "[M+H]+"},
        {"smiles": "COc1ccc(c(c1)C1CCCN1C(=O)CNC(c1cccs1)c1ccc(cc1)F)OC", "adduct": "[M+H]+"},
        {"smiles": "O=C(C1CC(=O)N(C1)Cc1ccco1)NCc1ccc(cc1)Oc1ccc(cc1)F", "adduct": "[M+H]+"},
        {"smiles": "Cc1ccc(cc1C(=O)N1CCCC(C1)c1nc2c(s1)cccc2)S(=O)(=O)N1CCCC1", "adduct": "[M+H]+"},
        {"smiles": "O=C(Nc1ccccc1C(=O)Nc1ccccc1)CNC(c1cccs1)c1ccccc1", "adduct": "[M+H]+"},
        {"smiles": "O=C1CCc2c(N1)ccc(c2)S(=O)(=O)NCC(c1ccccc1)(c1ccccc1)C", "adduct": "[M+H]+"},
        {"smiles": "COc1ccc(cc1C1CCCN1C(=O)CN(CC(=O)Nc1c(C)cc(cc1C)C)C)OC", "adduct": "[M+H]+"},
        {"smiles": "CC1CCCCN1S(=O)(=O)c1ccc(cc1)C(=O)OCc1noc(c1)c1ccco1", "adduct": "[M+H]+"},
        {"smiles": "COc1cc(cc(c1OCc1ccccc1)OC)C(=O)N1CCN(CC1c1ccccc1)C", "adduct": "[M+H]+"},
        {"smiles": "Fc1ccccc1Cn1c(CN(Cc2ccccc2)Cc2ccccc2)nc2c1c(=O)n(c(=O)n2C)C", "adduct": "[M+H]+"},
        {"smiles": "COc1cc(ccc1OC)c1nnn(n1)CC(=O)NC(c1ccccc1)c1ccccc1", "adduct": "[M+H]+"},
        {"smiles": "COCCn1c(=O)[nH]c(=O)c(c1N)N(C(=O)c1cn(nc1c1ccccc1)c1ccccc1)Cc1ccccc1", "adduct": "[M+H]+"},
        {"smiles": "O=C(Cn1nnc2c(c1=O)cccc2)Nc1ccc2c(c1)OCO2", "adduct": "[M+H]+"},
        {"smiles": "CC(NC(=O)Nc1ccccc1C(=O)OCC(=O)N1CCCCC1)C", "adduct": "[M+H]+"},
        {"smiles": "CCOc1cc2CC(Oc2cc1NC(=O)CN1C(=O)CCOc2c1cccc2)C", "adduct": "[M+H]+"},
        {"smiles": "Oc1cccc(c1)I", "adduct": "[M-H]-"},
        {"smiles": "CCc1cccc(c1)NC(=O)c1cscc1", "adduct": "[M-H]-"},
        {"smiles": "CCCNS(=O)(=O)c1ccc(cc1)C(C)C", "adduct": "[M-H]-"},
        {"smiles": "O=C(Nc1ccccc1C(C)C)CN1C(=O)CCC1=O", "adduct": "[M-H]-"},
        {"smiles": "COC1=CC(C=CC(O)=O)=CC(OC)=C1OC", "adduct": "[M-H]-"},
        {"smiles": "CC(N1CCOCC1)CNS(=O)(=O)c1ccccc1", "adduct": "[M-H]-"},
        {"smiles": "CCOc1ccc(cc1OC)C(=O)NCc1ccncc1", "adduct": "[M-H]-"},
        {"smiles": "O=C(c1cccc(c1O)C)NCc1ccc2c(c1)OCO2", "adduct": "[M-H]-"},
        {"smiles": "CCCCOC(=O)CN1C(=O)NC(C1=O)(CC)c1ccccc1", "adduct": "[M-H]-"},
        {"smiles": "O=C(c1c(C)noc1C)OC(c1nc2ccccc2c(=O)[nH]1)C", "adduct": "[M-H]-"},
        {"smiles": "O=C(C(N1CCCC(C1)C(=O)N)C)Nc1ccccc1C(F)(F)F", "adduct": "[M-H]-"},
        {"smiles": "Cc1cccc(c1)CN1C(=O)NC2(C1=O)CCc1c2cccc1", "adduct": "[M-H]-"},
        {"smiles": "O=C(Cn1cc(ccc1=O)C(F)(F)F)Nc1cnc2c(c1)cccc2", "adduct": "[M-H]-"},
        {"smiles": "Fc1ccc2c(c1)c(CCC(=O)N1CCOCC1)c([nH]2)c1ccc(cc1)F", "adduct": "[M-H]-"},
        {"smiles": "CC(N(C(=O)CN1C(=O)NC2(C1=O)CCCCCC2)C(C)C)C", "adduct": "[M-H]-"},
        {"smiles": "CC(=O)c1ccc(cc1)NC(=O)c1ccccc1c1ccccc1", "adduct": "[M-H]-"},
        {"smiles": "O=C(Nc1ccc(cc1)C)CN(S(=O)(=O)C)C1CCCCC1", "adduct": "[M-H]-"},
        {"smiles": "COc1ccc(cc1F)CSc1nnc([nH]1)c1ccccc1Br", "adduct": "[M-H]-"},
        {"smiles": "COc1ccccc1CC(=O)Nc1cccc(c1)S(=O)(=O)N(C)C", "adduct": "[M-H]-"},
        {"smiles": "O=C(C(c1ccccc1)Sc1nnnn1C1CCCC1)N1CCOCC1", "adduct": "[M-H]-"},
        {"smiles": "CN(C(=O)c1ccc(cc1)S(=O)(=O)N(C)C)Cc1ccccc1F", "adduct": "[M-H]-"},
        {"smiles": "CC(c1ccccc1)NC(=O)c1cc(nc2c1cccc2)C1CC1", "adduct": "[M-H]-"},
        {"smiles": "O=C(Nc1ccccc1Cc1ccccc1)CSc1nnnn1C1CC1", "adduct": "[M-H]-"},
        {"smiles": "N#CC1(CCCC1)NC(=O)CSc1nnc(s1)NCC1CCCO1", "adduct": "[M-H]-"},
        {"smiles": "CNc1snc(c1C(=O)N1CCN(CC1)S(=O)(=O)c1ccccc1F)C", "adduct": "[M-H]-"},
        {"smiles": "O=C(NC1CCCC(C1C)C)CSc1nnc2n1cccc2", "adduct": "[M-H]-"},
        {"smiles": "CCNS(=O)(=O)c1ccc(cc1)C(=O)NC1CCC(CC1)O", "adduct": "[M-H]-"},
        {"smiles": "O=C(c1cc(ccc1C)S(=O)(=O)N)NCCc1ccccc1C", "adduct": "[M-H]-"},
        {"smiles": "CC(=O)NC(C(=O)NCCc1ccccn1)Cc1c[nH]c2c1cccc2", "adduct": "[M-H]-"},
        {"smiles": "O=C(CN1CSCC1=O)Nc1ccc(cc1)SCc1cccnc1", "adduct": "[M-H]-"},
        {"smiles": "CCOC(=O)C(=Cc1ccc(c(c1)OC)OC)C1=Nn2c(SC1)nnc2C", "adduct": "[M-H]-"},
        {"smiles": "OCCN(S(=O)(=O)c1ccc(cc1)C(=O)NCc1ccccc1F)C", "adduct": "[M-H]-"},
        {"smiles": "COc1cc(ccc1OC(F)F)C(=O)OCN1C(=O)c2c(C1=O)cccc2", "adduct": "[M-H]-"},
        {"smiles": "O=C(c1cc(C)c(cc1C)C)CSc1nnc(n1C)c1cccs1", "adduct": "[M-H]-"},
        {"smiles": "COCCN1CCN(CC1)S(=O)(=O)c1ccc(cc1)C(=O)OCC", "adduct": "[M-H]-"},
        {"smiles": "COc1ccc(cc1)n1c(SCC(=O)Nc2ccccc2C)nc2c1cccc2", "adduct": "[M-H]-"},
        {"smiles": "Fc1ccc(cc1)CNC(=O)C1(CCN(CC1)C(=O)C)c1ccccc1", "adduct": "[M-H]-"},
        {"smiles": "COc1cc2c(cc1NC(=O)CCSc1ccccc1)oc1c2cccc1", "adduct": "[M-H]-"},
        {"smiles": "Cc1ccc(c(c1)S(=O)(=O)N1CCN(CC1)C(=O)Cc1csc(n1)Cc1ccccn1)C", "adduct": "[M-H]-"},
        {"smiles": "CC(C(C(=O)NCc1ccc(cc1)Cl)NC(=O)c1ccccc1Br)C", "adduct": "[M-H]-"},
        {"smiles": "CCC(=O)N(c1nnc(s1)SCc1nnc(o1)c1cc(oc1C)C)C1CC1", "adduct": "[M-H]-"},
        {"smiles": "O=C(Nc1cc(Cl)ccc1n1ncnc1)CN1C(=O)NC(C1=O)(C)c1ccc2c(c1)OCO2", "adduct": "[M-H]-"},
        {"smiles": "O=C(C(OC(=O)c1ccc(cc1)c1ccc(cc1)O)C)NCc1ccccc1", "adduct": "[M-H]-"},
        {"smiles": "O=C1NC(C(=O)N1Cc1onc(n1)c1cccs1)(C)c1ccc2c(c1)OCCO2", "adduct": "[M-H]-"},
        {"smiles": "CCCN(CC1=C(C(=O)OCC)C(NC(=O)N1)C)CC(=O)Nc1ccc(c(c1F)F)F", "adduct": "[M-H]-"},
        {"smiles": "NC(=O)C1Cc2ccccc2CN1C(=O)CCc1c([nH]c2c1cccc2)c1ccc(cc1)F", "adduct": "[M-H]-"},
        {"smiles": "O=C(C1CCN(CC1)c1nc2c(o1)cccc2)NCCc1ccccn1", "adduct": "[M-H]-"},
        {"smiles": "CC(Cn1c(SCc2nnc(o2)c2ccccc2Br)nnc1N1CCOCC1)C", "adduct": "[M-H]-"},
        {"smiles": "O=C(Nc1cc(ccc1Oc1ccc(cc1)F)S(=O)(=O)N1CCOCC1)Cc1cccs1", "adduct": "[M-H]-"},
        {"smiles": "COC(=O)c1cc(ccc1NS(=O)(=O)C1=Cc2c(CC1)cccc2)N1CCOCC1", "adduct": "[M-H]-"},
        {"smiles": "COc1ccc(cc1)N(S(=O)(=O)c1ccc2c(c1)OCCO2)Cc1noc(c1)c1ccco1", "adduct": "[M-H]-"},
        {"smiles": "CC(=O)Nc1ccc(c(c1)F)C1=NN=C(SC1)NC(C)(C)C", "adduct": "[M-H]-"},
        {"smiles": "CC(C(=O)N1CCc2c(C1)cccc2)Sc1nnc(n1c1ccccc1F)c1cccnc1", "adduct": "[M-H]-"},
        {"smiles": "CC(C12CC3CC(C2)CC(C1)C3)NC(=O)c1ccc(cc1)S(=O)(=O)N1CCCC1", "adduct": "[M-H]-"},
        {"smiles": "O=C(C1CC(=O)N(C1)Cc1ccco1)NCc1ccc(cc1)Cn1cncc1", "adduct": "[M-H]-"},
        {"smiles": "O=C(Nc1ccc(cc1)N1CCOCC1)COc1ccccc1C(=O)Nc1ccccc1", "adduct": "[M-H]-"},
        {"smiles": "Fc1cccc(c1)n1c(SCc2nnc(o2)c2ccccc2)nnc1N1CCOCC1", "adduct": "[M-H]-"},
        {"smiles": "COc1cc(ccc1OC)Cc1nnc(n1N)SCC(=O)Nc1ccc(cc1)c1ccccc1", "adduct": "[M-H]-"},
        {"smiles": "C=CCc1ccc(c(c1)OC)OCC(COc1cccc(c1)C(=O)NCc1ccccc1)O", "adduct": "[M-H]-"},
        {"smiles": "COc1ccc(c(c1)OC)NC(=O)CSc1oc(nc1S(=O)(=O)c1ccc(cc1)C)c1ccco1", "adduct": "[M-H]-"},
        {"smiles": "O=C(NC(c1cccc2c1cccc2)C)CN1C(=O)NC(C1=O)(Cc1ccccc1)Cc1ccccc1", "adduct": "[M-H]-"},
        {"smiles": "COc1ccc(cc1)C(c1ccccc1)NC(=O)Cn1cc(ccc1=O)S(=O)(=O)N1CCCC1", "adduct": "[M-H]-"},
        {"smiles": "COc1ccc(cc1)c1scc(n1)CN1CCN(CC1)C(=O)c1ccc2c(c1)OCO2", "adduct": "[M-H]-"},
        {"smiles": "O=C(c1cn(nc1c1ccccc1)c1ccccc1)NCC(c1cccs1)N1CCOCC1", "adduct": "[M-H]-"},
        {"smiles": "O=C(Cn1cnc2c(c1=O)c(cs2)c1ccc(cc1)C)NC1CCN(C1)Cc1ccccc1", "adduct": "[M-H]-"},
        {"smiles": "COc1ccc2c(c1)CCCC12NC(=O)N(C1=O)CN1CCN(CC1)c1ccccc1", "adduct": "[M-H]-"},
        {"smiles": "Clc1ccc(cc1)c1nnc(o1)C(=O)c1ccncc1", "adduct": "[M+Na]+"},
        {"smiles": "COc1ccc(cc1)OCCOc1ccc(cc1)NC(=O)C", "adduct": "[M+Na]+"},
        {"smiles": "O=C(N1CCCSc2c1cccc2)Cc1cccs1", "adduct": "[M+Na]+"},
        {"smiles": "C=CCN1CC2=C(C1=O)C(NC(=O)N2)c1ccc(cc1)O", "adduct": "[M+Na]+"},
        {"smiles": "CC(=O)N1CCN(CC1)c1ccc(c(c1)C(F)(F)F)[N+](=O)[O-]", "adduct": "[M+Na]+"},
        {"smiles": "O=C(c1cc(n(c1C)C)C)COC(=O)c1c(F)cccc1F", "adduct": "[M+Na]+"},
        {"smiles": "COC(=O)c1cccc(c1)OCC(=O)Nc1ccc2c(c1)OCCO2", "adduct": "[M+Na]+"},
        {"smiles": "Brc1ccc(cc1)S(=O)(=O)N1CCCC1c1ccccc1", "adduct": "[M+Na]+"},
        {"smiles": "NC(=O)C1CCCN1c1nc(nc2c1cccc2)c1cccnc1", "adduct": "[M+Na]+"},
        {"smiles": "CNC(=O)c1cccc(c1)NC(=O)c1ccc2c(c1)cccc2", "adduct": "[M+Na]+"},
        {"smiles": "Cc1ccc(c(c1)NC(=O)c1nc2n(n1)c(C)cc(n2)C)F", "adduct": "[M+Na]+"},
        {"smiles": "O=C(C1COc2c(O1)cccc2)OC(C(=O)N1CCc2c1cccc2)C", "adduct": "[M+Na]+"},
        {"smiles": "Brc1ccc(=O)n(c1)CC(=O)N(C(C)(C)C)Cc1ccccc1", "adduct": "[M+Na]+"},
        {"smiles": "COc1cccc(c1)CNC(=O)COC(=O)c1cc(Cl)c(cc1OC)N", "adduct": "[M+Na]+"},
        {"smiles": "CC(c1nnc(o1)c1cccs1)Sc1nnc(o1)c1ccco1", "adduct": "[M+Na]+"},
        {"smiles": "O=C(Cn1ccccc1=O)Nc1ccc2c(c1)OC1(O2)CCCC1", "adduct": "[M+Na]+"},
        {"smiles": "O=C(c1ccc(cc1)NC(=O)C)OCC(=O)NC1CCCCCC1", "adduct": "[M+Na]+"},
        {"smiles": "OCCNC(=O)c1ccc(cc1)S(=O)(=O)N(CC)CC", "adduct": "[M+Na]+"},
        {"smiles": "COc1ccc(cc1)OCCCn1cc(ccc1=O)S(=O)(=O)N1CCCC1", "adduct": "[M+Na]+"},
        {"smiles": "O=C(CNC(=O)C1CCCCC1)NCCSCc1ccc(s1)Br", "adduct": "[M+Na]+"},
        {"smiles": "CCc1ccccc1NC(=O)c1cc(ccc1Br)S(=O)(=O)N", "adduct": "[M+Na]+"},
        {"smiles": "O=C(CN1C(=O)NC(C1=O)(C)c1ccccc1Br)NC1CCCCC1", "adduct": "[M+Na]+"},
        {"smiles": "O=C(N(C1=CCCCC1)Cc1ccco1)CSc1nnnn1C1CCCC1", "adduct": "[M+Na]+"},
        {"smiles": "Clc1ccccc1c1nnc(o1)CN(C(=O)c1sccc1C)C1CC1", "adduct": "[M+Na]+"},
        {"smiles": "O=C(CN1C(=O)NC2(C1=O)CCc1c2cccc1)NC1CCCc2c1cccc2", "adduct": "[M+Na]+"},
        {"smiles": "O=C(C(Sc1nc(cn1N)c1ccccc1)C)Nc1ccccc1c1ccccc1", "adduct": "[M+Na]+"},
        {"smiles": "O=C(C1CC2CCCC(C1)C12SCCS1)NCCc1cccs1", "adduct": "[M+Na]+"},
        {"smiles": "O=C(Cn1c(=O)c(F)cn(c1=O)C1CCCO1)Nc1ccccc1Oc1ccccc1", "adduct": "[M+Na]+"},
        {"smiles": "CCN(C(=O)CN1C(=O)NC(C1=O)(C)c1ccc(cc1)C#N)CC(=C)C", "adduct": "[M+Na]+"},
        {"smiles": "CN(C(=O)c1ccc(cc1)c1ccccc1)CC1COc2c(O1)cccc2", "adduct": "[M+Na]+"},
        {"smiles": "O=C(CSc1nnc(n1c1ccccc1)c1ccc(cc1)F)NCC1CCCO1", "adduct": "[M+Na]+"},
        {"smiles": "CCOc1cccc2c1oc(c2)C(NS(=O)(=O)c1ccc2c(c1)OCCO2)C", "adduct": "[M+Na]+"},
        {"smiles": "O=C(C(c1ccccc1)c1ccccc1)NCc1ccccc1CN1CCOCC1", "adduct": "[M+Na]+"},
        {"smiles": "CC(C12CC3CC(C2)CC(C1)C3)NC(=O)Cc1ccc(cc1)NC(=O)C", "adduct": "[M+Na]+"},
        {"smiles": "O=C(NC12CC3CC(C2)CC(C1)C3)CSc1nc2c(s1)ccc(c2)Cl", "adduct": "[M+Na]+"},
        {"smiles": "O=C(N(C(c1ccccc1)C)C)CSc1nnc(n1CC1CCCO1)c1cccs1", "adduct": "[M+Na]+"},
        {"smiles": "CCCN(Cc1nnnn1c1ccccc1)Cc1nnc(o1)c1cccs1", "adduct": "[M+Na]+"},
        {"smiles": "CCCC1(NC(=O)N(C1=O)CC(=O)NC(c1ccc2c(c1)OCCO2)C)c1ccccc1", "adduct": "[M+Na]+"},
        {"smiles": "O=C(c1ccc(cc1)S(=O)(=O)N(C)C)NCCCn1ccc2c1cccc2", "adduct": "[M+Na]+"},
        {"smiles": "C=CCn1c(SC(C(=O)Nc2ccc3c(c2)OCO3)C)nc2c(c1=O)c(C)c(s2)C", "adduct": "[M+Na]+"},
        {"smiles": "O=C(Nc1ccc2c(c1)CCN2C(=O)C)COc1ccccc1c1ccccc1", "adduct": "[M+Na]+"},
        {"smiles": "O=C(C(Sc1nnc(n1c1ccccc1)c1ccco1)C)N1CCc2c(C1)cccc2", "adduct": "[M+Na]+"},
        {"smiles": "O=C(N(C1CCCc2c1cccc2)C)CSc1nnc(n1Cc1ccccc1)C1CC1", "adduct": "[M+Na]+"},
        {"smiles": "O=C(CN1CC(Oc2c1cccc2)C(=O)N1CCOCC1)NC(=O)NC1CCCC1", "adduct": "[M+Na]+"},
        {"smiles": "O=S1(=O)CCC(C1)Nc1nc(nc2c1c(cs2)c1ccccc1)c1cccnc1", "adduct": "[M+Na]+"},
        {"smiles": "O=C(C1CCCN(C1)C(=O)c1ccc(cc1)Cl)NCC(N1CCCCC1)(C)C", "adduct": "[M+Na]+"},
        {"smiles": "COc1ccc(cc1)NC(=O)CC1C(=O)Nc2c(N1CC(=O)N1CCCC1)cccc2", "adduct": "[M+Na]+"},
        {"smiles": "CCOc1ccc(cc1)C(=O)NC(C(=O)Nc1ccc(cc1)Cl)C(C)C", "adduct": "[M+Na]+"},
        {"smiles": "O=C(Nc1ccc(cc1)C)Cc1scc(n1)Cn1nc(oc1=O)c1ccc(cc1)F", "adduct": "[M+Na]+"},
        {"smiles": "O=S(=O)(c1cccc(c1)c1nnc(o1)SCCCOc1ccccc1)N1CCOCC1", "adduct": "[M+Na]+"},
        {"smiles": "COc1ccc(cc1)C1=NN(C(C1)c1ccco1)C(=O)Cn1nnc(n1)c1cccs1", "adduct": "[M+Na]+"},
        {"smiles": "O=C(N1CCCC1)CSc1ccccc1C(=O)N1CCN(CC1)c1ccccc1Cl", "adduct": "[M+Na]+"},
        {"smiles": "O=S(=O)(N(C1CCCC1)Cc1ccc2c(c1)OCO2)c1ccc(cc1)S(=O)(=O)N1CCCC1", "adduct": "[M+Na]+"},
        {"smiles": "Fc1cccc(c1)n1c(SCC(=O)N2CCC3C(C2)CCCC3)nnc1N1CCOCC1", "adduct": "[M+Na]+"},
        {"smiles": "O=C(N1CCC(CC1)Cc1ccccc1)CSc1nnc(o1)c1ccc2c(c1)OCO2", "adduct": "[M+Na]+"},
        {"smiles": "c1ccc(nc1)c1scc(n1)c1ccc2c(c1)OCO2", "adduct": "[M+Na]+"},
        {"smiles": "COc1cccc(c1)CNS(=O)(=O)c1cccs1", "adduct": "[M+Na]+"},
        {"smiles": "CCC(NS(=O)(=O)c1cc(F)cc(c1)F)C", "adduct": "[M+Na]+"},
        {"smiles": "CC(=O)N1CCN(CC1)c1cc2c(cc1F)c(=O)c(cn2C1CC1)C(=O)OCc1cccc(c1F)F", "adduct": "[M+Na]+"},
        {"smiles": "COc1cccc(c1)C(=O)Nc1ccccc1OC(F)F", "adduct": "[M+Na]+"},
        {"smiles": "Cc1nnc(o1)COc1ccc(cc1)Oc1ccccc1", "adduct": "[M+Na]+"},
        {"smiles": "COc1ccc(cc1)C(=O)COC(=O)c1cc2c([nH]1)cccc2", "adduct": "[M+Na]+"},
        {"smiles": "O=C(N(C1CC1)Cc1nnc(o1)c1ccccc1Br)c1ccccc1", "adduct": "[M+Na]+"},
        {"smiles": "CCN1CCN(CC1)C(=O)CCC1CCCCC1", "adduct": "[M+Na]+"},
        {"smiles": "O=C(N(Cc1ccccc1N1CCOCC1)C)C1CCN(CC1)C(=O)c1cccs1", "adduct": "[M+Na]+"},
        {"smiles": "O=C(N1CCOCC1)C1CCN(CC1)C(=O)c1ccc(o1)c1nc2c(s1)cccc2", "adduct": "[M+Na]+"},
        {"smiles": "N#Cc1cccc(c1)C(=O)N1CCN(CC1)Cc1ccccc1", "adduct": "[M+Na]+"},

        # Additional new tesst cases - 420 compounds (140 per adduct)
        # [M+H]+ cases (+140)
        {"smiles": "Clc1ccc(nn1)N1CCCC1", "adduct": "[M+H]+"},
        {"smiles": "N#CC(=Cc1cscc1)c1ccccn1", "adduct": "[M+H]+"},
        {"smiles": "Cc1ccc(cc1)Nc1ncnc2c1nc[nH]2", "adduct": "[M+H]+"},
        {"smiles": "O=C(C(=O)N1CCNCC1)N1CCOCC1", "adduct": "[M+H]+"},
        {"smiles": "O=C(N(C)C)COC(=O)c1ccccc1I", "adduct": "[M+H]+"},
        {"smiles": "Clc1cccc(c1)c1nc(c(s1)C(=O)O)C", "adduct": "[M+H]+"},
        {"smiles": "O=C(Nc1ccccc1)CCN1C(=O)CCC1=O", "adduct": "[M+H]+"},
        {"smiles": "Cc1ccc(c(c1)S(=O)(=O)Nc1ccccn1)C", "adduct": "[M+H]+"},
        {"smiles": "CCNC(=O)CSc1cc(C)c2c(n1)cccc2", "adduct": "[M+H]+"},
        {"smiles": "COC(=O)C1=C(C)NC(=S)NC1c1cscc1", "adduct": "[M+H]+"},
        {"smiles": "CCC(NC(=O)CN1C(=O)C2C(C1=O)CC=CC2)C", "adduct": "[M+H]+"},
        {"smiles": "COc1ccc(c(c1)C(=O)Nc1ncc(s1)C)Br", "adduct": "[M+H]+"},
        {"smiles": "CC(CC(C(=O)N(C)C)NC(=O)c1cnccn1)C", "adduct": "[M+H]+"},
        {"smiles": "COCCCNC(=O)C1OC(=O)c2c(C1)cccc2", "adduct": "[M+H]+"},
        {"smiles": "N#CC(Sc1nnc(s1)Nc1ccccc1Cl)C", "adduct": "[M+H]+"},
        {"smiles": "Cc1ccc(cc1)Oc1ccccc1NC(=O)C1CC1", "adduct": "[M+H]+"},
        {"smiles": "Cc1ccc(cc1F)c1noc(n1)CN1CCOCC1", "adduct": "[M+H]+"},
        {"smiles": "Cc1ccc(nc1)NS(=O)(=O)c1cccc(c1Cl)Cl", "adduct": "[M+H]+"},
        {"smiles": "O=C(Nc1sc2c(n1)CCC2)CSc1ccccc1", "adduct": "[M+H]+"},
        {"smiles": "C=CCc1cc(cc(c1OC)OC)C(=O)NC1CC1", "adduct": "[M+H]+"},
        {"smiles": "N#Cc1ccsc1NC(=O)c1cccnc1N1CCCC1", "adduct": "[M+H]+"},
        {"smiles": "COc1ccc(c(c1)C(=O)OCc1onc(c1)C)O", "adduct": "[M+H]+"},
        {"smiles": "Cc1ccc(cc1)CN1C(=O)NC(C1=O)(C)c1cccs1", "adduct": "[M+H]+"},
        {"smiles": "CCOC(=O)NC(=O)CSc1ncnc2c1cccc2", "adduct": "[M+H]+"},
        {"smiles": "O=C(N1CCCC1)CSc1nn2c(n1)nc(cc2C)C", "adduct": "[M+H]+"},
        {"smiles": "Cc1c(C)sc2c1c(=O)n1c(n2)C(=Cc2ccsc2)CC1", "adduct": "[M+H]+"},
        {"smiles": "Clc1ccc(cc1)OCc1nnc(o1)c1ccccc1", "adduct": "[M+H]+"},
        {"smiles": "CN(C(=O)C1CC2CC1C=C2)Cc1cccc(c1OC)OC", "adduct": "[M+H]+"},
        {"smiles": "O=C(c1ccc(cc1)N(C)C)Nc1nc2c([nH]1)cccc2", "adduct": "[M+H]+"},
        {"smiles": "COCCN1C(=O)NC(C(=C1C)C(=O)OCC)c1cccs1", "adduct": "[M+H]+"},
        {"smiles": "CCn1cc(c2c1cccc2)C=C1C(=O)Nc2c1cccc2", "adduct": "[M+H]+"},
        {"smiles": "O=C1c2ccccc2C(=O)C21Cc1ccccc1N1C2CCCC1", "adduct": "[M+H]+"},
        {"smiles": "Cc1ccc(cc1)S(=O)(=O)Nc1ncnc2c1c1CCCc1s2", "adduct": "[M+H]+"},
        {"smiles": "CCCNC(=O)C(NC(=O)Cc1ccc(cc1)OC)C", "adduct": "[M+H]+"},
        {"smiles": "Cc1cccc(c1)NC(=O)C(N1CCc2c1cccc2)C", "adduct": "[M+H]+"},
        {"smiles": "O=C(Nc1c(oc2c1cccc2)C(=O)N)Cc1ccc(cc1)Cl", "adduct": "[M+H]+"},
        {"smiles": "CC1C(C)CCCC1NC(=O)COC(=O)c1ccccc1", "adduct": "[M+H]+"},
        {"smiles": "CC1Cc2c(N1C(=O)Cn1cnc3c1cccc3)cccc2", "adduct": "[M+H]+"},
        {"smiles": "OC1CCC(CC1)Nc1nc(nc2c1cccc2)C(F)(F)F", "adduct": "[M+H]+"},
        {"smiles": "O=C(c1cc(F)ccc1Br)NC1CC(C)(C)NC(C1)(C)C", "adduct": "[M+H]+"},
        {"smiles": "COc1cc(cc(c1)OC)NC(=O)c1cc(C)ccc1C", "adduct": "[M+H]+"},
        {"smiles": "O=C(Nc1scc(n1)c1ccccn1)COc1noc(c1)C", "adduct": "[M+H]+"},
        {"smiles": "CC1CCc2c(C1)sc(c2C(=O)N)NC(=O)c1cccn1C", "adduct": "[M+H]+"},
        {"smiles": "Clc1cccc(c1)OCc1nnc(o1)c1ccccc1Br", "adduct": "[M+H]+"},
        {"smiles": "O=C(Nc1ncc(c(c1Cl)C)Cl)COC(=O)c1cccn1C", "adduct": "[M+H]+"},
        {"smiles": "Fc1ccc(cc1)CS(=O)(=O)NCCc1nc2n(c1)cccc2", "adduct": "[M+H]+"},
        {"smiles": "Brc1ccc(o1)C(=O)N1N=C(CC1c1ccco1)c1ccco1", "adduct": "[M+H]+"},
        {"smiles": "CCONC(=O)C(C(C)C)NC1=NS(=O)(=O)c2c1cccc2", "adduct": "[M+H]+"},
        {"smiles": "Cc1ccc(cc1)n1nnnc1SC(C(=O)N1CCCC1)C", "adduct": "[M+H]+"},
        {"smiles": "O=C(N1CCCCC1)CN1C(=O)NC2(C1=O)CCc1c2cccc1", "adduct": "[M+H]+"},
        {"smiles": "CNC(=O)c1ccc(cc1)NC(=O)CNc1ccccc1", "adduct": "[M+H]+"},
        {"smiles": "Cc1onc(c1)NC(=O)C(OC(=O)c1ccc2c(c1)OCO2)C", "adduct": "[M+H]+"},
        {"smiles": "O=C(C1CCN(CC1)C(=O)c1ccc(cc1)F)NC1=NCCS1", "adduct": "[M+H]+"},
        {"smiles": "O=C(COc1ccc(cc1)c1nnco1)NCCc1c[nH]c2c1cccc2", "adduct": "[M+H]+"},
        {"smiles": "Clc1ccc(cc1)C1(CC1)C(=O)NCc1ccncc1", "adduct": "[M+H]+"},
        {"smiles": "CCCN1C(=S)NC(=Cc2ccc(c(c2)OCC)O)C1=O", "adduct": "[M+H]+"},
        {"smiles": "Cc1ccc(cc1)C(=O)C(=O)Nc1scc(n1)c1ccccc1", "adduct": "[M+H]+"},
        {"smiles": "O=C1OCCC1Sc1nnc(n1c1ccccc1Cl)N1CCCC1", "adduct": "[M+H]+"},
        {"smiles": "COc1cc(cc(c1)OC)C(c1nccn1C)NC(=O)C1CC1", "adduct": "[M+H]+"},
        {"smiles": "CCC(c1nnc(s1)NC(=O)Cc1ccccc1OC)CC", "adduct": "[M+H]+"},
        {"smiles": "NC(=O)CSc1ccccc1NC(=O)c1cc2c(s1)cccc2", "adduct": "[M+H]+"},
        {"smiles": "OC(=O)c1cn(nc1c1cccc2c1cccc2)c1ccccc1", "adduct": "[M+H]+"},
        {"smiles": "CCOc1ccccc1OCCC(=O)Nc1nc2c(s1)cccc2", "adduct": "[M+H]+"},
        {"smiles": "Cc1oc(nc1CN1CCC(CC1)c1ccn[nH]1)c1ccccc1", "adduct": "[M+H]+"},
        {"smiles": "CC1CCCN(C1)C(=O)CSc1nnc(o1)c1ccoc1C", "adduct": "[M+H]+"},
        {"smiles": "NC(=O)CSc1ncc(n1Cc1ccccc1)c1ccc(cc1)F", "adduct": "[M+H]+"},
        {"smiles": "O=C(N1CCCC1=O)CSc1nnc(s1)Nc1ccccc1", "adduct": "[M+H]+"},
        {"smiles": "CCCCCNC(=O)c1[nH]cc(c1)c1csc(n1)C", "adduct": "[M+H]+"},
        {"smiles": "COCc1c(oc2c1cccc2)C(=O)Nc1scc(n1)c1ccc(o1)C", "adduct": "[M+H]+"},
        {"smiles": "O=C(C1=C(C)NC(=S)NC1c1ccc(cc1)C)N1CCOCC1", "adduct": "[M+H]+"},
        {"smiles": "Oc1ccc(cc1)N1C(=O)C(=C(C1c1ccc(cc1)Br)C(=O)C)O", "adduct": "[M+H]+"},
        {"smiles": "O=C(c1ccc(cc1)c1ccccc1)NCCCn1cccn1", "adduct": "[M+H]+"},
        {"smiles": "CCN(CC(=O)N1CCc2c1cccc2)CC(=O)NC(C)C", "adduct": "[M+H]+"},
        {"smiles": "O=C(N1CCN(CC1)S(=O)(=O)c1ccc2c(c1)cccc2)C1CC1", "adduct": "[M+H]+"},
        {"smiles": "COC(=O)CCSc1ncnc2c1cc(s2)c1ccccc1", "adduct": "[M+H]+"},
        {"smiles": "CSCCC1NC(=O)N(C1=O)CC(=O)NCCc1ccc(cc1)Cl", "adduct": "[M+H]+"},
        {"smiles": "COc1cc2CN(CCc2cc1OC)S(=O)(=O)c1ccccc1Cl", "adduct": "[M+H]+"},
        {"smiles": "CSc1ccc(cc1)CN1C(=O)NC2(C1=O)CCc1c2cccc1", "adduct": "[M+H]+"},
        {"smiles": "O=C(C1CCCCN1S(=O)(=O)c1ccccc1)Nc1ccc(cc1)F", "adduct": "[M+H]+"},
        {"smiles": "O=C(CN1C(=O)NC(C1=O)C(C)C)Nc1cccc(c1)C(F)(F)F", "adduct": "[M+H]+"},
        {"smiles": "CC(C(=O)Nc1ccccc1)OC(=O)c1nc2n(n1)c(C)ccn2", "adduct": "[M+H]+"},
        {"smiles": "O=C(Nc1ccccc1c1ccccc1)COC(=O)c1cnc(cn1)C", "adduct": "[M+H]+"},
        {"smiles": "COc1ccc(cc1N(S(=O)(=O)C)C)S(=O)(=O)NCCC(C)C", "adduct": "[M+H]+"},
        {"smiles": "CCCn1c(NC(=O)c2cccc(c2)NC(=O)C)nc2c1cccc2", "adduct": "[M+H]+"},
        {"smiles": "O=C(NCC1(CC1)c1ccccc1)CCCn1cnc2c(c1=O)cccc2", "adduct": "[M+H]+"},
        {"smiles": "CCCNC(=O)CN(CC(=O)Nc1ccccc1C(=O)C)C", "adduct": "[M+H]+"},
        {"smiles": "Cc1ccc2c(c1)c(=O)cc(o2)C(=O)N(C(c1ccc(c(c1)F)F)C)C", "adduct": "[M+H]+"},
        {"smiles": "CCn1c(SCC(=O)c2ccc[nH]2)nnc1c1ccc(cc1)Cl", "adduct": "[M+H]+"},
        {"smiles": "Cc1ccc(cc1)S(=O)(=O)N1CCCCC1C(=O)NCc1ccco1", "adduct": "[M+H]+"},
        {"smiles": "CC1CC(C)CN(C1)C(=O)CN1C(=O)NC(C1=O)(C)c1ccccc1Cl", "adduct": "[M+H]+"},
        {"smiles": "COc1ccccc1c1scc(n1)Cn1nnn(c1=O)c1ccccc1", "adduct": "[M+H]+"},
        {"smiles": "O=C(NC(c1ccccc1)c1ccccc1)CCc1c[nH]c2c1cccc2", "adduct": "[M+H]+"},
        {"smiles": "CCCN(C(=O)c1ccc(o1)C)Cc1nnc(o1)c1ccccc1Cl", "adduct": "[M+H]+"},
        {"smiles": "COCCOC(=O)C1=C(C)Nc2n(C1c1ccc(cc1)Cl)ncn2", "adduct": "[M+H]+"},
        {"smiles": "CC(c1nc2ccccc2c(=O)[nH]1)OC(=O)c1cc2c(s1)cccc2", "adduct": "[M+H]+"},
        {"smiles": "CC(c1nnc(o1)c1ccccc1)Sc1nnnn1C1CCCC1", "adduct": "[M+H]+"},
        {"smiles": "COc1ccccc1n1nnnc1SCCOc1ccccc1C(=O)C", "adduct": "[M+H]+"},
        {"smiles": "COC(=O)c1sccc1S(=O)(=O)Nc1ccc2c(c1)sc(n2)C", "adduct": "[M+H]+"},
        {"smiles": "O=C(NC(c1ccc(cc1)F)C)CSc1ncnc2c1c(C)c(s2)C", "adduct": "[M+H]+"},
        {"smiles": "COC(=O)N1CCN(CC1)C(=O)C(Cc1c[nH]c2c1cccc2)NC(=O)C", "adduct": "[M+H]+"},
        {"smiles": "O=C(c1ccc(cc1)n1cccn1)N(Cc1cccs1)Cc1ccco1", "adduct": "[M+H]+"},
        {"smiles": "Cc1cccc(c1)N1CCN(CC1)C(=O)CCc1c(C)nc2n(c1C)ncn2", "adduct": "[M+H]+"},
        {"smiles": "O=C(c1ccccc1Br)Nc1cc(ccc1N1CCOCC1)C(F)(F)F", "adduct": "[M+H]+"},
        {"smiles": "COc1cc2CN(CCc2cc1OC)C(=O)Nc1ccc(cc1F)F", "adduct": "[M+H]+"},
        {"smiles": "O=C1NC(C(=O)N1CC(=O)C(C)(C)C)(C)c1ccc2c(c1)cccc2", "adduct": "[M+H]+"},
        {"smiles": "O=C1NC2(C(=O)N1Cc1onc(n1)c1cccs1)CCOc1c2cccc1", "adduct": "[M+H]+"},
        {"smiles": "COc1cc(c(cc1OC)C)S(=O)(=O)NCCN1CCCC(C1)C", "adduct": "[M+H]+"},
        {"smiles": "O=C(NC1CCCC1)COC(=O)C1c2ccccc2Oc2c1cccc2", "adduct": "[M+H]+"},
        {"smiles": "COc1c(ccc(c1OC)OC)C1N(C)C(=O)c2c1c(=O)c1c(o2)cccc1", "adduct": "[M+H]+"},
        {"smiles": "CCCc1ccc(cc1)C(=O)CN1C(=O)NC2(C1=O)CCCCC2", "adduct": "[M+H]+"},
        {"smiles": "COc1ccc(cc1)CN(C(=O)c1ccc(cc1)C#N)C(C1CC1)C", "adduct": "[M+H]+"},
        {"smiles": "CNC(=O)C1Oc2ccccc2N(C1)C(=O)COc1ccc(c(c1)Cl)Cl", "adduct": "[M+H]+"},
        {"smiles": "Fc1ccc(cc1)S(=O)(=O)N(Cc1cnn(c1)Cc1ccccc1)C", "adduct": "[M+H]+"},
        {"smiles": "Brc1ccc(c(c1)F)Cn1cnc2c(c1=O)cnn2c1ccccc1", "adduct": "[M+H]+"},
        {"smiles": "CCn1c2nnc(n2c2c(c1=O)cccc2)COc1ccc2c(c1)CCC2", "adduct": "[M+H]+"},
        {"smiles": "O=C(N1CCc2c1cccc2)CSc1nnc([nH]1)c1ccccc1Br", "adduct": "[M+H]+"},
        {"smiles": "O=C(Nc1ccc(cc1C)I)CSc1nncn1C(C)C", "adduct": "[M+H]+"},
        {"smiles": "Cc1ccc(c(c1)C)n1nnnc1SCC(=O)c1cc(C)ccc1C", "adduct": "[M+H]+"},
        {"smiles": "C=CCN(c1nc2c(s1)cccc2)C(=O)CN1C(=O)c2c(C1=O)cccc2", "adduct": "[M+H]+"},
        {"smiles": "O=C(NC(c1ccc(cc1)c1ccccc1)C)COc1ccc(cc1)n1cnnn1", "adduct": "[M+H]+"},
        {"smiles": "CCCCCOc1ccc(cc1)C(=O)CCC(=O)N1CCCC1", "adduct": "[M+H]+"},
        {"smiles": "O=C1Nc2ccccc2N(C1)C(=O)CSc1nnc(o1)c1ccccc1", "adduct": "[M+H]+"},
        {"smiles": "O=C(Nc1cc2OCOc2cc1C(=O)C)COc1ccccc1c1ccccc1", "adduct": "[M+H]+"},
        {"smiles": "O=C(C(=O)c1c(C)nn(c1C)c1ccccc1)OCC(=O)c1[nH]ccc1", "adduct": "[M+H]+"},
        {"smiles": "O=C(Nc1ccccc1N1CCOCC1)Cc1csc(n1)c1cscc1", "adduct": "[M+H]+"},
        {"smiles": "N#CC1(CCCCC1)NC(=O)C(Sc1nc2c([nH]1)ccc(c2)OC)C", "adduct": "[M+H]+"},
        {"smiles": "CNc1snc(c1C(=O)OCC(=O)c1cc(n(c1C)C(C)C)C)C", "adduct": "[M+H]+"},
        {"smiles": "COc1cccc(c1O)C1CC(=NN1C(=O)C)c1ccccc1NS(=O)(=O)C", "adduct": "[M+H]+"},
        {"smiles": "O=C(c1cc(nn1c1ccccc1)c1ccco1)Nc1c[nH]nc1", "adduct": "[M+H]+"},
        {"smiles": "O=C(c1ccc(cc1)CS(=O)(=O)C)Nc1cccc2c1CCCC2", "adduct": "[M+H]+"},
        {"smiles": "CCC(NC(=O)CC(c1ccccc1)NC(=O)c1ccccc1Cl)C", "adduct": "[M+H]+"},
        {"smiles": "Brc1cccc(c1)c1noc(n1)CSc1nnnn1C1CCCC1", "adduct": "[M+H]+"},
        {"smiles": "O=C(NC(c1cccc(c1)Cl)C)COC(=O)c1cccnc1SC(F)F", "adduct": "[M+H]+"},
        {"smiles": "CCCC(=O)Nc1ccc(c(c1)C(=O)C)OCc1ccc(cc1)C#N", "adduct": "[M+H]+"},
        {"smiles": "O=S(=O)(c1cccc(c1)c1nnc2n1CCCCC2)N1CCOCC1", "adduct": "[M+H]+"},
        {"smiles": "O=C(Nc1ccc2c(c1)OCCCO2)COc1ccccc1c1ccccc1", "adduct": "[M+H]+"},
        {"smiles": "Fc1ccc(cc1)COc1ccccc1C(=O)Nc1scc(n1)c1cccnc1", "adduct": "[M+H]+"},
        {"smiles": "CCN(C(=O)COC(=O)c1ccc(cc1)C(=O)c1ccccc1)CC", "adduct": "[M+H]+"},
        {"smiles": "COc1cc(OC)c(cc1C(=O)N(CCOc1cccc(c1)C)C)OC", "adduct": "[M+H]+"},
        {"smiles": "CCC(NC(=O)CN1C(=O)NC2(C1=O)CCOc1c2cccc1)(C#C)CC", "adduct": "[M+H]+"},

        # [M-H]- cases (+140)
        {"smiles": "Cc1ccc(cc1F)S(=O)(=O)N", "adduct": "[M-H]-"},
        {"smiles": "O=C(c1ccc[nH]1)CSc1ccc(c(c1)F)F", "adduct": "[M-H]-"},
        {"smiles": "N#CCN1C(=O)NC2(C1=O)CCc1c2cccc1", "adduct": "[M-H]-"},
        {"smiles": "CC1(NC(=O)N(C1=O)CC#N)c1cc(F)ccc1F", "adduct": "[M-H]-"},
        {"smiles": "Fc1ccc(cc1)CNC(=O)c1ccccc1O", "adduct": "[M-H]-"},
        {"smiles": "CCC(NS(=O)(=O)c1ccc(cc1)C(=O)C)C", "adduct": "[M-H]-"},
        {"smiles": "Cc1ccc(cc1)NC(=O)c1cccc(c1)O", "adduct": "[M-H]-"},
        {"smiles": "Brc1ccc(o1)C(=O)Nc1ccccc1OC(F)(F)F", "adduct": "[M-H]-"},
        {"smiles": "Fc1ccc(c(c1)Br)C(=O)NC1CCCCC1", "adduct": "[M-H]-"},
        {"smiles": "CNc1snc(c1C(=O)NCc1cccnc1)C", "adduct": "[M-H]-"},
        {"smiles": "O=C(NC(=O)N)CNc1ccccc1c1ccccc1", "adduct": "[M-H]-"},
        {"smiles": "NC1=Nc2nc3c(n2C(N1)c1csc(n1)C)cccc3", "adduct": "[M-H]-"},
        {"smiles": "CCCCNC(=O)CN1C(=O)c2c(C1=O)c(N)ccc2", "adduct": "[M-H]-"},
        {"smiles": "CC(=O)c1ccc(cc1)OCc1coc(n1)c1cccs1", "adduct": "[M-H]-"},
        {"smiles": "CCCNC(=O)C(NC(=O)c1cccc(c1)Cl)C", "adduct": "[M-H]-"},
        {"smiles": "Brc1ccc(cc1)C(NS(=O)(=O)c1cccs1)C", "adduct": "[M-H]-"},
        {"smiles": "N#CC(Sc1nnc(s1)Nc1ccc(cc1)C)C", "adduct": "[M-H]-"},
        {"smiles": "Ic1cc(ccc1C)NC(=O)c1c(C)noc1C", "adduct": "[M-H]-"},
        {"smiles": "CCc1cccc(c1NC(=O)CCN1C(=O)CCC1=O)C", "adduct": "[M-H]-"},
        {"smiles": "Cc1cc(C)n2c(n1)nc(n2)C(=O)Nc1ccc(c(c1)F)F", "adduct": "[M-H]-"},
        {"smiles": "CSc1cccc(c1)NS(=O)(=O)c1ccc(c(c1)Cl)F", "adduct": "[M-H]-"},
        {"smiles": "O=C(N1CCCC1)CSc1nn2c(n1)nc(cc2C)C", "adduct": "[M-H]-"},
        {"smiles": "Clc1ccc(c(c1)O)C(=O)NCCN1CCOCC1", "adduct": "[M-H]-"},
        {"smiles": "O=C(Nc1cccc(c1)NC(=O)C)CC1CCCC1", "adduct": "[M-H]-"},
        {"smiles": "CCC(NC(=O)CSc1nnc2n1c(C)cc(n2)C)CC", "adduct": "[M-H]-"},
        {"smiles": "O=C(c1c(Cl)cccc1Cl)NC1CCOc2c1cccc2", "adduct": "[M-H]-"},
        {"smiles": "Clc1ccc(c(c1)F)C(=O)NCCc1c[nH]c2c1cccc2", "adduct": "[M-H]-"},
        {"smiles": "COc1cc(ccc1OC(F)F)C=C1C(=O)Nc2c1cccc2", "adduct": "[M-H]-"},
        {"smiles": "CCn1cc(c2c1cccc2)C=C1C(=O)Nc2c1cccc2", "adduct": "[M-H]-"},
        {"smiles": "CCOc1ccccc1NC(=O)C=Cc1cccc(c1)F", "adduct": "[M-H]-"},
        {"smiles": "CCOc1ccccc1c1noc(n1)CSc1nnnn1C", "adduct": "[M-H]-"},
        {"smiles": "O=C(Cc1c(C)noc1C)Nc1ccc2c(c1)nc(o2)C", "adduct": "[M-H]-"},
        {"smiles": "O=C(N1CCCC1)CCNC(=O)c1cc2ccccc2oc1=O", "adduct": "[M-H]-"},
        {"smiles": "COc1ccc(c(c1)C(=O)C=Cc1ccccc1OC)O", "adduct": "[M-H]-"},
        {"smiles": "O=C(Nc1c(C)cccc1C)CN(C(=O)c1cccc(c1)F)C", "adduct": "[M-H]-"},
        {"smiles": "O=C(c1[nH]c(c(c1C)C(=O)C)C)NCc1ccc(n1C)C", "adduct": "[M-H]-"},
        {"smiles": "O=C(Nc1ccc(cc1F)F)CN1C(=O)NC2(C1=O)CCCC2", "adduct": "[M-H]-"},
        {"smiles": "CC(c1ccccc1)NC(=O)C(C1CCCC1)NC(=O)C", "adduct": "[M-H]-"},
        {"smiles": "O=C(Nc1cc(C)ccc1C)CN1CCCCC1C(=O)N", "adduct": "[M-H]-"},
        {"smiles": "Cc1cccc(c1)NC(=O)C(N1CCc2c1cccc2)C", "adduct": "[M-H]-"},
        {"smiles": "O=C(NC(c1ccco1)C)CSc1nnc(n1C1CC1)C1CC1", "adduct": "[M-H]-"},
        {"smiles": "Fc1ccc(c(c1)C1(C)NC(=O)N(C1=O)CC(=O)c1ccc(s1)Cl)F", "adduct": "[M-H]-"},
        {"smiles": "O=C(Cn1cnc2c1c(=O)n(C)c(=O)n2C)NC1CC1", "adduct": "[M-H]-"},
        {"smiles": "CN(Cc1ccccc1)CC(=O)Nc1ccc(cc1Cl)C", "adduct": "[M-H]-"},
        {"smiles": "O=C(CN1C(=O)NC(C1=O)C(C)C)Nc1cccc(c1)C(F)(F)F", "adduct": "[M-H]-"},
        {"smiles": "Clc1ccc(c(c1)N)C(=O)OCc1nc2c(s1)cccc2", "adduct": "[M-H]-"},
        {"smiles": "O=C1CSc2c(N1)cc(cc2)C(=O)N(Cc1ccc(cc1)OC(F)F)C", "adduct": "[M-H]-"},
        {"smiles": "Cc1onc(c1)COC(=O)c1ccc(cc1)c1ccccc1", "adduct": "[M-H]-"},
        {"smiles": "O=C(c1ccc(nc1)C(F)(F)F)OCc1nnnn1c1ccccc1", "adduct": "[M-H]-"},
        {"smiles": "CCCCCNC(=O)c1[nH]cc(c1)c1csc(n1)C", "adduct": "[M-H]-"},
        {"smiles": "COc1ccc(cc1NCC(=O)Nc1cccc(c1)C)C", "adduct": "[M-H]-"},
        {"smiles": "CCCC1NC(=O)N(C1=O)CN(CC(F)(F)F)Cc1ccccc1", "adduct": "[M-H]-"},
        {"smiles": "CCCNC(=O)CN1C(=O)NC(C1=O)(C)c1cccc(c1)Br", "adduct": "[M-H]-"},
        {"smiles": "O=C(C(C)C)NCCNC(=O)c1cc2c([nH]1)ccc(c2)Cl", "adduct": "[M-H]-"},
        {"smiles": "CC(NC(=O)CNC(=O)CCNC(=O)c1ccccc1)C", "adduct": "[M-H]-"},
        {"smiles": "CC(CNC(=O)c1oc2c(c1CSc1nnc[nH]1)cccc2)C", "adduct": "[M-H]-"},
        {"smiles": "O=C(NCc1ccc2c(c1)OCO2)CCCc1c[nH]c2c1cccc2", "adduct": "[M-H]-"},
        {"smiles": "Cn1ccc(c1)C(=O)CSc1nnnn1c1ccc(c(c1)C)C", "adduct": "[M-H]-"},
        {"smiles": "NC(=O)CSc1ccccc1NC(=O)c1cc2c(s1)cccc2", "adduct": "[M-H]-"},
        {"smiles": "O=C(c1ccc2c(c1)OCCCO2)Nc1nc2c(s1)cccc2", "adduct": "[M-H]-"},
        {"smiles": "O=C(C(C)(C)C)C=c1sc(=Cc2ccc(cc2)O)c(=O)[nH]1", "adduct": "[M-H]-"},
        {"smiles": "CCCNC(=O)CNC(=O)c1ccc(cc1)C(C)(C)C", "adduct": "[M-H]-"},
        {"smiles": "Brc1ccccc1S(=O)(=O)Nc1cccc(c1)c1nnnn1C1CC1", "adduct": "[M-H]-"},
        {"smiles": "O=C(Nc1ccccc1c1ccccc1)COC(=O)c1cnc(cn1)C", "adduct": "[M-H]-"},
        {"smiles": "Cc1onc(c1)NC(=O)C(OC(=O)c1ccc2c(c1)OCO2)C", "adduct": "[M-H]-"},
        {"smiles": "O=C(NC1CCCCCC1)CNC(=O)c1ccc2c(c1)OCO2", "adduct": "[M-H]-"},
        {"smiles": "COc1ccc(cc1C(=O)N1CC(C)CC(C1)C)S(=O)(=O)N", "adduct": "[M-H]-"},
        {"smiles": "COCCNC(=O)NC(=O)CSc1ccc2c(c1)OCCO2", "adduct": "[M-H]-"},
        {"smiles": "COC(=O)c1ccccc1NC(=O)CNc1ccc(cc1)C(=O)C", "adduct": "[M-H]-"},
        {"smiles": "O=C(c1sc2c(c1)n1c(n2)scc1)N1CCN(CC1)c1ncccn1", "adduct": "[M-H]-"},
        {"smiles": "COc1cc2CN(CCc2cc1OC)CC(COCC1CCCO1)O", "adduct": "[M-H]-"},
        {"smiles": "COc1ccccc1C(=O)CN1C(=O)NC(C1=O)(C)c1cc2c(o1)cccc2", "adduct": "[M-H]-"},
        {"smiles": "Brc1cccc(c1)c1noc(n1)CSc1nnnn1C1CCCC1", "adduct": "[M-H]-"},
        {"smiles": "CCOC(=O)c1sc(nc1C)NC(=O)C=Cc1cscc1", "adduct": "[M-H]-"},
        {"smiles": "COCCn1nc(ccc1=O)C(=O)NCC=C", "adduct": "[M-H]-"},
        {"smiles": "O=C(C(Sc1nc2sccc2c(=O)[nH]1)C)Nc1cccc(c1)C(=O)C", "adduct": "[M-H]-"},
        {"smiles": "COc1cccc(c1)NC(=O)c1cc(OC)c(c(c1)OC)Br", "adduct": "[M-H]-"},
        {"smiles": "Brc1ccc(c(c1)Cl)S(=O)(=O)Nc1ccccc1N1CCOCC1", "adduct": "[M-H]-"},
        {"smiles": "CCOC(=O)Cc1csc(n1)SCC(=O)NC1CCCCC1", "adduct": "[M-H]-"},
        {"smiles": "COc1ccc(cc1)S(=O)(=O)NCCn1c(C)nc2c1cccc2", "adduct": "[M-H]-"},
        {"smiles": "O=C(c1ccc(cc1)c1ccccc1)NCCCn1cccn1", "adduct": "[M-H]-"},
        {"smiles": "O=C(Nc1cc2OCOc2cc1C(=O)C)CSc1nnnn1C1CCCC1", "adduct": "[M-H]-"},
        {"smiles": "CC(N(C(=O)COC(=O)c1cc(=O)[nH]c2c1cccc2)C(C)C)C", "adduct": "[M-H]-"},
        {"smiles": "Clc1ccccc1Oc1ccc(cc1)Nc1ccc2n(n1)nnn2", "adduct": "[M-H]-"},
        {"smiles": "NC(=O)C1Cc2ccccc2CN1S(=O)(=O)c1ccc(cc1)Br", "adduct": "[M-H]-"},
        {"smiles": "O=C(CSc1nc(C)cc(n1)C)NCCc1c[nH]c2c1cccc2", "adduct": "[M-H]-"},
        {"smiles": "O=C(NC(c1ccccc1)c1ccccc1)CCc1c[nH]c2c1cccc2", "adduct": "[M-H]-"},
        {"smiles": "Clc1ccccc1c1noc(n1)CSc1nnc2n1CCCCC2", "adduct": "[M-H]-"},
        {"smiles": "NC(=O)CSc1nnc(n1c1ccccc1)Cc1cccs1", "adduct": "[M-H]-"},
        {"smiles": "COc1cc(ccc1OC)c1nn(cc1C(=O)N(Cc1ccco1)C)c1ccccc1", "adduct": "[M-H]-"},
        {"smiles": "Cc1ccc(cc1)c1nnc2n1N=C(CS2)c1cc2c(o1)cccc2", "adduct": "[M-H]-"},
        {"smiles": "COC(=O)c1cn(c2c1cccc2)Cc1cc(ccc1OC)C(=O)C", "adduct": "[M-H]-"},
        {"smiles": "O=C(N1CCCCCCC1)CN1C(=O)NC2(C1=O)CCCCCCC2", "adduct": "[M-H]-"},
        {"smiles": "Brc1ccc(o1)C(=O)Nc1ccc2c(c1)OC1(O2)CCCC1", "adduct": "[M-H]-"},
        {"smiles": "COc1cc(cc(c1O)Br)C=C(c1nc2ccccc2c(=O)[nH]1)Cl", "adduct": "[M-H]-"},
        {"smiles": "COc1ccc(c(c1)C(=O)Nc1scc(n1)c1ccccc1)Br", "adduct": "[M-H]-"},
        {"smiles": "COc1ccccc1C=CC(=O)Nc1ccccc1C(C)(C)C", "adduct": "[M-H]-"},
        {"smiles": "COc1ccc(cc1C(NC(=O)COC(=O)c1c(NC)snc1C)C)OC", "adduct": "[M-H]-"},
        {"smiles": "COc1ccc(cc1COc1ccc(cc1)NC(=O)C)C(=O)C", "adduct": "[M-H]-"},
        {"smiles": "O=C(CN1C(=O)NC(C1=O)(C)C)OC(c1nnc(o1)c1ccc(cc1)C)C", "adduct": "[M-H]-"},
        {"smiles": "C=CCN(C(=O)CN1C(=O)NC2(C1=O)CCCCc1c2cccc1)CC=C", "adduct": "[M-H]-"},
        {"smiles": "CCC(c1ccc(cc1)NC(=O)COc1ccc(cc1)C)C", "adduct": "[M-H]-"},
        {"smiles": "CC(c1nc2ccccc2c(=O)[nH]1)OC(=O)c1cc2c(s1)cccc2", "adduct": "[M-H]-"},
        {"smiles": "O=C(c1[nH]c(c(c1C)C(=O)C)C)Cn1nnn(c1=O)c1ccccc1", "adduct": "[M-H]-"},
        {"smiles": "O=C(CN1C(=O)NC(C1=O)(C)c1cc(F)ccc1F)NCC1CCCCC1", "adduct": "[M-H]-"},
        {"smiles": "CC(=O)c1ccccc1OCCSc1n[nH]c(=O)n1CC1CCCO1", "adduct": "[M-H]-"},
        {"smiles": "O=C(NC(=O)NC(C)(C)C)CCSc1nccc(n1)C(F)(F)F", "adduct": "[M-H]-"},
        {"smiles": "O=C(c1cn(c2c1cccc2)C)OCc1cc(=O)n2c(n1)scc2", "adduct": "[M-H]-"},
        {"smiles": "O=C(NC1CCOc2c1cccc2)CCNC(=O)OC(C)(C)C", "adduct": "[M-H]-"},
        {"smiles": "NC(=O)CSc1nnc(n1c1ccc(cc1)Cl)Cc1cccs1", "adduct": "[M-H]-"},
        {"smiles": "CC(c1nc2ccccc2c(=O)[nH]1)Sc1nnc(n1C1CC1)c1ccncc1", "adduct": "[M-H]-"},
        {"smiles": "O=C(NC(c1ccccc1)Cc1ccccc1)CCc1c(C)noc1C", "adduct": "[M-H]-"},
        {"smiles": "O=C(NC(=O)NC(C)(C)C)COC(=O)C12CC3CC(C1)CC(C2)(C3)O", "adduct": "[M-H]-"},
        {"smiles": "Cc1ccc(cc1)C1CC(=NN1C(=O)c1ccco1)c1ccccc1O", "adduct": "[M-H]-"},
        {"smiles": "COc1cc(Cl)c(cc1N1C(=O)CC(C1=O)N1CCCCC1)C", "adduct": "[M-H]-"},
        {"smiles": "COCc1c(oc2c1cccc2)C(=O)Nc1scc(n1)c1ccc(o1)C", "adduct": "[M-H]-"},
        {"smiles": "O=C1NC2(C(=O)N1CC(=O)N1CCc3c1cccc3)CCOc1c2cccc1", "adduct": "[M-H]-"},
        {"smiles": "CCOc1ccc(cc1)NC(=O)CSc1ncnc2c1cccc2", "adduct": "[M-H]-"},
        {"smiles": "O=C(CN1C(=O)NC2(C1=O)CCCc1c2cccc1)NC1CCCCC1", "adduct": "[M-H]-"},
        {"smiles": "O=C(N1CCc2c(C1)ccs2)CCC(=O)c1ccc2c(c1)CCC2", "adduct": "[M-H]-"},
        {"smiles": "CSc1ccc(c(c1)OC)C(=O)OCC(=O)c1ccc2c(c1)cccc2", "adduct": "[M-H]-"},
        {"smiles": "O=C(CN1C(=O)N(C(C1=O)C)c1ccc(cc1)C)NCCC1=CCCCC1", "adduct": "[M-H]-"},
        {"smiles": "FC(Oc1ccc(cc1)C1(C)NC(=O)N(C1=O)CC(=O)NC1CCCC1)F", "adduct": "[M-H]-"},
        {"smiles": "CNC(=O)C1CCCCN1C(C(=O)Nc1ccccc1c1ccccc1)C", "adduct": "[M-H]-"},
        {"smiles": "O=C(CN1C(=O)NC(C1=O)(C)c1ccc(cc1)C)NC1CCCCCC1", "adduct": "[M-H]-"},
        {"smiles": "NC(=O)c1ccc(cc1)CN1C(=O)NC2(C1=O)CCCCc1c2cccc1", "adduct": "[M-H]-"},
        {"smiles": "CCCNC(=O)NC(=O)COC(=O)c1oc2c(c1C)cc(cc2)OC", "adduct": "[M-H]-"},
        {"smiles": "COC(=O)c1ccc(cc1)C=CC(=O)Nc1cc(C)ccc1OC", "adduct": "[M-H]-"},
        {"smiles": "COC(=O)CNC(=O)C1CCCN1S(=O)(=O)c1cc(C)c(cc1C)Cl", "adduct": "[M-H]-"},
        {"smiles": "N#CC(=c1sc(=Cc2ccccc2OC)c(=O)n1CC)C(=O)N1CCOCC1", "adduct": "[M-H]-"},
        {"smiles": "O=C(NCc1ccccc1OC(F)F)CCCn1c(=O)oc2c1cccc2", "adduct": "[M-H]-"},
        {"smiles": "O=C(CCNC1=NS(=O)(=O)c2c1cccc2)NCCc1ccc(cc1)C(C)C", "adduct": "[M-H]-"},
        {"smiles": "CCOC(=O)CCCNC(=O)CNC(=O)c1ccc(c(c1)Cl)Cl", "adduct": "[M-H]-"},
        {"smiles": "COc1cc(OC)c(cc1C(=O)N(CCOc1cccc(c1)C)C)OC", "adduct": "[M-H]-"},
        {"smiles": "O=C(N1CCOCC1)CN1C(=O)NC(C1=O)(Cc1ccccc1)Cc1ccccc1", "adduct": "[M-H]-"},
        {"smiles": "CCN(S(=O)(=O)c1cccc(c1)NS(=O)(=O)c1ccc2c(c1)cccc2)CC", "adduct": "[M-H]-"},
        {"smiles": "CNC(=O)C1Cc2ccccc2CN1C(=O)COc1cccc(c1)F", "adduct": "[M-H]-"},
        {"smiles": "CC1CCC2(CC1)NC(=O)N(C2=O)CN1CCN(CC1)c1ccccc1C", "adduct": "[M-H]-"},
        {"smiles": "O=C(C(Sc1nn2c(n1)nc(cc2C)C)C)Nc1cc(Cl)ccc1Cl", "adduct": "[M-H]-"},
        {"smiles": "OCC1OC(OCc2ccc(cc2)OC)C(C(C1O)O)NC(=O)C", "adduct": "[M-H]-"},

        # [M+Na]+ cases (+140)
        {"smiles": "COC(=O)C1CCCN1S(=O)(=O)C", "adduct": "[M+Na]+"},
        {"smiles": "Fc1ccc2c(c1F)n1CCCCCCc1n2", "adduct": "[M+Na]+"},
        {"smiles": "CONC(=O)NCc1ccc2c(c1)ccc(c2)OC", "adduct": "[M+Na]+"},
        {"smiles": "O=C(Cn1cc(ccc1=O)C(F)(F)F)Nc1ccccc1", "adduct": "[M+Na]+"},
        {"smiles": "C=CCN(S(=O)(=O)c1ccccc1C(F)(F)F)CC(=O)O", "adduct": "[M+Na]+"},
        {"smiles": "CONC(=O)c1cc(nn1c1ccccc1)C1CC1", "adduct": "[M+Na]+"},
        {"smiles": "O=C(Cc1ccccc1C)NCc1ccc(cc1)S(=O)(=O)N", "adduct": "[M+Na]+"},
        {"smiles": "O=C(Cc1c(C)noc1C)NCc1ccc(cc1)C", "adduct": "[M+Na]+"},
        {"smiles": "O=C1N(C)c2c(C1=Cc1ccc(cc1)C(=O)O)cccc2", "adduct": "[M+Na]+"},
        {"smiles": "COc1cc(C=CC(=O)c2ccc(cc2)O)cc(c1O)OC", "adduct": "[M+Na]+"},
        {"smiles": "COc1ccc(cc1C(=O)NCC(F)(F)F)S(=O)(=O)N", "adduct": "[M+Na]+"},
        {"smiles": "FC(Oc1ccc(cc1)CN(C(=O)c1ccc(nc1)O)C)F", "adduct": "[M+Na]+"},
        {"smiles": "O=C(N(C)C)CN1C(=O)NC(C1=O)(C)c1ccccc1Br", "adduct": "[M+Na]+"},
        {"smiles": "FC(Oc1ccccc1C(=O)Nc1ccccc1c1ccccc1)F", "adduct": "[M+Na]+"},
        {"smiles": "O=C(Cc1ccc(cc1)C)NCC1COc2c(O1)cccc2", "adduct": "[M+Na]+"},
        {"smiles": "COc1c(OC)cccc1C(=O)NCCSCc1ccco1", "adduct": "[M+Na]+"},
        {"smiles": "CCOc1ccccc1NC(=O)C1(CC1)c1ccccc1", "adduct": "[M+Na]+"},
        {"smiles": "O=C(Cn1cc(ccc1=O)C(F)(F)F)NCc1ccc(cc1)F", "adduct": "[M+Na]+"},
        {"smiles": "COc1ccc(cc1C(=O)NCC(C)C)S(=O)(=O)N", "adduct": "[M+Na]+"},
        {"smiles": "CCC(NC(=O)CN1C(=O)C2C(C1=O)CCCC2)CC", "adduct": "[M+Na]+"},
        {"smiles": "COc1ccccc1CCC(=O)Nc1ccc(cc1)N(C)C", "adduct": "[M+Na]+"},
        {"smiles": "CCC(C(C(=O)OC)NS(=O)(=O)c1ccc(cc1)C)C", "adduct": "[M+Na]+"},
        {"smiles": "NC(=O)C1CC(=NN1c1ccccc1)C(=O)NCc1ccco1", "adduct": "[M+Na]+"},
        {"smiles": "Fc1ccc(c(c1)F)C(=O)NC(c1cccc(c1)S(=O)(=O)N)C", "adduct": "[M+Na]+"},
        {"smiles": "COc1ccc(cc1)C(=O)COC(=O)CCc1c[nH]c2c1cccc2", "adduct": "[M+Na]+"},
        {"smiles": "O=C(NCC1COc2c(O1)cccc2)NCCOc1ccc(cc1)Cl", "adduct": "[M+Na]+"},
        {"smiles": "CN(C(=O)c1cccc(c1)S(=O)(=O)C)CCOc1ccccc1C", "adduct": "[M+Na]+"},
        {"smiles": "COc1ccccc1CNC(=O)c1ccc2c(c1)nn[nH]2", "adduct": "[M+Na]+"},
        {"smiles": "COc1ccc(cc1OC)CNC(=O)C(Oc1ccc2c(c1)OCCO2)C", "adduct": "[M+Na]+"},
        {"smiles": "COc1ccccc1CNC(=O)CSc1nnc2n1c1ccccc1[nH]2", "adduct": "[M+Na]+"},
        {"smiles": "Clc1ccc(s1)S(=O)(=O)N1CCN(CC1)C(=O)c1ccccn1", "adduct": "[M+Na]+"},
        {"smiles": "COC(=O)C1(CCCCC1)NS(=O)(=O)N1CCCC(C1)C", "adduct": "[M+Na]+"},
        {"smiles": "COc1cc(C)ccc1OCc1cccc(c1)S(=O)(=O)N(C)C", "adduct": "[M+Na]+"},
        {"smiles": "O=C(C(NC(=O)C(C)(C)C)C)Nc1ccc(cc1C)Cl", "adduct": "[M+Na]+"},
        {"smiles": "O=C(C(Sc1nc2ccccc2c2n1nc(n2)C)C)NCC1CCCO1", "adduct": "[M+Na]+"},
        {"smiles": "O=C(CN(S(=O)(=O)c1cccs1)C)NC1CCCc2c1cccc2", "adduct": "[M+Na]+"},
        {"smiles": "O=C(NC(C)(C)C)COC(=O)C1(CCCC1)c1cccc(c1)F", "adduct": "[M+Na]+"},
        {"smiles": "O=C(C1(CC1)c1ccccc1)Nc1cccc(c1)c1nnco1", "adduct": "[M+Na]+"},
        {"smiles": "O=C(C1CC1)N1CCN(CC1)C(=O)c1c(C)onc1c1c(F)cccc1Cl", "adduct": "[M+Na]+"},
        {"smiles": "O=C(Nc1ncc(c(c1Cl)C)Cl)COC(=O)c1cccn1C", "adduct": "[M+Na]+"},
        {"smiles": "O=C(N1CCCCC1c1nc2c(s1)cccc2)C1CCCC1", "adduct": "[M+Na]+"},
        {"smiles": "O=C1Nc2ccccc2N(C1)C(=O)Cc1coc2c1cc(C(C)C)c(c2)C", "adduct": "[M+Na]+"},
        {"smiles": "CCOc1ccc(cc1CN1C(=O)NC(C1=O)Cc1ccccc1)C(=O)C", "adduct": "[M+Na]+"},
        {"smiles": "CC(Cc1ccc(cc1)c1ccccc1)NC(=O)CNC(=O)C1CC1", "adduct": "[M+Na]+"},
        {"smiles": "CN(C(=O)COC(=O)c1c(sc(c1C)C)NC(=O)C(C)(C)C)C", "adduct": "[M+Na]+"},
        {"smiles": "CC(C1NC(=O)N(C1=O)Cc1cc(Cl)c2c(c1)OCCCO2)C", "adduct": "[M+Na]+"},
        {"smiles": "COc1ccc(cc1)CN(C(C(=O)Nc1ccccc1C)C)C", "adduct": "[M+Na]+"},
        {"smiles": "CCC1(NC(=O)N(C1=O)CC(=O)c1cccn1C)c1ccccc1", "adduct": "[M+Na]+"},
        {"smiles": "Cc1ccc(cc1C)S(=O)(=O)N1CCC(CC1)NS(=O)(=O)C", "adduct": "[M+Na]+"},
        {"smiles": "O=C(NC(c1ccccn1)C)CSc1nnc(n1C)c1ccccc1F", "adduct": "[M+Na]+"},
        {"smiles": "O=C(CSc1nnc(n1Cc1ccco1)C1CC1)NCCc1ccccc1", "adduct": "[M+Na]+"},
        {"smiles": "O=C(C1COc2c(O1)cccc2)NC1CC2C(C1(C)CC2)(C)C", "adduct": "[M+Na]+"},
        {"smiles": "Fc1ccc(cc1)C(=O)Cn1cc(ccc1=O)S(=O)(=O)N(C)C", "adduct": "[M+Na]+"},
        {"smiles": "CCC(NC(=O)CN1C(=O)NC(C1=O)(C)c1ccccc1)(C#C)CC", "adduct": "[M+Na]+"},
        {"smiles": "O=C(NCc1ccc2c(c1)OCO2)COC(=O)C1(CCCC1)c1cccc(c1)F", "adduct": "[M+Na]+"},
        {"smiles": "CCc1ccc(cc1)C(=O)N1CCN(CC1)C(=O)c1ccccc1", "adduct": "[M+Na]+"},
        {"smiles": "O=C(C1CCCN1S(=O)(=O)c1ccccc1C(F)(F)F)NCc1ccc(cc1)C", "adduct": "[M+Na]+"},
        {"smiles": "Brc1ccc(o1)C(=O)OC(C(=O)c1c(C)[nH]c2c1cccc2)C", "adduct": "[M+Na]+"},
        {"smiles": "CCOC(=O)c1csc(n1)NC(=O)C=Cc1ccc(o1)Br", "adduct": "[M+Na]+"},
        {"smiles": "O=C(C1CC(=O)N(C1)C1CC1)Nc1cc(Cl)ccc1Oc1ccccc1", "adduct": "[M+Na]+"},
        {"smiles": "O=C(N(C1CCCCC1)C)C(Sc1nc2ccsc2c(=O)n1C)C", "adduct": "[M+Na]+"},
        {"smiles": "Nc1nc(COc2ccc3c(c2)CCC3)nc(n1)Nc1ccc(cc1)C", "adduct": "[M+Na]+"},
        {"smiles": "CC1c2ccsc2CCN1c1nc(nc2c1cccc2)c1cccnc1", "adduct": "[M+Na]+"},
        {"smiles": "CC(C(=O)N(c1ccccc1)C)SC1=C(C#N)C(C(C(=O)N1)C#N)(C)C", "adduct": "[M+Na]+"},
        {"smiles": "N#CC1(CCCCC1)NC(=O)CSc1nncn1c1ccccc1C", "adduct": "[M+Na]+"},
        {"smiles": "Brc1ccc(cc1)C1=NN(C(C1)c1ccccc1O)C(=O)C", "adduct": "[M+Na]+"},
        {"smiles": "O=C(C1OC(=O)c2c(C1)cccc2)NCc1nnc2n1cc(cc2)C(F)(F)F", "adduct": "[M+Na]+"},
        {"smiles": "NC1=Nc2nc3c(n2C(N1)c1cc(C)c(c(c1)C)O)cccc3", "adduct": "[M+Na]+"},
        {"smiles": "COc1cc(ccc1OC)NC(=O)c1cccc(c1)S(=O)(=O)N1CCCC1", "adduct": "[M+Na]+"},
        {"smiles": "N#CC1(CCCCC1)NC(=O)C(Oc1ccc2c(c1)oc(=O)cc2C)C", "adduct": "[M+Na]+"},
        {"smiles": "Cc1nn2c(s1)nc(cc2=O)CN(c1ccccc1c1ccccc1)C", "adduct": "[M+Na]+"},
        {"smiles": "COc1ccc(cc1)CNC(=O)C(OC(=O)c1cc2c([nH]1)cccc2)C", "adduct": "[M+Na]+"},
        {"smiles": "O=C(NC(C12CC3CC(C2)CC(C1)C3)C)CSc1nccn1C", "adduct": "[M+Na]+"},
        {"smiles": "O=C(COc1ccccc1c1ccccc1)NCC1CCCCC1", "adduct": "[M+Na]+"},
        {"smiles": "O=C(NC1CCCCCC1)CCN1C(=O)C2C(C1=O)CCCC2", "adduct": "[M+Na]+"},
        {"smiles": "CCN(c1ccccc1)C(=O)COC(=O)c1ccc(cc1OC)OC", "adduct": "[M+Na]+"},
        {"smiles": "O=C(NC1CCCC1)COC(=O)C1c2ccccc2Oc2c1cccc2", "adduct": "[M+Na]+"},
        {"smiles": "CCCC(=O)N1N=C(CC1c1ccccc1O)c1ccc(cc1)C", "adduct": "[M+Na]+"},
        {"smiles": "O=C(NC12CC3CC(C2)CC(C1)C3)COC(=O)c1ccccc1", "adduct": "[M+Na]+"},
        {"smiles": "O=C(CN1C(=O)NC2(C1=O)CCCCC2)NC(=O)NC(C)(C)C", "adduct": "[M+Na]+"},
        {"smiles": "COc1cccc(c1)N1CC(CC1=O)C(=O)NCc1ccc(cc1)Cn1cncc1", "adduct": "[M+Na]+"},
        {"smiles": "O=C1OCCC1Sc1nnc(n1C1CCCCC1)c1ccccc1F", "adduct": "[M+Na]+"},
        {"smiles": "CCOC(=O)COc1ccccc1C1=NN(C(C1)c1ccccc1)C(=O)C", "adduct": "[M+Na]+"},
        {"smiles": "O=C(Nc1cccc2c1cccc2)Nc1ccc2c(c1)CCN2C(=O)C", "adduct": "[M+Na]+"},
        {"smiles": "CCCCn1nnnc1CN1C(=O)N(C(C1=O)C)c1ccc(cc1)F", "adduct": "[M+Na]+"},
        {"smiles": "CCOc1ncccc1C(=O)Nc1nnc(s1)Cc1ccccc1", "adduct": "[M+Na]+"},
        {"smiles": "NCC(=O)OCc1ccccc1", "adduct": "[M+Na]+"},
        {"smiles": "Clc1ccc2c(c1)nc(n(c2=O)Cc1ccco1)SCc1ccccn1", "adduct": "[M+Na]+"},
        {"smiles": "O=C1CCCN1c1ccccc1NC(=O)c1ccc(cc1)NC(=O)c1ccco1", "adduct": "[M+Na]+"},
        {"smiles": "N#CC1(CCCC1)NC(=O)CN1CCCCC1c1nc2c(s1)cccc2", "adduct": "[M+Na]+"},
        {"smiles": "CCN(C(=O)c1oc2c(c1C)cc(cc2)Cl)C1CCS(=O)(=O)C1", "adduct": "[M+Na]+"},
        {"smiles": "CC(c1nnc(o1)c1ccccc1)Sc1nnnn1Cc1cccs1", "adduct": "[M+Na]+"},
        {"smiles": "COc1ccc(cc1NC(=O)C1=NN(C(C1)C(=O)N)c1ccccc1)C", "adduct": "[M+Na]+"},
        {"smiles": "CCCOc1ccc(cc1OCC)C(=O)Nc1ccccc1", "adduct": "[M+Na]+"},
        {"smiles": "O=C(N1CCN(CC1)C(=O)c1nn(c(cc1=O)C)c1ccccc1C(F)(F)F)C1CC1", "adduct": "[M+Na]+"},
        {"smiles": "O=C(CN1C(=O)NC2(C1=O)CCc1c2cccc1)NC1CCCCC1", "adduct": "[M+Na]+"},
        {"smiles": "CCC(=O)N1N=C(CC1c1ccccc1O)c1ccc(cc1)Br", "adduct": "[M+Na]+"},
        {"smiles": "O=C(NC(C12CC3CC(C2)CC(C1)C3)C)CCNC(=O)c1cccs1", "adduct": "[M+Na]+"},
        {"smiles": "COc1ccccc1N1CCN(CC1)CC(=O)NC(c1ccccc1)C", "adduct": "[M+Na]+"},
        {"smiles": "O=C(Nc1nnc(s1)C1CC1)COc1c(C)cc(cc1C)C", "adduct": "[M+Na]+"},
        {"smiles": "COc1ccc(cc1)c1scc(n1)CC(=O)NCC(c1cccs1)N1CCOCC1", "adduct": "[M+Na]+"},
        {"smiles": "O=C1CC(CN1c1ccc(c(c1)Cl)C)c1onc(n1)c1ccc2c(c1)OCO2	", "adduct": "[M+Na]+"},
        {"smiles": "CC1CCC2(CC1)NC(=O)N(C2=O)CC(=O)NC1CCCc2c1cccc2", "adduct": "[M+Na]+"},
        {"smiles": "N#CC1(CCCCC1)NC(=O)Cn1cnc2c(c1=O)cc(cc2Cl)Cl", "adduct": "[M+Na]+"},
        {"smiles": "Cc1ccc(c(c1)C)n1nnnc1SCC(=O)c1cc(C)ccc1C", "adduct": "[M+Na]+"},
        {"smiles": "O=C(N1CCN(CC1)c1ccccc1F)CSc1nnc(n1C1CC1)C", "adduct": "[M+Na]+"},
        {"smiles": "O=C(N(C1=CCCCC1)Cc1ccco1)Cn1nnn(c1=O)c1ccccc1", "adduct": "[M+Na]+"},
        {"smiles": "CCS(=O)(=O)N1CCC(CC1)NC(=O)C(C(C)C)NC(=O)OCc1ccccc1", "adduct": "[M+Na]+"},
        {"smiles": "COc1ccc(cc1)CN(C(=O)c1ccc(cc1)C#N)C(C1CC1)C", "adduct": "[M+Na]+"},
        {"smiles": "Cc1cc(C)n2c(n1)nc(n2)C(=O)OCC(=O)c1cc(n(c1C)C)C", "adduct": "[M+Na]+"},
        {"smiles": "CCC(NC(=O)c1cc(ccc1Br)S(=O)(=O)N1CCCCC1)C", "adduct": "[M+Na]+"},
        {"smiles": "COc1ccc(cc1)N1CCN(CC1)C(=O)CNC(=O)c1ccc(cc1)F", "adduct": "[M+Na]+"},
        {"smiles": "Cc1ccc(cc1)S(=O)(=O)N1CCN(CC1)Cn1sc2c(c1=O)cccc2", "adduct": "[M+Na]+"},
        {"smiles": "O=C(c1c[nH]nc1c1ccc(cc1)F)N1CCC(CC1)C(=O)N1CCCCCC1", "adduct": "[M+Na]+"},
        {"smiles": "COc1ccc(cc1)N1CCN(CC1)CN1C(=O)NC2(C1=O)CCCc1c2ccs1", "adduct": "[M+Na]+"},
        {"smiles": "O=C(NCc1cn(nc1c1ccc(cc1)F)c1ccccc1)CNC(=O)C1CC1C", "adduct": "[M+Na]+"},
        {"smiles": "O=C(NC(C(=O)C)Cc1ccccc1)CSc1nnc(n1C)c1ccccc1F", "adduct": "[M+Na]+"},
        {"smiles": "O=C(CN(C(=O)c1cc(C)cc(c1)C)C)Nc1ccc(cc1)Cl", "adduct": "[M+Na]+"},
        {"smiles": "Clc1ccc(cc1)c1nnn(n1)CC(=O)c1cccc(c1)N1CCCC1=O", "adduct": "[M+Na]+"},
        {"smiles": "O=C(Nc1ccc(cc1)C)Cc1scc(n1)CN(C1CC1)Cc1ccc(cc1)F", "adduct": "[M+Na]+"},
        {"smiles": "CC(=O)OCC(=O)N1N=C(CC1c1ccccc1)c1ccc2c(c1)cccc2", "adduct": "[M+Na]+"},
        {"smiles": "COc1cc2c(cc1OC)CCN(C2c1ccccc1)C(=O)COC(=O)C1CC1", "adduct": "[M+Na]+"},
        {"smiles": "C=CCn1c(SCC(=O)NC(=O)Nc2ccc3c(c2)OCCO3)nnc1C", "adduct": "[M+Na]+"},
        {"smiles": "CN(S(=O)(=O)c1cccs1)CCS(=O)(=O)NCC1CCCCC1", "adduct": "[M+Na]+"},
        {"smiles": "O=C(CN1C(=O)NC2(C1=O)CCc1c2cccc1)NC1CCCCC1C", "adduct": "[M+Na]+"},
        {"smiles": "O=C(N1CCCCCCC1)CN1C(=O)NC(C1=O)(C)c1cc2c(o1)cccc2", "adduct": "[M+Na]+"},
        {"smiles": "COc1ccc(cc1)CNc1nnc(s1)SCC(=O)N1CCSC1c1ccccc1F", "adduct": "[M+Na]+"},
        {"smiles": "O=C(N1C(COc2c1cccc2)c1ccccc1)CSc1nnc(o1)c1cccs1", "adduct": "[M+Na]+"},
        {"smiles": "CCOC(=O)c1c(C)[nH]c(c1C(=O)COC(=O)c1ccc2c(c1)OCCO2)C", "adduct": "[M+Na]+"},
        {"smiles": "O=C(c1cn(nc1c1ccccc1)Cc1ccccc1Cl)N1CCOCC1", "adduct": "[M+Na]+"},
        {"smiles": "COc1ccc(cc1)C(=O)N[C@H](C(=O)Nc1cccc(c1)C(=O)C)C(C)C", "adduct": "[M+Na]+"},
        {"smiles": "COC(=O)c1ccc(cc1)OCC(=O)Nc1ccccc1C(=O)NC(c1ccccc1)C", "adduct": "[M+Na]+"},
        {"smiles": "Oc1ccc(cc1)c1ccc(cc1)C(=O)OCC(=O)c1ccccc1", "adduct": "[M+Na]+"},
        {"smiles": "O=C(NC(c1ccccc1)c1ccccc1)CCS(=O)(=O)c1ccccc1Cl", "adduct": "[M+Na]+"},
        {"smiles": "O=C(CN1CCCC1c1ccc2c(c1)OCCCO2)NC(=O)NC1CCCCC1", "adduct": "[M+Na]+"},
        {"smiles": "COCCOC(=O)C1=C(C)N(c2ccccc2)C(=S)NC1c1ccccc1C", "adduct": "[M+Na]+"},
        {"smiles": "CCCNC(=O)C(NC(=O)c1ccc2c(c1)c1CCCCc1[nH]2)C", "adduct": "[M+Na]+"},
        {"smiles": "O=C(NCc1nnc2n1cc(cc2)C(F)(F)F)COc1ccccc1c1ccccc1", "adduct": "[M+Na]+"},
        {"smiles": "Cc1cc(ccc1C)S(=O)(=O)N(Cc1nnnn1C1CC1)Cc1ccccc1", "adduct": "[M+Na]+"},
        {"smiles": "CCSc1nnc(s1)NC(=O)[C@H](C(C)C)NC(=O)c1ccccc1Br", "adduct": "[M+Na]+"},
    ]

    prompts_func = {
        "simple": lambda s, a: construir_prompt_simple(s, a),
        "con conocimiento experto": lambda s, a: construir_prompt_ce(s, a),
        "con ejemplos": lambda s, a: construir_prompt_ejemplos(s, a, DATOS_TRAIN, n=5),
        "completo": lambda s, a: construir_prompt_completo(s, a, DATOS_TRAIN, n=5),
    }

    variantes = list(prompts_func.keys())

    # Inicializar Excel con timestamp para no sobrescribir ejecuciones previas
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    excel_path = f"../data/results/predicciones_prompts_{timestamp}.xlsx"
    os.makedirs(os.path.dirname(excel_path), exist_ok=True)
    inicializar_excel(excel_path, variantes)

    print("=" * 70)
    print(f"TEST DE PROMPTS\nInicio de prueba: {datetime.now()}")

    for i, caso in enumerate(test_cases, 1):
        smiles, adduct = caso["smiles"], caso["adduct"]
        print(f"{'='*70}")
        print(f"COMPUESTO {i} | Aducto = {adduct}")
        print(f"SMILES: {smiles[:60]}...")
        print(f"{'='*70}")

        # Limpiar caché entre predicciones
        torch.cuda.empty_cache()  # por si acaso aunque estés en CPU
        if hasattr(MODEL, 'reset_cache'):
            MODEL.reset_cache()

        # Crear la fila base del compuesto (predicciones vacías, se rellenarán abajo)
        fila_base = [i, smiles, adduct] + [None] * len(variantes)
        guardar_fila_excel(excel_path, fila_base)
        # Cabecera = fila 1, así que el compuesto i va a fila i+1
        fila_idx = i + 1

        for j, (nombre, fn_prompt) in enumerate(prompts_func.items()):
            prompt = fn_prompt(smiles, adduct)
            print(f"Prompt {nombre}: ")
            print(f" - Hora de ejecución: {datetime.now().strftime('%H:%M:%S')}")

            ejemplos = seleccionar_ejemplos(DATOS_TRAIN, smiles, adduct, n=5)
            resultado = predecir_ccs(MODEL, TOKENIZER, prompt, STATS, ejemplos)

            fallback_str = " [FALLBACK]" if resultado["fallback"] else ""
            print(f" - CCS = {resultado['predicted_ccs']:.2f} Å²{fallback_str}")
            print(f" - Reasoning: {resultado['reasoning'][:80]}")

            # Guardar resultado inmediatamente en el Excel
            # Columnas: 1=id, 2=smiles, 3=aducto, 4..=variantes
            col_idx = 4 + j
            actualizar_celda_excel(excel_path, fila_idx, col_idx, resultado["predicted_ccs"])

    print(f"\n{'='*70}")
    print(f"TEST FINALIZADO!\nFin de prueba: {datetime.now()}")
    print("=" * 70)


# Para archivar resultados
class Logger:
    def __init__(self, filepath):
        self.terminal = sys.stdout
        os.makedirs(os.path.dirname(filepath), exist_ok=True)
        self.log = open(filepath, 'w', encoding='utf-8')

    def write(self, message):
        self.terminal.write(message)
        self.log.write(message)
        self.log.flush()

    def flush(self):
        self.terminal.flush()
        self.log.flush()

    def close(self):
        self.log.close()


if __name__ == '__main__':
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    log_path = f"../data/results/log_{timestamp}.txt"
    sys.stdout = Logger(log_path)
    print(f"Log guardado en: {log_path}")

    os.environ["TRANSFORMERS_VERBOSITY"] = "error"
    if inicializar_app():
        test_prompts()
    else:
        print("ERROR en la inicialización. Verifica la configuración.")

    sys.stdout.close()
    sys.stdout = sys.stdout.terminal
