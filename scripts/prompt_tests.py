import sys
from datetime import datetime

from flask import Flask
import os
import torch
from transformers import AutoModelForCausalLM, AutoTokenizer
import json
import csv
import re

app = Flask(__name__)

MODEL = None
TOKENIZER = None
TRAIN_DATA = None
STATS = None

def extract_features(smiles):
    return {
        "length": len(smiles),
        "num_rings": sum(c.isdigit() for c in smiles),
        "num_branches": smiles.count("("),
        "double_bonds": smiles.count("="),
        "triple_bonds": smiles.count("#"),
        "aromatic_atoms": sum(smiles.count(c) for c in "cnosp"),
        "num_N": smiles.count("N") + smiles.count("n"),
        "num_O": smiles.count("O") + smiles.count("o"),
        "num_S": smiles.count("S") + smiles.count("s"),
        "num_P": smiles.count("P"),
        "num_F": smiles.count('F'),
        "num_Cl": smiles.count('Cl'),
        "num_Br": smiles.count('Br'),
        "stereocenters": smiles.count("@"),
        "net_charge": smiles.count("+") - smiles.count("-"),
    }


def read_csv(filepath):
    data = []
    with open(filepath, 'r', encoding='utf-8') as f:
        reader = csv.DictReader(f)
        for row in reader:
            try:
                data.append({
                    'smiles': row['smiles'].strip(),
                    'adduct': row['Adduct'].strip(),
                    'ccs': float(row['CCS_AVG'])
                })
            except (ValueError, KeyError):
                continue
    return data


def analyze_data(data):
    ccs_values = [d['ccs'] for d in data]
    stats = {
        'ccs_min': min(ccs_values),
        'ccs_max': max(ccs_values),
        'ccs_avg': sum(ccs_values) / len(ccs_values)
    }
    return stats


def normalize_adduct(adduct):
    # Strips the trailing charge character to unify adduct formats.
    return adduct.rstrip('+-').strip()

# Selects N structurally similar molecules from the dataset for the given input.
def select_examples(data, input_smiles, input_adduct, n):

    # Step 1: Compute descriptors for the input compound.
    input_features = extract_features(input_smiles)
    adduct_norm = normalize_adduct(input_adduct)

    # Step 2: Filter the training set by adduct.
    filtered_data = [d for d in data if normalize_adduct(d['adduct']) == adduct_norm]
    # Fallback: if not enough examples for the same adduct, use the full dataset.
    if len(filtered_data) < n:
        print(f"Warning: only {len(filtered_data)} examples for adduct {adduct_norm}, using full dataset")
        filtered_data = data

    similarities = []
    for d in filtered_data:
        features = extract_features(d['smiles'])

        # Step 3: Compute a similarity score for each structural descriptor.
        #  - For each molecule in the filtered set
        #  - Compare its descriptors against those of the input compound
        #  - Produce a score between 0 and 1
        sim_length    = 1 / (1 + abs(features['length']        - input_features['length']) / 10)
        sim_rings     = 1 / (1 + abs(features['num_rings']     - input_features['num_rings']))
        sim_aromatic  = 1 / (1 + abs(features['aromatic_atoms']- input_features['aromatic_atoms']) / 3)

        # Step 4: Combine the three scores with weights (structure > length > aromaticity).
        similarity = 0.50 * sim_rings + 0.30 * sim_length + 0.20 * sim_aromatic
        similarities.append((similarity, d))

    # Step 5: Sort by score and return the top n.
    similarities.sort(reverse=True, key=lambda x: x[0])
    return [d for _, d in similarities[:n]]


def search_dataset(smiles, adduct, dataset):
    adduct_norm = normalize_adduct(adduct)
    for row in dataset:
        if row["smiles"] == smiles and normalize_adduct(row["Adduct"]) == adduct_norm:
            return row["ccs"]
    return None


ADDUCT_INFO = {
    '[M+H]':     {'charge': 1,  'mass_add': 1.007,  'effect': 'protonated, baseline reference'},
    '[M+Na]':    {'charge': 1,  'mass_add': 22.989, 'effect': 'sodium adduct, ~3-5 Å² larger than [M+H]+'},
    '[M-H]':     {'charge': -1, 'mass_add': -1.007, 'effect': 'deprotonated, ~2-5 Å² smaller than [M+H]+'},
}

# Iteration 1: Builds a simple prompt using only the target compound information.
def build_prompt_simple(smiles, adduct):
    adduct_norm = normalize_adduct(adduct)
    info = ADDUCT_INFO.get(adduct_norm, {
        'charge': 1, 'mass_add': 0, 'effect': 'unknown adduct type'
    })

    prompt = f"""You are an expert in ion mobility mass spectrometry.
Your task is to predict the Collision Cross Section (CCS, in Å²) of the molecule from its SMILES structure and adduct.

To estimate CCS, follow these steps:
  1. Read the SMILES and identify rings, chains, branches, and functional groups.
  2. Decide if the structure is compact (mostly fused rings, rigid) or extended (chains, flexible).
  3. Apply the adduct correction described in the molecule line below.
  4. Output a single CCS value, typically in the range 100-300 Å² for small molecules.

Molecule:
   SMILES: {smiles}
   Adduct: {adduct_norm} ({info['effect']})

You MUST end your answer with the line "Final CCS: <number>", do not leave the number blank.
<think>
Let me apply the 4 steps to estimate the CCS.

Step 1 (size): """
    return prompt

# Iteration 2: Builds a prompt using expert metabolomics knowledge and the target compound information.
def build_prompt_expert(smiles, adduct):
    adduct_norm = normalize_adduct(adduct)
    info = ADDUCT_INFO.get(adduct_norm, {
        'charge': 1, 'mass_add': 0, 'effect': 'unknown adduct type'
    })

    prompt = f"""You are an expert in ion mobility mass spectrometry.
Your task is to predict the Collision Cross Section (CCS, in Å²) of the molecule from its SMILES structure and adduct.

CCS reflects the rotationally-averaged area a molecule presents while colliding with buffer gas in an ion mobility cell. It is determined primarily by the three-dimensional shape of the ion, not by its mass.

Expert knowledge - structural factors that determine CCS:

1. Molecular size and rigidity:
   - Compact, rigid structures (fused aromatic systems, polycyclic cores) yield LOWER CCS for their mass.
   - Extended, flexible structures (long aliphatic chains, single-bond rotations) yield HIGHER CCS for their mass.

2. Branching and substitution:
   - Branched molecules pack more compactly than linear isomers of the same mass → lower CCS.
   - Bulky substituents (long side chains, multiple rings off a central atom) increase CCS.

3. Heteroatoms and functional groups:
   - Heteroatoms (N, O, S) and halogens add mass but contribute relatively little to molecular volume.
   - Polar groups (-OH, -NH2, -COOH) can promote intramolecular interactions, slightly reducing CCS.

4. Adduct effect:
   - The adduct alters charge state and ion geometry.
   - Use the adduct effect described in the molecule line below as a small final adjustment.

To estimate CCS, follow these steps:
  1. Read the SMILES and identify rings, chains, branches, and functional groups.
  2. Decide if the structure is compact (mostly fused rings, rigid) or extended (chains, flexible).
  3. Apply the adduct correction described in the molecule line below.
  4. Output a single CCS value, typically in the range 100-300 Å² for small molecules.

Molecule:
  SMILES: {smiles}
  Adduct: {adduct_norm} ({info['effect']})

You MUST end your answer with the line "Final CCS: <number>", do not leave the number blank.
<think>
Let me apply the 4 steps to estimate the CCS.

Step 1 (structure): """
    return prompt

# Iteration 3: Builds a prompt using database examples and the target compound information.
def build_prompt_examples(smiles, adduct, TRAIN_DATA, n):
    adduct_norm = normalize_adduct(adduct)
    info = ADDUCT_INFO.get(adduct_norm, {
        'charge': 1, 'mass_add': 0, 'effect': 'unknown adduct type'
    })

    # Select N structurally similar molecules from the dataset
    examples = select_examples(TRAIN_DATA, smiles, adduct, n)

    # Build the examples block (same format as the target: SMILES + Adduct + CCS)
    examples_text = ""
    for i, ex in enumerate(examples, 1):
        ex_adduct_norm = normalize_adduct(ex['adduct'])
        examples_text += (
            f"  Example {i}:\n"
            f"    SMILES: {ex['smiles']}\n"
            f"    Adduct: {ex_adduct_norm}\n"
            f"    CCS: {ex['ccs']:.2f} Å²\n\n"
        )

    prompt = f"""You are an expert in ion mobility mass spectrometry.
Your task is to predict the Collision Cross Section (CCS, in Å²) of the molecule from its SMILES structure and adduct.

Below are {n} reference molecules with experimentally measured CCS values, selected for their structural similarity to the target. Use them to identify patterns and infer the CCS of the target molecule.

Reference molecules:

{examples_text}To estimate CCS, follow these steps:
  1. Read the SMILES and identify rings, chains, branches, and functional groups.
  2. Decide if the structure is compact (mostly fused rings, rigid) or extended (chains, flexible).
  3. Compare with the reference molecules above. Identify the most similar ones and use their CCS values as anchors.
  4. Apply the adduct correction described in the molecule line below.
  5. Output a single CCS value, typically in the range 100-300 Å² for small molecules.

Molecule:
   SMILES: {smiles}
   Adduct: {adduct_norm} ({info['effect']})

You MUST end your answer with the line "Final CCS: <number>", do not leave the number blank.
<think>
Let me apply the 5 steps to estimate the CCS.

Step 1 (size): """
    return prompt

# Iteration 4: Builds a prompt using expert knowledge, database examples, and the target compound information.
def build_prompt_combined(smiles, adduct, TRAIN_DATA, n):
    adduct_norm = normalize_adduct(adduct)
    info = ADDUCT_INFO.get(adduct_norm, {
        'charge': 1, 'mass_add': 0, 'effect': 'unknown adduct type'
    })

    # Select N structurally similar molecules from the dataset
    examples = select_examples(TRAIN_DATA, smiles, adduct, n)

    # Build the examples block (same format as the target: SMILES + Adduct + CCS)
    examples_text = ""
    for i, ex in enumerate(examples, 1):
        ex_adduct_norm = normalize_adduct(ex['adduct'])
        examples_text += (
            f"  Example {i}:\n"
            f"    SMILES: {ex['smiles']}\n"
            f"    Adduct: {ex_adduct_norm}\n"
            f"    CCS: {ex['ccs']:.2f} Å²\n\n"
        )

    prompt = f"""You are an expert in ion mobility mass spectrometry.
Your task is to predict the Collision Cross Section (CCS, in Å²) of the molecule from its SMILES structure and adduct.

CCS reflects the rotationally-averaged area a molecule presents while colliding with buffer gas in an ion mobility cell. It is determined primarily by the three-dimensional shape of the ion, not by its mass.

Expert knowledge - structural factors that determine CCS:

1. Molecular size and rigidity:
   - Compact, rigid structures (fused aromatic systems, polycyclic cores) yield LOWER CCS for their mass.
   - Extended, flexible structures (long aliphatic chains, single-bond rotations) yield HIGHER CCS for their mass.

2. Branching and substitution:
   - Branched molecules pack more compactly than linear isomers of the same mass → lower CCS.
   - Bulky substituents (long side chains, multiple rings off a central atom) increase CCS.

3. Heteroatoms and functional groups:
   - Heteroatoms (N, O, S) and halogens add mass but contribute relatively little to molecular volume.
   - Polar groups (-OH, -NH2, -COOH) can promote intramolecular interactions, slightly reducing CCS.

4. Adduct effect:
   - The adduct alters charge state and ion geometry.
   - Use the adduct effect described in the molecule line below as a small final adjustment.

Below are {n} reference molecules with experimentally measured CCS values, selected for their structural similarity to the target. Use them to identify patterns and infer the CCS of the target molecule.

Reference molecules:

{examples_text}To estimate CCS, follow these steps:
  1. Read the SMILES and identify rings, chains, branches, and functional groups.
  2. Decide if the structure is compact (mostly fused rings, rigid) or extended (chains, flexible).
  3. Compare with the reference molecules above. Identify the most similar ones and use their CCS values as anchors.
  4. Apply the adduct correction described in the molecule line below.
  5. Output a single CCS value, typically in the range 100-300 Å² for small molecules.

Molecule:
   SMILES: {smiles}
   Adduct: {adduct_norm} ({info['effect']})

You MUST end your answer with the line "Final CCS: <number>", do not leave the number blank.
<think>
Let me apply the 5 steps to estimate the CCS.

Step 1 (size): """
    return prompt


# Extracts the predicted CCS from the model's raw response.
def parse_response(raw_response):

    # Initial cleanup
    # Remove markdown bold/italic that interferes with regex
    text = raw_response.replace('**', '').replace('*', '')
    # Remove common LaTeX artifacts: $...$, $$...$$, \[ \], \( \)
    text = re.sub(r'\$+', '', text)
    text = re.sub(r'\\[\[\]\(\)]', '', text)
    text = text.strip()

    # Check the <think> block
    if '<think>' in text and '</think>' not in text:
        return {'predicted_ccs': None, 'fallback': True, 'source': 'think_unclosed'}
    if '</think>' not in text:
        return {'predicted_ccs': None, 'fallback': True, 'source': 'no_think_tag'}

    # From here on, work only with the post-think section
    post_text = text.split('</think>')[-1].strip()

    def accept(value, source):
        if 50 <= value <= 500:  # reasonable CCS range
            return {'predicted_ccs': round(value, 2), 'fallback': False, 'source': source}
        return None

    NUM = r'([0-9]+(?:\.[0-9]+)?)'

    # STRATEGY 1: "Final CCS: <number>" in variants (with/without \boxed{})
    final_patterns = [
        rf'Final\s+CCS\s*[:=]?\s*\\?boxed\{{?\s*{NUM}\s*\}}?',  # "Final CCS: \boxed{200}" or "Final CCS: 200"
        rf'Final\s+CCS\s*[:=]\s*{NUM}',                          # "Final CCS: 200" or "Final CCS = 200"
        rf'Final\s+CCS\s+{NUM}'                                   # "Final CCS 200" (no separator)
    ]
    for pattern in final_patterns:
        match = re.search(pattern, post_text, re.IGNORECASE)
        if match:
            value = float(match.group(1))
            result = accept(value, 'final_ccs_tag')
            if result:
                return result

    # STRATEGY 2: \boxed{<number>} without a preceding "Final CCS:"
    match = re.search(rf'\\?boxed\{{?\s*{NUM}\s*\}}?', post_text)
    if match:
        value = float(match.group(1))
        result = accept(value, 'boxed_tag')
        if result:
            return result

    # STRATEGY 3: last plausible number in the CCS range
    matches = re.findall(r'\b([0-9]{2,3}(?:\.[0-9]+)?)\b', post_text)
    plausible = [float(m) for m in matches if 50 <= float(m) <= 500]
    if plausible:
        return {
            'predicted_ccs': round(plausible[-1], 2),
            'fallback': False,
            'source': 'last_plausible',
        }

    return {'predicted_ccs': None, 'fallback': True, 'source': 'no_number_found'}

# Classifies whether the prediction is valid or a degenerate value.
def classify_prediction(ccs_pred, examples, stats):

    reference_ccs = {round(ex['ccs'], 2) for ex in examples}
    avg = round(stats['ccs_avg'])

    if abs(ccs_pred - avg) <= 0.5:
        return 'dataset_avg', False

    if any(abs(ccs_pred - ref) <= 0.02 for ref in reference_ccs):
        return 'exact_copy', False

    return 'interpolated', True


def predict_ccs(model, tokenizer, prompt, stats, examples):

    inputs = tokenizer(prompt, return_tensors="pt", truncation=True, max_length=5000)

    with torch.inference_mode():
        outputs = model.generate(
            **inputs,
            do_sample = False,           # 1.5B -> False ; Bigger models (7B) -> True
            temperature = None,          # 1.5B -> None  ; Bigger models (7B) -> 0.3
            top_p = None,                # 1.5B -> None  ; Bigger models (7B) -> 0.9
            repetition_penalty = 1.15,   # Penalizes repeated tokens to prevent looping during long reasoning traces.
            max_new_tokens = 10000,      # Extra margin for the <think> block
            pad_token_id = tokenizer.eos_token_id
        )

    n_prompt_tokens = inputs['input_ids'].shape[1]
    n_total_tokens = outputs.shape[1]
    n_generated_tokens = n_total_tokens - n_prompt_tokens

    print(f" - Prompt tokens: {n_prompt_tokens}")
    print(f" - Total output tokens: {n_total_tokens}")
    print(f" - Generated tokens: {n_generated_tokens}")

    full_response = tokenizer.decode(outputs[0], skip_special_tokens=True)
    prompt_text = tokenizer.decode(inputs['input_ids'][0], skip_special_tokens=True)
    response = full_response[len(prompt_text):].strip()

    print(" - FULL RESPONSE: " + json.dumps(full_response))
    print(" - RAW RESPONSE: " + json.dumps(response))

    result = parse_response(response)

    # Fallback on parse failure
    if result['fallback']:
        result['predicted_ccs'] = 0.0
        result['pred_type'] = 'heuristic_fallback'
        result['reasoning'] = "Heuristic fallback: parser found no valid number"
        return result

    # Classify the prediction
    pred_type, is_valid = classify_prediction(result['predicted_ccs'], examples, stats)
    result['predicted_ccs_raw'] = result['predicted_ccs']
    result['pred_type'] = pred_type

    if is_valid:
        result['reasoning'] = f"Model interpolation ({result['predicted_ccs']})"
    else:
        result['pred_type'] = pred_type  # 'exact_copy' or 'dataset_avg', no heuristic fallback

        if pred_type == 'exact_copy':
            result['reasoning'] = f"Model output accepted (reference copy: {result['predicted_ccs_raw']})"
        else:
            result['reasoning'] = f"Model output accepted (dataset avg: {result['predicted_ccs_raw']})"

    return result


def load_model():

    os.environ['CUDA_VISIBLE_DEVICES'] = ''
    model_path = r"D:\Modelos TFG\DeepSeek-R1-Distill-Qwen-1.5B"  # Local path
    print(f" - Path: {model_path}")

    tokenizer = AutoTokenizer.from_pretrained(model_path, trust_remote_code=True)
    if tokenizer.pad_token is None:
        tokenizer.pad_token = tokenizer.eos_token

    model = AutoModelForCausalLM.from_pretrained(
        model_path,
        device_map = "cpu",
        trust_remote_code = True,
        dtype = torch.float16,
        low_cpu_mem_usage = True
    )

    model.eval()  # Inference mode: disables dropout and gradients

    if hasattr(torch, 'compile'):
        print(" - Compiling model with torch.compile...")
        model = torch.compile(model, mode="reduce-overhead")
        print(" - Compilation complete")

    return model, tokenizer


def initialize_app():
    global MODEL, TOKENIZER, TRAIN_DATA, STATS

    print("=" * 70)
    print("Initializing CCS prediction application.")
    print("=" * 70)

    # Load data
    csv_path = r"../data/processed/train.csv"
    if not os.path.exists(csv_path):
        print(f"WARNING: File {csv_path} not found")
        print("Please make sure train.csv is at the correct path")
        return False

    print(f"Loading dataset: {csv_path}")
    TRAIN_DATA = read_csv(csv_path)
    print(f" - Dataset loaded: {len(TRAIN_DATA)} compounds")

    # Compute statistics
    print("Analyzing dataset statistics...")
    STATS = analyze_data(TRAIN_DATA)
    print(f" - CCS range: {STATS['ccs_min']:.1f} - {STATS['ccs_max']:.1f} Å²")

    # Load model
    print("Loading DeepSeek model...")
    MODEL, TOKENIZER = load_model()
    print(" - Model loaded and ready")
    print(" - Application ready to receive requests")
    print("=" * 70)
    return True


from openpyxl import Workbook, load_workbook

def initialize_excel(filepath, variants):
    """Creates the Excel file with headers if it does not exist."""
    if os.path.exists(filepath):
        return
    wb = Workbook()
    ws = wb.active
    ws.title = "CCS Predictions"
    headers = ["", "SMILES", "Adduct"]
    headers.extend(variants)
    ws.append(headers)
    wb.save(filepath)
    print(f" - Excel created: {filepath}")

def save_row_excel(filepath, row):
    """Appends a row to the Excel file and saves immediately."""
    wb = load_workbook(filepath)
    ws = wb.active
    ws.append(row)
    wb.save(filepath)

def update_cell_excel(filepath, row_idx, col_idx, value):
    """Updates a specific cell and saves immediately."""
    wb = load_workbook(filepath)
    ws = wb.active
    ws.cell(row=row_idx, column=col_idx, value=value)
    wb.save(filepath)


def test_prompts():

    global MODEL, TOKENIZER, TRAIN_DATA, STATS

    # Total of 390 compounds extracted from test.csv
    test_cases = [

        # [M+H]+ cases (130 compounds)
        {"smiles": "O=C([C@@H](NS(=O)(=O)c1ccc(cc1)Cl)Cc1c[nH]c2c1cccc2)NC1CCCC1", "adduct": "[M+H]+"},
        {"smiles": "COc1cccc(c1)[C@@H]1N(Cc2ccc(cc2)F)C(=O)c2c([C@@H]1C(=O)O)cccc2", "adduct": "[M+H]+"},
        {"smiles": "CCn1c(N)c(Br)c(=O)[nH]c1=O", "adduct": "[M+H]+"},
        {"smiles": "CC(=O)c1sc(nc1C)NCc1ccccc1", "adduct": "[M+H]+"},
        {"smiles": "Cc1ccc(cc1)n1sc(=N)nc1c1ccccc1", "adduct": "[M+H]+"},
        {"smiles": "C=CCNCCCOc1c(C)cc(cc1Cl)Cl", "adduct": "[M+H]+"},
        {"smiles": "O=C(NC1(CCCC1)C(=O)O)Cc1c[nH]c2c1cccc2", "adduct": "[M+H]+"},
        {"smiles": "ON=Cc1cn(nc1c1ccncc1)c1ccccc1", "adduct": "[M+H]+"},
        {"smiles": "Cc1sc2c(c1C)c(ncn2)Sc1nn2c(n1)nccc2", "adduct": "[M+H]+"},
        {"smiles": "Cc1ccc2c(c1)oc(=O)cc2COC(=O)c1cccs1", "adduct": "[M+H]+"},
        {"smiles": "O=C1NC(C(=O)N1CC(=O)N1CCCC1)Cc1c[nH]c2c1cccc2", "adduct": "[M+H]+"},
        {"smiles": "NC(=O)C1Cc2ccccc2CN1S(=O)(=O)c1c(C)noc1C", "adduct": "[M+H]+"},
        {"smiles": "CCN1CCN(CC1)C(=O)CCC1CCCCC1", "adduct": "[M+H]+"},
        {"smiles": "CCOc1ccccc1NC(=O)CNc1ccccc1C", "adduct": "[M+H]+"},
        {"smiles": "CC1CCN(CC1)CN1C(=O)NC2(C1=O)CCCc1c2cccc1", "adduct": "[M+H]+"},
        {"smiles": "O=C(C1CCN(CC1)c1nc2c(o1)cccc2)NCCc1ccccn1", "adduct": "[M+H]+"},
        {"smiles": "CN(C(=O)C(c1ccccc1)NCCc1c[nH]c2c1cccc2)C", "adduct": "[M+H]+"},
        {"smiles": "CCN(S(=O)(=O)c1ccccc1C#N)Cc1ccc2c(c1)OCO2", "adduct": "[M+H]+"},
        {"smiles": "O=C(N1CCc2c1cccc2)Cn1c(=O)cnc2c1cccc2", "adduct": "[M+H]+"},
        {"smiles": "CC1CC1C(=O)NCC(=O)NCc1ccc(cc1)C(F)(F)F", "adduct": "[M+H]+"},
        {"smiles": "Cc1ccc(cc1)C1(C)NC(=O)N(C1=O)Cc1ccc2c(c1)OCO2", "adduct": "[M+H]+"},
        {"smiles": "Cc1ccc(c(c1)C)C(=O)COc1ccc(cc1)c1nnco1", "adduct": "[M+H]+"},
        {"smiles": "CC(CC(C(=O)O)NC(=O)[C@H](C(C)C)NC(=O)c1ccco1)C", "adduct": "[M+H]+"},
        {"smiles": "CC(=O)N1N=C(CC1c1ccccc1Cl)c1ccccc1NS(=O)(=O)C", "adduct": "[M+H]+"},
        {"smiles": "O=C(N(CC1COc2c(O1)cccc2)C)Cc1ccc2c(c1)CCC2", "adduct": "[M+H]+"},
        {"smiles": "CC(CNC(=O)NC(=O)CN1C(=O)NC2(C1=O)CCC(CC2)C)C", "adduct": "[M+H]+"},
        {"smiles": "CCCN(C(=O)c1ccc(cc1)n1cccn1)Cc1nnc(o1)c1ccco1", "adduct": "[M+H]+"},
        {"smiles": "CCc1nc2c(n1CC(=O)Nc1ccccc1c1ccccc1)cccc2", "adduct": "[M+H]+"},
        {"smiles": "Fc1ccc2c(c1)c(CCC(=O)N1CCOCC1)c([nH]2)c1ccc(cc1)F", "adduct": "[M+H]+"},
        {"smiles": "CCc1ccccc1NC(=O)c1ccccc1c1ccc(cc1)Cn1cccn1", "adduct": "[M+H]+"},
        {"smiles": "Brc1ccc(cc1)S(=O)(=O)NCCC(=O)N(Cc1cccs1)C", "adduct": "[M+H]+"},
        {"smiles": "Clc1ccc(cc1)c1cnc(o1)CN1CCN(CC1)c1ccccn1", "adduct": "[M+H]+"},
        {"smiles": "CC(CC(C(=O)NCCNC(=O)C)NC(=O)c1ccc(cc1)Cl)C", "adduct": "[M+H]+"},
        {"smiles": "CCOC(=O)c1ccc(cc1)OCC(CN1CC(C)OC(C1)C)O", "adduct": "[M+H]+"},
        {"smiles": "O=C(Nc1c(C)cccc1C)CN1CCN(CC1)C(=O)Cc1ccccc1", "adduct": "[M+H]+"},
        {"smiles": "Cn1ccnc1Sc1ccc(cc1)NS(=O)(=O)c1ccccc1C(F)(F)F", "adduct": "[M+H]+"},
        {"smiles": "O=C(c1cncn1c1ccccc1)N1CCCC(C1)c1nc2c(s1)cccc2", "adduct": "[M+H]+"},
        {"smiles": "CCCn1c(NC(=O)c2ccccc2SCc2c(C)onc2C)nc2c1cccc2", "adduct": "[M+H]+"},
        {"smiles": "COc1cccc(c1)C(N1CCCC1)CNC(=O)c1ccc(cc1)Br", "adduct": "[M+H]+"},
        {"smiles": "CCn1c(SCC(=O)N2CCN(CC2)c2ncccn2)nnc1C1CC1", "adduct": "[M+H]+"},
        {"smiles": "Clc1ccccc1c1nc2ccccc2c(c1)C(=O)OCc1c(C)noc1C", "adduct": "[M+H]+"},
        {"smiles": "O=C(CSc1nnc(n1Cc1ccccc1)C1CC1)NCC1CCCO1", "adduct": "[M+H]+"},
        {"smiles": "COc1cc(OC)ccc1C1CCCN1CC(=O)Nc1ccnn1C1CCCC1", "adduct": "[M+H]+"},
        {"smiles": "CC1CN(CCCNC(=O)COc2ccccc2c2ccccc2)CC(C1)C", "adduct": "[M+H]+"},
        {"smiles": "O=C(c1cccnc1Sc1ccc2c(c1)OCCO2)N1CCN(CC1)c1ccccn1", "adduct": "[M+H]+"},
        {"smiles": "O=C(N1CCC(CC1)c1nc2c(o1)cccc2)C=Cc1ccc2c(c1)OCO2", "adduct": "[M+H]+"},
        {"smiles": "Cc1cccc(c1)n1c(SCc2nnc(o2)c2ccco2)nnc1N1CCOCC1", "adduct": "[M+H]+"},
        {"smiles": "COc1ccccc1c1nnc(n1CC)SCC(=O)N1CC(C)CC(C1)C", "adduct": "[M+H]+"},
        {"smiles": "COc1ccccc1c1nn(cc1C(=O)OC1CC(OC1=O)C)Cc1ccccc1", "adduct": "[M+H]+"},
        {"smiles": "COc1cccc(c1C(=O)N(c1ccccc1)C1CCN(CC1)Cc1ccccc1)OC", "adduct": "[M+H]+"},
        {"smiles": "CCOc1ccc(cc1)c1n[nH]c(=S)n1CC(=O)NCC1(CC1)c1ccccc1", "adduct": "[M+H]+"},
        {"smiles": "O=C(CSc1nnc(n1Cc1ccco1)c1cccc(c1)C)NCC1COc2c(O1)cccc2", "adduct": "[M+H]+"},
        {"smiles": "O=C(N(c1sc2c(n1)CCCC2)Cc1ccccc1)Cn1nnc(n1)c1ccccc1", "adduct": "[M+H]+"},
        {"smiles": "O=C(N1CCC(CC1)c1ncc([nH]1)c1ccc(c(c1)F)F)NCC(c1ccco1)N1CCCC1", "adduct": "[M+H]+"},
        {"smiles": "O=C(Nc1ccc(cc1)N1CCOCC1)CN1CCCC1c1ccc(cc1)Cl", "adduct": "[M+H]+"},
        {"smiles": "COc1ccc(c(c1)C1CCCN1C(=O)CNC(c1cccs1)c1ccc(cc1)F)OC", "adduct": "[M+H]+"},
        {"smiles": "O=C(C1CC(=O)N(C1)Cc1ccco1)NCc1ccc(cc1)Oc1ccc(cc1)F", "adduct": "[M+H]+"},
        {"smiles": "Cc1ccc(cc1C(=O)N1CCCC(C1)c1nc2c(s1)cccc2)S(=O)(=O)N1CCCC1", "adduct": "[M+H]+"},
        {"smiles": "O=C(Nc1ccccc1C(=O)Nc1ccccc1)CNC(c1cccs1)c1ccccc1", "adduct": "[M+H]+"},
        {"smiles": "O=C1CCc2c(N1)ccc(c2)S(=O)(=O)NCC(c1ccccc1)(c1ccccc1)C", "adduct": "[M+H]+"},
        {"smiles": "COc1ccc(cc1C1CCCN1C(=O)CN(CC(=O)Nc1c(C)cc(cc1C)C)C)OC", "adduct": "[M+H]+"},
        {"smiles": "CC1CCCCN1S(=O)(=O)c1ccc(cc1)C(=O)OCc1noc(c1)c1ccco1", "adduct": "[M+H]+"},
        {"smiles": "COc1cc(cc(c1OCc1ccccc1)OC)C(=O)N1CCN(CC1c1ccccc1)C", "adduct": "[M+H]+"},
        {"smiles": "Fc1ccccc1Cn1c(CN(Cc2ccccc2)Cc2ccccc2)nc2c1c(=O)n(c(=O)n2C)C", "adduct": "[M+H]+"},
        {"smiles": "COc1cc(ccc1OC)c1nnn(n1)CC(=O)NC(c1ccccc1)c1ccccc1", "adduct": "[M+H]+"},
        {"smiles": "COCCn1c(=O)[nH]c(=O)c(c1N)N(C(=O)c1cn(nc1c1ccccc1)c1ccccc1)Cc1ccccc1", "adduct": "[M+H]+"},
        {"smiles": "COc1ccccc1N(S(=O)(=O)c1ccc(cc1)Cl)Cc1onc(n1)c1ccsc1", "adduct": "[M+H]+"},
        {"smiles": "Cc1onc(c1)C(=O)N(C1CC1)Cc1nnc(o1)c1ccccc1Cl", "adduct": "[M+H]+"},
        {"smiles": "NCCC(c1ccccc1)c1ccccc1", "adduct": "[M+H]+"},
        {"smiles": "O=C(N1N=C(CC1c1ccco1)c1cccs1)Cn1nnc(n1)c1ccccc1C", "adduct": "[M+H]+"},
        {"smiles": "c1ccc(cc1)Cn1nc(c(c1)CNC1CC1)c1cccnc1", "adduct": "[M+H]+"},
        {"smiles": "O=C(NCC1(CC1)c1ccccc1)CCCn1cnc2c(c1=O)cccc2", "adduct": "[M+H]+"},
        {"smiles": "COc1ccc(cc1)c1cc(ccc1OC)NC(=O)Cc1c(C)noc1C", "adduct": "[M+H]+"},
        {"smiles": "COc1cccc(c1)C(=O)N1CCN(CC1)Cc1csc(n1)C1CCCCC1", "adduct": "[M+H]+"},
        {"smiles": "COC(=O)C1=C(C)NC(=C(C1C(=O)OCC(=O)Nc1onc(c1)C(C)(C)C)C(=O)OC)C", "adduct": "[M+H]+"},
        {"smiles": "CCOc1ccc(c(c1)O)[C@H]1N(Cc2ccccc2)C(=O)c2c([C@H]1C(=O)O)cccc2", "adduct": "[M+H]+"},
        {"smiles": "CC1CCCCN1C(=O)c1ccc(cc1)S(=O)(=O)N1CCCCC1C", "adduct": "[M+H]+"},
        {"smiles": "NC(=O)CSc1nnc(n1c1ccccc1Cl)N1CCCC1", "adduct": "[M+H]+"},
        {"smiles": "O=C(N1CCCC1c1ccc2c(c1)OCCCO2)CCc1c(C)nc2n(c1C)nc(n2)C(F)(F)F", "adduct": "[M+H]+"},
        {"smiles": "CCOC(=O)c1c(C)[nH]c(c1C)C(=O)OCc1csc(n1)COc1ccc(cc1)F", "adduct": "[M+H]+"},
        {"smiles": "CN(C(=O)c1cc(ccc1N1CCCC1)S(=O)(=O)N(C)C)Cc1c(C)nn(c1C)c1ccccc1", "adduct": "[M+H]+"},
        {"smiles": "O=C(NC(c1cccs1)c1ccccc1)CSc1nccn1Cc1ccccc1", "adduct": "[M+H]+"},
        {"smiles": "O=C(c1ccc(cc1)n1cnc2c1cccc2)NCC(c1ccco1)N1CCCC1", "adduct": "[M+H]+"},
        {"smiles": "COCCOC(=O)C1=C(C)N(C(=O)NC1c1ccc(cc1)C#N)C(COC)C", "adduct": "[M+H]+"},
        {"smiles": "N#Cc1ccccc1c1ccc(cc1)COC(=O)[C@H](Cc1c[nH]c2c1cccc2)NC(=O)C", "adduct": "[M+H]+"},
        {"smiles": "Brc1ccc(o1)C(=O)Nc1ccccc1OC(F)(F)F", "adduct": "[M+H]+"},
        {"smiles": "O=C(C1CCCN1S(=O)(=O)c1ccc(cc1)OC(F)(F)F)N(C)C", "adduct": "[M+H]+"},
        {"smiles": "OC(=O)c1ccc(cc1)OCC(F)(F)F", "adduct": "[M+H]+"},
        {"smiles": "N#Cc1c(C)c(n(c1NC(=O)CN1CCCC1c1cccs1)CC1CCCO1)C", "adduct": "[M+H]+"},
        {"smiles": "CC(NC(=O)CCn1cnc2c(c1=O)c1CCCCc1s2)C", "adduct": "[M+H]+"},
        {"smiles": "COc1cc(cc(c1OC)OC)c1nnc(o1)SCC(=O)c1cc(n(c1C)CC1CCCO1)C", "adduct": "[M+H]+"},
        {"smiles": "O=C(N1CCC(CC1)Cc1ccccc1)Cn1c(C)nc2c(c1=O)cccc2", "adduct": "[M+H]+"},
        {"smiles": "O=C(Cc1c(C)noc1C)NCC(c1c[nH]c2c1cccc2)c1ccccc1Cl", "adduct": "[M+H]+"},
        {"smiles": "COC(=O)c1c(C)[nH]c(c1C)C(=O)NCC12CC3CC(C2)CC(C1)C3", "adduct": "[M+H]+"},
        {"smiles": "CCOCCCNCC(=O)N1N=C(CC1c1ccc(cc1)OC)c1ccc2c(c1)cccc2", "adduct": "[M+H]+"},
        {"smiles": "O=C(N1CCCC1c1cccs1)Cn1nnn(c1=O)c1ccccc1", "adduct": "[M+H]+"},
        {"smiles": "CNC(=O)COC(=O)c1c(C)nn(c1C)c1ccccc1", "adduct": "[M+H]+"},
        {"smiles": "CCS(=O)(=O)Nc1cccc(c1)Oc1ccccc1", "adduct": "[M+H]+"},
        {"smiles": "Fc1ccc(c(c1)C(=O)Nc1nc2c(s1)cccc2)Br", "adduct": "[M+H]+"},
        {"smiles": "Clc1ccc(cc1)c1cnc(o1)CN1CCN(CC1)c1nccs1", "adduct": "[M+H]+"},
        {"smiles": "O=C(c1cc(nc2c1cccc2)c1ccncc1)NCc1ccc(cc1)S(=O)(=O)N1CCOCC1", "adduct": "[M+H]+"},
        {"smiles": "CCC1(NC(=O)N(C1=O)CC(=O)N1CCCCC1C)c1ccc(cc1)F", "adduct": "[M+H]+"},
        {"smiles": "CC1CC(C(=O)O1)Sc1nnc(n1c1ccccc1)c1ccc2c(c1)OCO2", "adduct": "[M+H]+"},
        {"smiles": "O=C(NCc1ccc(nc1)n1cncc1)NOCc1ccccc1", "adduct": "[M+H]+"},
        {"smiles": "O=C(C(c1ccccc1)Sc1nnc(o1)c1ccccc1F)N1CCCCC1", "adduct": "[M+H]+"},
        {"smiles": "COc1ccccc1N(S(=O)(=O)c1ccc(cc1)C(=O)Nc1ncc(s1)C)C", "adduct": "[M+H]+"},
        {"smiles": "CC(C(C(=O)NC1CC1)NS(=O)(=O)c1ccc2c(c1)OCCO2)C", "adduct": "[M+H]+"},
        {"smiles": "O=C(c1n[nH]c(c1)c1ccccc1)N1CCN(CC1)S(=O)(=O)c1ccc2c(c1)OCCO2", "adduct": "[M+H]+"},
        {"smiles": "OCC1CCCN(C1)CC(=O)NC(c1ccc2c(c1)OCCCO2)C(C)C", "adduct": "[M+H]+"},
        {"smiles": "CCCC1(NC(=O)N(C1=O)CC(=O)NC12CC3CC(C2)CC(C1)C3)c1ccccc1", "adduct": "[M+H]+"},
        {"smiles": "O=C(Nc1ccccc1C(=O)Nc1ccc2c(c1)OCO2)CSc1nnc(n1C1CC1)C1CC1", "adduct": "[M+H]+"},
        {"smiles": "O=C(c1ccc2c(c1)c(on2)c1ccccc1)NC(c1ccccc1)CN1CCOCC1", "adduct": "[M+H]+"},
        {"smiles": "O=C(c1ccc(=O)n(n1)C)NCC1(CCCCC1)N1CCCCC1", "adduct": "[M+H]+"},
        {"smiles": "c1ccc(cc1)CSc1nnc([nH]1)c1cccs1", "adduct": "[M+H]+"},
        {"smiles": "O=C(N1CCN(CC1)C(=O)c1ccc2c(c1)cccc2)CN1C(=O)NC2(C1=O)CCCC2", "adduct": "[M+H]+"},
        {"smiles": "CC(CC(N1CCOCC1)CNC(=O)C1CCCO1)C", "adduct": "[M+H]+"},
        {"smiles": "CCN(S(=O)(=O)c1cccc(c1)c1nnc(n1C)SCC(=O)c1ccc(cc1)F)CC", "adduct": "[M+H]+"},
        {"smiles": "O=C(c1ccc2c(c1)CCCC2)CSc1nnc(o1)c1ccccc1Br", "adduct": "[M+H]+"},
        {"smiles": "O=C(c1ccc(cc1)S(=O)(=O)N1CCCC1)Nc1ccnn1Cc1cccs1", "adduct": "[M+H]+"},
        {"smiles": "O=C(c1ccc(nc1)C(F)(F)F)OCc1nnnn1c1ccccc1", "adduct": "[M+H]+"},
        {"smiles": "CC([C@H](N1C(=O)c2c(C1=O)cccc2)C(=O)NCC(c1ccccc1)N(C)C)C", "adduct": "[M+H]+"},
        {"smiles": "CN(Cc1cc(=O)n2c(n1)scc2)CC1COc2c(O1)cccc2", "adduct": "[M+H]+"},
        {"smiles": "O=C(CC1Sc2ccccc2NC1=O)NCc1cccnc1", "adduct": "[M+H]+"},
        {"smiles": "COc1cc2CN(CCc2cc1OC)S(=O)(=O)c1ccccc1Cl", "adduct": "[M+H]+"},
        {"smiles": "Clc1ccc(cc1)OCc1nnc(o1)c1ccccc1", "adduct": "[M+H]+"},
        {"smiles": "O=C(C1CCCN(C1)C(=O)c1ccco1)Nc1ccccc1N1CCCC1=O", "adduct": "[M+H]+"},
        {"smiles": "O=C1N(Cc2nnc(o2)c2ccccc2Br)C(=O)C2C1CC=CC2", "adduct": "[M+H]+"},
        {"smiles": "O=C(N1CCN(CC1)CC(=O)N1CCCC1)CCc1cccs1", "adduct": "[M+H]+"},
        {"smiles": "O=C(C1CCCN1S(=O)(=O)c1c(C)noc1C)Nc1cccc(c1)S(=O)(=O)N(C)C", "adduct": "[M+H]+"},
        {"smiles": "CC(C(=O)N1CCOCC1)Oc1ccc(cc1)C(=O)c1ccccc1", "adduct": "[M+H]+"},

        # [M-H]- cases (130 compounds)
        {"smiles": "OC(=O)/C=C/c1ccc(cc1)OC(F)(F)F", "adduct": "[M-H]-"},
        {"smiles": "O=C1N[C@@H]2[C@H](N1)[C@@H](SC2)CCCCC(=O)N1CCC(CC1)C(=O)Nc1ccc2c(c1)OCO2", "adduct": "[M-H]-"},
        {"smiles": "Oc1cccc(c1)I", "adduct": "[M-H]-"},
        {"smiles": "CCc1cccc(c1)NC(=O)c1cscc1", "adduct": "[M-H]-"},
        {"smiles": "CCCNS(=O)(=O)c1ccc(cc1)C(C)C", "adduct": "[M-H]-"},
        {"smiles": "O=C(Nc1ccccc1C(C)C)CN1C(=O)CCC1=O", "adduct": "[M-H]-"},
        {"smiles": "COC1=CC(C=CC(O)=O)=CC(OC)=C1OC", "adduct": "[M-H]-"},
        {"smiles": "CC(N1CCOCC1)CNS(=O)(=O)c1ccccc1", "adduct": "[M-H]-"},
        {"smiles": "CCOc1ccc(cc1OC)C(=O)NCc1ccncc1", "adduct": "[M-H]-"},
        {"smiles": "O=C(c1cccc(c1O)C)NCc1ccc2c(c1)OCO2", "adduct": "[M-H]-"},
        {"smiles": "CCCCOC(=O)CN1C(=O)NC(C1=O)(CC)c1ccccc1", "adduct": "[M-H]-"},
        {"smiles": "O=C(c1c(C)noc1C)OC(c1nc2ccccc2c(=O)[nH]1)C", "adduct": "[M-H]-"},
        {"smiles": "O=C(C(N1CCCC(C1)C(=O)N)C)Nc1ccccc1C(F)(F)F", "adduct": "[M-H]-"},
        {"smiles": "Cc1cccc(c1)CN1C(=O)NC2(C1=O)CCc1c2cccc1", "adduct": "[M-H]-"},
        {"smiles": "O=C(Cn1cc(ccc1=O)C(F)(F)F)Nc1cnc2c(c1)cccc2", "adduct": "[M-H]-"},
        {"smiles": "Fc1ccc2c(c1)c(CCC(=O)N1CCOCC1)c([nH]2)c1ccc(cc1)F", "adduct": "[M-H]-"},
        {"smiles": "CC(N(C(=O)CN1C(=O)NC2(C1=O)CCCCCC2)C(C)C)C", "adduct": "[M-H]-"},
        {"smiles": "CC(=O)c1ccc(cc1)NC(=O)c1ccccc1c1ccccc1", "adduct": "[M-H]-"},
        {"smiles": "O=C(Nc1ccc(cc1)C)CN(S(=O)(=O)C)C1CCCCC1", "adduct": "[M-H]-"},
        {"smiles": "COc1ccc(cc1F)CSc1nnc([nH]1)c1ccccc1Br", "adduct": "[M-H]-"},
        {"smiles": "COc1ccccc1CC(=O)Nc1cccc(c1)S(=O)(=O)N(C)C", "adduct": "[M-H]-"},
        {"smiles": "O=C(C(c1ccccc1)Sc1nnnn1C1CCCC1)N1CCOCC1", "adduct": "[M-H]-"},
        {"smiles": "CN(C(=O)c1ccc(cc1)S(=O)(=O)N(C)C)Cc1ccccc1F", "adduct": "[M-H]-"},
        {"smiles": "CC(c1ccccc1)NC(=O)c1cc(nc2c1cccc2)C1CC1", "adduct": "[M-H]-"},
        {"smiles": "O=C(Nc1ccccc1Cc1ccccc1)CSc1nnnn1C1CC1", "adduct": "[M-H]-"},
        {"smiles": "N#CC1(CCCC1)NC(=O)CSc1nnc(s1)NCC1CCCO1", "adduct": "[M-H]-"},
        {"smiles": "CNc1snc(c1C(=O)N1CCN(CC1)S(=O)(=O)c1ccccc1F)C", "adduct": "[M-H]-"},
        {"smiles": "O=C(NC1CCCC(C1C)C)CSc1nnc2n1cccc2", "adduct": "[M-H]-"},
        {"smiles": "CCNS(=O)(=O)c1ccc(cc1)C(=O)NC1CCC(CC1)O", "adduct": "[M-H]-"},
        {"smiles": "O=C(c1cc(ccc1C)S(=O)(=O)N)NCCc1ccccc1C", "adduct": "[M-H]-"},
        {"smiles": "CC(=O)NC(C(=O)NCCc1ccccn1)Cc1c[nH]c2c1cccc2", "adduct": "[M-H]-"},
        {"smiles": "O=C(CN1CSCC1=O)Nc1ccc(cc1)SCc1cccnc1", "adduct": "[M-H]-"},
        {"smiles": "CCOC(=O)C(=Cc1ccc(c(c1)OC)OC)C1=Nn2c(SC1)nnc2C", "adduct": "[M-H]-"},
        {"smiles": "OCCN(S(=O)(=O)c1ccc(cc1)C(=O)NCc1ccccc1F)C", "adduct": "[M-H]-"},
        {"smiles": "COc1cc(ccc1OC(F)F)C(=O)OCN1C(=O)c2c(C1=O)cccc2", "adduct": "[M-H]-"},
        {"smiles": "O=C(c1cc(C)c(cc1C)C)CSc1nnc(n1C)c1cccs1", "adduct": "[M-H]-"},
        {"smiles": "COCCN1CCN(CC1)S(=O)(=O)c1ccc(cc1)C(=O)OCC", "adduct": "[M-H]-"},
        {"smiles": "COc1ccc(cc1)n1c(SCC(=O)Nc2ccccc2C)nc2c1cccc2", "adduct": "[M-H]-"},
        {"smiles": "Fc1ccc(cc1)CNC(=O)C1(CCN(CC1)C(=O)C)c1ccccc1", "adduct": "[M-H]-"},
        {"smiles": "COc1cc2c(cc1NC(=O)CCSc1ccccc1)oc1c2cccc1", "adduct": "[M-H]-"},
        {"smiles": "Cc1ccc(c(c1)S(=O)(=O)N1CCN(CC1)C(=O)Cc1csc(n1)Cc1ccccn1)C", "adduct": "[M-H]-"},
        {"smiles": "CC(C(C(=O)NCc1ccc(cc1)Cl)NC(=O)c1ccccc1Br)C", "adduct": "[M-H]-"},
        {"smiles": "CCC(=O)N(c1nnc(s1)SCc1nnc(o1)c1cc(oc1C)C)C1CC1", "adduct": "[M-H]-"},
        {"smiles": "O=C(Nc1cc(Cl)ccc1n1ncnc1)CN1C(=O)NC(C1=O)(C)c1ccc2c(c1)OCO2", "adduct": "[M-H]-"},
        {"smiles": "O=C(C(OC(=O)c1ccc(cc1)c1ccc(cc1)O)C)NCc1ccccc1", "adduct": "[M-H]-"},
        {"smiles": "O=C1NC(C(=O)N1Cc1onc(n1)c1cccs1)(C)c1ccc2c(c1)OCCO2", "adduct": "[M-H]-"},
        {"smiles": "CCCN(CC1=C(C(=O)OCC)C(NC(=O)N1)C)CC(=O)Nc1ccc(c(c1F)F)F", "adduct": "[M-H]-"},
        {"smiles": "NC(=O)C1Cc2ccccc2CN1C(=O)CCc1c([nH]c2c1cccc2)c1ccc(cc1)F", "adduct": "[M-H]-"},
        {"smiles": "O=C(C1CCN(CC1)c1nc2c(o1)cccc2)NCCc1ccccn1", "adduct": "[M-H]-"},
        {"smiles": "CC(Cn1c(SCc2nnc(o2)c2ccccc2Br)nnc1N1CCOCC1)C", "adduct": "[M-H]-"},
        {"smiles": "O=C(Nc1cc(ccc1Oc1ccc(cc1)F)S(=O)(=O)N1CCOCC1)Cc1cccs1", "adduct": "[M-H]-"},
        {"smiles": "COC(=O)c1cc(ccc1NS(=O)(=O)C1=Cc2c(CC1)cccc2)N1CCOCC1", "adduct": "[M-H]-"},
        {"smiles": "COc1ccc(cc1)N(S(=O)(=O)c1ccc2c(c1)OCCO2)Cc1noc(c1)c1ccco1", "adduct": "[M-H]-"},
        {"smiles": "CC(C(=O)N1CCc2c(C1)cccc2)Sc1nnc(n1c1ccccc1F)c1cccnc1", "adduct": "[M-H]-"},
        {"smiles": "CC(C12CC3CC(C2)CC(C1)C3)NC(=O)c1ccc(cc1)S(=O)(=O)N1CCCC1", "adduct": "[M-H]-"},
        {"smiles": "O=C(C1CC(=O)N(C1)Cc1ccco1)NCc1ccc(cc1)Cn1cncc1", "adduct": "[M-H]-"},
        {"smiles": "O=C(Nc1ccc(cc1)N1CCOCC1)COc1ccccc1C(=O)Nc1ccccc1", "adduct": "[M-H]-"},
        {"smiles": "Fc1cccc(c1)n1c(SCc2nnc(o2)c2ccccc2)nnc1N1CCOCC1", "adduct": "[M-H]-"},
        {"smiles": "COc1cc(ccc1OC)Cc1nnc(n1N)SCC(=O)Nc1ccc(cc1)c1ccccc1", "adduct": "[M-H]-"},
        {"smiles": "C=CCc1ccc(c(c1)OC)OCC(COc1cccc(c1)C(=O)NCc1ccccc1)O", "adduct": "[M-H]-"},
        {"smiles": "COc1ccc(c(c1)OC)NC(=O)CSc1oc(nc1S(=O)(=O)c1ccc(cc1)C)c1ccco1", "adduct": "[M-H]-"},
        {"smiles": "O=C(NC(c1cccc2c1cccc2)C)CN1C(=O)NC(C1=O)(Cc1ccccc1)Cc1ccccc1", "adduct": "[M-H]-"},
        {"smiles": "COc1ccc(cc1)C(c1ccccc1)NC(=O)Cn1cc(ccc1=O)S(=O)(=O)N1CCCC1", "adduct": "[M-H]-"},
        {"smiles": "COc1ccc(cc1)c1scc(n1)CN1CCN(CC1)C(=O)c1ccc2c(c1)OCO2", "adduct": "[M-H]-"},
        {"smiles": "O=C(c1cn(nc1c1ccccc1)c1ccccc1)NCC(c1cccs1)N1CCOCC1", "adduct": "[M-H]-"},
        {"smiles": "O=C(Cn1cnc2c(c1=O)c(cs2)c1ccc(cc1)C)NC1CCN(C1)Cc1ccccc1", "adduct": "[M-H]-"},
        {"smiles": "O=C(CN1C(=O)CSc2c1cccc2)Nc1ccc(cc1)C(=O)N", "adduct": "[M-H]-"},
        {"smiles": "Cc1cccc(c1)N(C(=O)Cn1nnc(n1)c1cccs1)Cc1ccccc1", "adduct": "[M-H]-"},
        {"smiles": "Ic1cc(ccc1C)NC(=O)c1c(C)noc1C", "adduct": "[M-H]-"},
        {"smiles": "CCOc1ccc(cc1)S(=O)(=O)N1CCNC(=O)C1", "adduct": "[M-H]-"},
        {"smiles": "CCCCN(c1c(=O)[nH]c(=O)n(c1N)Cc1ccccc1)C(=O)CSc1nnc(o1)Cc1csc(n1)C", "adduct": "[M-H]-"},
        {"smiles": "O=C(Cc1ccc(cc1)NC(=O)C)NCC(=O)Nc1ccc(c(c1F)F)F", "adduct": "[M-H]-"},
        {"smiles": "O=C(NCc1ccc(cc1)NC(=O)C1CC1)NCCc1ccccc1F", "adduct": "[M-H]-"},
        {"smiles": "O=C(N1CCc2c1cccc2)CSc1nnc([nH]1)c1ccccc1Br", "adduct": "[M-H]-"},
        {"smiles": "Fc1ccc(cc1)NC(=O)C(n1cnc2c1c(=O)n(C)c(=O)n2C)C", "adduct": "[M-H]-"},
        {"smiles": "COc1ccc(cc1S(=O)(=O)N(C)C)NS(=O)(=O)c1ccc2c(c1)CCCC2", "adduct": "[M-H]-"},
        {"smiles": "N#CC(=c1[nH]nc2n1CCCCC2)C(=O)CSc1ncccn1", "adduct": "[M-H]-"},
        {"smiles": "CCN(CCNC(=O)c1ccc(cc1)NC(=O)c1ccc(cc1)Br)CC", "adduct": "[M-H]-"},
        {"smiles": "CCCN(C(=O)c1ccc(cc1)NC(=O)c1cccs1)Cc1nnc(o1)c1ccco1", "adduct": "[M-H]-"},
        {"smiles": "O=C(NCC1COc2c(O1)cccc2)COc1ccc(cc1)c1ccccc1", "adduct": "[M-H]-"},
        {"smiles": "CSc1ccc(cc1)NC(=O)C12CC3CC(C1)CC(C2)(C3)O", "adduct": "[M-H]-"},
        {"smiles": "CSCCC(C(=O)NCc1cccnc1)NC(=O)c1ccccc1Br", "adduct": "[M-H]-"},
        {"smiles": "O=C(Nc1ccccc1)CCN1C(=O)CCC1=O", "adduct": "[M-H]-"},
        {"smiles": "COc1ccc(cc1)NC(=O)CN(c1nc2cc(nn2c2c1cccc2)C)C", "adduct": "[M-H]-"},
        {"smiles": "CCOc1ccc(cc1)S(=O)(=O)N(c1ccc(cc1)OCC(=O)Nc1scc(n1)c1cc(OC)ccc1OC)C", "adduct": "[M-H]-"},
        {"smiles": "O=C(C1CCCN1S(=O)(=O)c1c(C)noc1C)Nc1ccc(c(c1)Cl)C", "adduct": "[M-H]-"},
        {"smiles": "O=C(C1CN(C(=O)C1)c1ccc2c(c1)OCCO2)Nc1ccccc1N1CCOCC1", "adduct": "[M-H]-"},
        {"smiles": "Cc1ccc(c(c1)CN1C(=O)NC(C1=O)(C)c1ccccc1Cl)C", "adduct": "[M-H]-"},
        {"smiles": "O=C(Nc1sc(c(c1C(=O)N1CCOCC1)C)c1ccccc1)CN1CCc2c(C1)cccc2", "adduct": "[M-H]-"},
        {"smiles": "CC1OCCN(C1)Cn1nc(n(c1=S)c1ccccc1)c1ccncc1", "adduct": "[M-H]-"},
        {"smiles": "C=CCn1c(SCC(=O)c2ccc(s2)Cl)nnc1COc1ccccc1F", "adduct": "[M-H]-"},
        {"smiles": "Cc1ccc(s1)CN(C(=O)Cn1cnc2c1c(=O)n(C)c(=O)n2C)CCc1ccccc1", "adduct": "[M-H]-"},
        {"smiles": "CC(c1nnc(o1)c1ccc(cc1)[N+](=O)[O-])Sc1nnc(o1)c1ccc2c(c1)OCO2", "adduct": "[M-H]-"},
        {"smiles": "O=S(=O)(c1cccc(c1)c1nnc(o1)SCCc1ccccc1)N1CCOCC1", "adduct": "[M-H]-"},
        {"smiles": "CCN(C(=O)C1CCN(CC1)c1ccc2n(n1)nnn2)CC", "adduct": "[M-H]-"},
        {"smiles": "O=C(C1CCN(CC1)S(=O)(=O)c1c(C)noc1C)NC1CCCCCC1", "adduct": "[M-H]-"},
        {"smiles": "CCN1C(=O)C(=O)N(C1=O)CC(=O)c1ccc(cc1)C(C)C", "adduct": "[M-H]-"},
        {"smiles": "COc1ccc(cc1)c1noc(c1)Cn1nnc(n1)c1ccc(cc1)Cl", "adduct": "[M-H]-"},
        {"smiles": "Brc1ccc(cc1)c1ccc(cc1)OS(=O)(=O)c1ccc2c(c1)oc(=O)n2C", "adduct": "[M-H]-"},
        {"smiles": "COc1cc2CN(CCc2cc1OC)CC(COCC1CCCO1)O", "adduct": "[M-H]-"},
        {"smiles": "COc1ccccc1CSc1nc2c([nH]1)cc(cc2)Cl", "adduct": "[M-H]-"},
        {"smiles": "O=C(C1CCN(CC1)c1ccc2n(n1)c(nn2)C(F)(F)F)NCCc1ccc2c(c1)OCCO2", "adduct": "[M-H]-"},
        {"smiles": "CCCCn1ccnc1C(C(F)(F)F)O", "adduct": "[M-H]-"},
        {"smiles": "C=CCn1c(SCC(=O)NC(=O)Nc2ccc3c(c2)OCCO3)nnc1C", "adduct": "[M-H]-"},
        {"smiles": "CCCCc1ccc(cc1)NC(=O)CNC(=O)c1ccc(c(c1)OC)OC", "adduct": "[M-H]-"},
        {"smiles": "COC(=O)CNC(=O)C1CCCN1S(=O)(=O)c1cc(C)c(cc1C)Cl", "adduct": "[M-H]-"},
        {"smiles": "O=C(N1CCCC1c1nc2c(s1)cccc2)c1cccc(c1)S(=O)(=O)NCC1CCCO1", "adduct": "[M-H]-"},
        {"smiles": "O=C(c1ccc(c(c1)S(=O)(=O)Nc1ccccc1N1CCOCC1)Cl)N1CCOCC1", "adduct": "[M-H]-"},
        {"smiles": "Clc1c(cnn(c1=O)c1ccccc1)N1CCN(CC1)Cc1ccc2c(c1)OCO2", "adduct": "[M-H]-"},
        {"smiles": "O=C(C(Sc1nc2ccsc2c(=O)n1C)C)NC1CCCCCC1", "adduct": "[M-H]-"},
        {"smiles": "O=C(Nc1nnc(s1)c1cccnc1)COc1ccc(cc1)Cl", "adduct": "[M-H]-"},
        {"smiles": "CCOc1ccc(cc1)C(=O)N[C@H](C(=O)Nc1cccc(c1)OC)C(C)C", "adduct": "[M-H]-"},
        {"smiles": "CC1CC(C)(C)CC2(C1)NC(=O)N(C2=O)CC(=O)Nc1ccc(c(c1F)F)F", "adduct": "[M-H]-"},
        {"smiles": "O=C(CN1C(=O)NC2(C1=O)CCc1c2cccc1)NC1CCCCC1", "adduct": "[M-H]-"},
        {"smiles": "CCc1cccc(c1NC(=O)C1CN(C(=O)C1)Cc1ccccc1)CC", "adduct": "[M-H]-"},
        {"smiles": "O=C([C@@H]1CCCN1S(=O)(=O)c1ccccc1F)Nc1ccc(c(c1)F)C", "adduct": "[M-H]-"},
        {"smiles": "CC(NC(=O)COC(=O)c1c(C)c(nc2c1cccc2)c1cccc(c1)Cl)C", "adduct": "[M-H]-"},
        {"smiles": "O=C(COc1ccc(cc1)c1ccccc1)NCC(c1ccc(cc1)F)N1CCOCC1", "adduct": "[M-H]-"},
        {"smiles": "O=C(NC12CC3CC(C2)CC(C1)C3)CN1C(=O)N(C(C1=O)C)c1ccc(cc1)C", "adduct": "[M-H]-"},
        {"smiles": "O=C(CN1C(=O)NC2(C1=O)CCCCC2)NC(=O)NC(C)(C)C", "adduct": "[M-H]-"},
        {"smiles": "CCOc1cc(NC(=O)c2ccccc2)c(cc1NC(=O)Cn1ccccc1=O)OCC", "adduct": "[M-H]-"},
        {"smiles": "CCCCCCn1nc(C(=O)OCc2cc(=O)oc3c2ccc(c3)OC)c2c(c1=O)cccc2", "adduct": "[M-H]-"},
        {"smiles": "COc1ccccc1C1CC(=NN1S(=O)(=O)C)c1ccccc1", "adduct": "[M-H]-"},
        {"smiles": "O=C(C1=C(C)NC(=S)NC1c1ccc(cc1)C)N1CCOCC1", "adduct": "[M-H]-"},
        {"smiles": "COc1cccc(c1)c1nnn(n1)c1nc(CN2CCOCC2)nc2c1c(C)c(s2)C", "adduct": "[M-H]-"},
        {"smiles": "CCN1C(=O)CC(=O)N(C1=S)CC", "adduct": "[M-H]-"},
        {"smiles": "CC(=O)Nc1ccc(cc1)NC(c1nnc(o1)c1ccccc1)C", "adduct": "[M-H]-"},
        {"smiles": "O=C(NC(C12CC3CC(C2)CC(C1)C3)C)CSc1nnc(o1)c1cccc(c1)S(=O)(=O)N1CCCCC1", "adduct": "[M-H]-"},
        {"smiles": "Cc1ccc(c(c1)C)CN(C(=O)Cc1c[nH]c2c1cccc2)C", "adduct": "[M-H]-"},
        {"smiles": "CCCN(C(=O)COc1ccccc1C#N)Cc1nnc(o1)c1ccccc1Cl", "adduct": "[M-H]-"},

        # [M+Na]+ cases (130 compounds)
        {"smiles": "Clc1ccc(cc1)c1occ(n1)CSc1nnnn1CCc1cccs1", "adduct": "[M+Na]+"},
        {"smiles": "N#CC1(CCCC1)NC(=O)CSc1ccc(cn1)S(=O)(=O)N1CCCCC1", "adduct": "[M+Na]+"},
        {"smiles": "Clc1ccc(cc1)c1nnc(o1)C(=O)c1ccncc1", "adduct": "[M+Na]+"},
        {"smiles": "COc1ccc(cc1)OCCOc1ccc(cc1)NC(=O)C", "adduct": "[M+Na]+"},
        {"smiles": "O=C(N1CCCSc2c1cccc2)Cc1cccs1", "adduct": "[M+Na]+"},
        {"smiles": "C=CCN1CC2=C(C1=O)C(NC(=O)N2)c1ccc(cc1)O", "adduct": "[M+Na]+"},
        {"smiles": "CC(=O)N1CCN(CC1)c1ccc(c(c1)C(F)(F)F)[N+](=O)[O-]", "adduct": "[M+Na]+"},
        {"smiles": "O=C(c1cc(n(c1C)C)C)COC(=O)c1c(F)cccc1F", "adduct": "[M+Na]+"},
        {"smiles": "COC(=O)c1cccc(c1)OCC(=O)Nc1ccc2c(c1)OCCO2", "adduct": "[M+Na]+"},
        {"smiles": "Brc1ccc(cc1)S(=O)(=O)N1CCCC1c1ccccc1", "adduct": "[M+Na]+"},
        {"smiles": "NC(=O)C1CCCN1c1nc(nc2c1cccc2)c1cccnc1", "adduct": "[M+Na]+"},
        {"smiles": "CNC(=O)c1cccc(c1)NC(=O)c1ccc2c(c1)cccc2", "adduct": "[M+Na]+"},
        {"smiles": "Cc1ccc(c(c1)NC(=O)c1nc2n(n1)c(C)cc(n2)C)F", "adduct": "[M+Na]+"},
        {"smiles": "O=C(C1COc2c(O1)cccc2)OC(C(=O)N1CCc2c1cccc2)C", "adduct": "[M+Na]+"},
        {"smiles": "Brc1ccc(=O)n(c1)CC(=O)N(C(C)(C)C)Cc1ccccc1", "adduct": "[M+Na]+"},
        {"smiles": "COc1cccc(c1)CNC(=O)COC(=O)c1cc(Cl)c(cc1OC)N", "adduct": "[M+Na]+"},
        {"smiles": "CC(c1nnc(o1)c1cccs1)Sc1nnc(o1)c1ccco1", "adduct": "[M+Na]+"},
        {"smiles": "O=C(Cn1ccccc1=O)Nc1ccc2c(c1)OC1(O2)CCCC1", "adduct": "[M+Na]+"},
        {"smiles": "O=C(c1ccc(cc1)NC(=O)C)OCC(=O)NC1CCCCCC1", "adduct": "[M+Na]+"},
        {"smiles": "OCCNC(=O)c1ccc(cc1)S(=O)(=O)N(CC)CC", "adduct": "[M+Na]+"},
        {"smiles": "COc1ccc(cc1)OCCCn1cc(ccc1=O)S(=O)(=O)N1CCCC1", "adduct": "[M+Na]+"},
        {"smiles": "O=C(CNC(=O)C1CCCCC1)NCCSCc1ccc(s1)Br", "adduct": "[M+Na]+"},
        {"smiles": "CCc1ccccc1NC(=O)c1cc(ccc1Br)S(=O)(=O)N", "adduct": "[M+Na]+"},
        {"smiles": "O=C(CN1C(=O)NC(C1=O)(C)c1ccccc1Br)NC1CCCCC1", "adduct": "[M+Na]+"},
        {"smiles": "O=C(N(C1=CCCCC1)Cc1ccco1)CSc1nnnn1C1CCCC1", "adduct": "[M+Na]+"},
        {"smiles": "Clc1ccccc1c1nnc(o1)CN(C(=O)c1sccc1C)C1CC1", "adduct": "[M+Na]+"},
        {"smiles": "O=C(CN1C(=O)NC2(C1=O)CCc1c2cccc1)NC1CCCc2c1cccc2", "adduct": "[M+Na]+"},
        {"smiles": "O=C(C(Sc1nc(cn1N)c1ccccc1)C)Nc1ccccc1c1ccccc1", "adduct": "[M+Na]+"},
        {"smiles": "O=C(C1CC2CCCC(C1)C12SCCS1)NCCc1cccs1", "adduct": "[M+Na]+"},
        {"smiles": "O=C(Cn1c(=O)c(F)cn(c1=O)C1CCCO1)Nc1ccccc1Oc1ccccc1", "adduct": "[M+Na]+"},
        {"smiles": "CCN(C(=O)CN1C(=O)NC(C1=O)(C)c1ccc(cc1)C#N)CC(=C)C", "adduct": "[M+Na]+"},
        {"smiles": "CN(C(=O)c1ccc(cc1)c1ccccc1)CC1COc2c(O1)cccc2", "adduct": "[M+Na]+"},
        {"smiles": "O=C(CSc1nnc(n1c1ccccc1)c1ccc(cc1)F)NCC1CCCO1", "adduct": "[M+Na]+"},
        {"smiles": "CCOc1cccc2c1oc(c2)C(NS(=O)(=O)c1ccc2c(c1)OCCO2)C", "adduct": "[M+Na]+"},
        {"smiles": "O=C(C(c1ccccc1)c1ccccc1)NCc1ccccc1CN1CCOCC1", "adduct": "[M+Na]+"},
        {"smiles": "CC(C12CC3CC(C2)CC(C1)C3)NC(=O)Cc1ccc(cc1)NC(=O)C", "adduct": "[M+Na]+"},
        {"smiles": "O=C(NC12CC3CC(C2)CC(C1)C3)CSc1nc2c(s1)ccc(c2)Cl", "adduct": "[M+Na]+"},
        {"smiles": "O=C(N(C(c1ccccc1)C)C)CSc1nnc(n1CC1CCCO1)c1cccs1", "adduct": "[M+Na]+"},
        {"smiles": "CCCN(Cc1nnnn1c1ccccc1)Cc1nnc(o1)c1cccs1", "adduct": "[M+Na]+"},
        {"smiles": "CCCC1(NC(=O)N(C1=O)CC(=O)NC(c1ccc2c(c1)OCCO2)C)c1ccccc1", "adduct": "[M+Na]+"},
        {"smiles": "O=C(c1ccc(cc1)S(=O)(=O)N(C)C)NCCCn1ccc2c1cccc2", "adduct": "[M+Na]+"},
        {"smiles": "C=CCn1c(SC(C(=O)Nc2ccc3c(c2)OCO3)C)nc2c(c1=O)c(C)c(s2)C", "adduct": "[M+Na]+"},
        {"smiles": "O=C(Nc1ccc2c(c1)CCN2C(=O)C)COc1ccccc1c1ccccc1", "adduct": "[M+Na]+"},
        {"smiles": "O=C(C(Sc1nnc(n1c1ccccc1)c1ccco1)C)N1CCc2c(C1)cccc2", "adduct": "[M+Na]+"},
        {"smiles": "O=C(N(C1CCCc2c1cccc2)C)CSc1nnc(n1Cc1ccccc1)C1CC1", "adduct": "[M+Na]+"},
        {"smiles": "O=C(CN1CC(Oc2c1cccc2)C(=O)N1CCOCC1)NC(=O)NC1CCCC1", "adduct": "[M+Na]+"},
        {"smiles": "O=S1(=O)CCC(C1)Nc1nc(nc2c1c(cs2)c1ccccc1)c1cccnc1", "adduct": "[M+Na]+"},
        {"smiles": "O=C(C1CCCN(C1)C(=O)c1ccc(cc1)Cl)NCC(N1CCCCC1)(C)C", "adduct": "[M+Na]+"},
        {"smiles": "COc1ccc(cc1)NC(=O)CC1C(=O)Nc2c(N1CC(=O)N1CCCC1)cccc2", "adduct": "[M+Na]+"},
        {"smiles": "CCOc1ccc(cc1)C(=O)NC(C(=O)Nc1ccc(cc1)Cl)C(C)C", "adduct": "[M+Na]+"},
        {"smiles": "O=C(Nc1ccc(cc1)C)Cc1scc(n1)Cn1nc(oc1=O)c1ccc(cc1)F", "adduct": "[M+Na]+"},
        {"smiles": "O=S(=O)(c1cccc(c1)c1nnc(o1)SCCCOc1ccccc1)N1CCOCC1", "adduct": "[M+Na]+"},
        {"smiles": "COc1ccc(cc1)C1=NN(C(C1)c1ccco1)C(=O)Cn1nnc(n1)c1cccs1", "adduct": "[M+Na]+"},
        {"smiles": "O=C(N1CCCC1)CSc1ccccc1C(=O)N1CCN(CC1)c1ccccc1Cl", "adduct": "[M+Na]+"},
        {"smiles": "O=S(=O)(N(C1CCCC1)Cc1ccc2c(c1)OCO2)c1ccc(cc1)S(=O)(=O)N1CCCC1", "adduct": "[M+Na]+"},
        {"smiles": "Fc1cccc(c1)n1c(SCC(=O)N2CCC3C(C2)CCCC3)nnc1N1CCOCC1", "adduct": "[M+Na]+"},
        {"smiles": "O=C(N1CCC(CC1)Cc1ccccc1)CSc1nnc(o1)c1ccc2c(c1)OCO2", "adduct": "[M+Na]+"},
        {"smiles": "CC(=O)N1CCN(CC1)c1cc2c(cc1F)c(=O)c(cn2C1CC1)C(=O)OCc1cccc(c1F)F", "adduct": "[M+Na]+"},
        {"smiles": "COc1ccc(cc1)n1c2nnc(n2c2c(c1=O)cccc2)SCC(=O)N1CCCCC1C", "adduct": "[M+Na]+"},
        {"smiles": "Brc1ccc(s1)c1nnc(o1)CN1C(=O)NC(C1=O)(C)c1ccc2c(c1)OCO2", "adduct": "[M+Na]+"},
        {"smiles": "COc1cc2CN(CCc2cc1OC)C(=O)CC12CC3CC(C2)CC(C1)C3", "adduct": "[M+Na]+"},
        {"smiles": "O=c1cc(CN2C(=O)NC(C2=O)(C)c2ccc3c(c2)OCO3)c2c(o1)cc1c(c2)CCC1", "adduct": "[M+Na]+"},
        {"smiles": "O=C(N1CCSC1c1ccccc1F)CSc1nncn1c1ccccc1", "adduct": "[M+Na]+"},
        {"smiles": "Cc1ccc(cc1)c1nnn(n1)CC(=O)c1cc(n(c1C)CC1CCCO1)C", "adduct": "[M+Na]+"},
        {"smiles": "O=C(N(Cc1ccccc1)Cc1ccccc1)COc1ccc(cc1)c1nnco1", "adduct": "[M+Na]+"},
        {"smiles": "O=C(NC(C(=O)C)Cc1ccccc1)CSc1nnc(n1C)c1ccccc1F", "adduct": "[M+Na]+"},
        {"smiles": "COc1cc(cc(c1OC)OC)C(=O)NCC(=O)N1CCCCCCC1", "adduct": "[M+Na]+"},
        {"smiles": "O=C(CN1C(=O)NC2(C1=O)CCCc1c2cccc1)NC1CCCCC1", "adduct": "[M+Na]+"},
        {"smiles": "COc1ccc(cc1S(=O)(=O)NC1CC1)C(=O)Nc1ccc(cc1)c1ccccc1", "adduct": "[M+Na]+"},
        {"smiles": "COc1cc(ccc1OC)C(=O)Nc1ccccc1C(=O)NCc1ccco1", "adduct": "[M+Na]+"},
        {"smiles": "CN(Cc1ccccc1)CC(=O)Nc1ccc(cc1Cl)C", "adduct": "[M+Na]+"},
        {"smiles": "COc1ccc(cc1S(=O)(=O)NC1CC1)C(=O)NCCC(C)C", "adduct": "[M+Na]+"},
        {"smiles": "O=C1N(c2ccccc2)C(Nc2c1cccc2)c1c[nH]nc1c1cccnc1", "adduct": "[M+Na]+"},
        {"smiles": "Fc1ccc(cc1)S(=O)(=O)N1CCCC1C(=O)N1CCN(CC1)S(=O)(=O)c1cccs1", "adduct": "[M+Na]+"},
        {"smiles": "COc1cc(Br)cc2c1oc(=O)c(c2)C(=O)N1CCc2c(C1)cccc2", "adduct": "[M+Na]+"},
        {"smiles": "COc1cc2CN(CCc2cc1OC)S(=O)(=O)c1ccccc1Cl", "adduct": "[M+Na]+"},
        {"smiles": "OC(CN1C(=O)NC2(C1=O)CCCCC2)COc1c(C)cccc1C", "adduct": "[M+Na]+"},
        {"smiles": "CC(=O)N1CCN(CC1)c1ccc(cc1)Oc1nc(nc2c1cccc2)c1cccnc1", "adduct": "[M+Na]+"},
        {"smiles": "CC(=O)NC(C(=O)Nc1ccc(cc1)c1nnc2n1CCCCC2)C(C)C", "adduct": "[M+Na]+"},
        {"smiles": "COc1ccc(cc1)N1CCN(CC1)CCC#N", "adduct": "[M+Na]+"},
        {"smiles": "CCN(C(=O)CCNC(=O)N(Cc1ccc(cc1)c1ccccc1)C)CC", "adduct": "[M+Na]+"},
        {"smiles": "O=C1NC2(C(=O)N1Cc1onc(n1)c1cccs1)CCOc1c2cccc1", "adduct": "[M+Na]+"},
        {"smiles": "COc1cc(ccc1OC(F)F)C=C1C(=O)c2c(C1=O)cccc2", "adduct": "[M+Na]+"},
        {"smiles": "Cc1sc2c(c1C)c(nc(n2)CN1CCOCC1)n1nnc(n1)c1ccccc1", "adduct": "[M+Na]+"},
        {"smiles": "Cc1cccc(c1C)n1c(Sc2nnnn2c2ccccc2)nc2c(c1=O)cccc2", "adduct": "[M+Na]+"},
        {"smiles": "CONC(=O)c1cc(nn1c1ccccc1)C1CC1", "adduct": "[M+Na]+"},
        {"smiles": "O=C(Nc1cccc(c1C)C)CN(C(=O)CCc1ccc(cc1)C)C", "adduct": "[M+Na]+"},
        {"smiles": "COc1cc(CC(=O)NC2CCCc3c2cccc3)cc(c1OC)OC", "adduct": "[M+Na]+"},
        {"smiles": "COc1ccc(cc1OC)CNC(=O)CN(S(=O)(=O)c1ccc2c(c1)OCCO2)C", "adduct": "[M+Na]+"},
        {"smiles": "Fc1ccc(cc1)C(N1CCOCC1)CNC(=O)c1cccnc1Sc1ccc(cc1)Cl", "adduct": "[M+Na]+"},
        {"smiles": "O=C(N1CCN(CC1)C(=O)c1oc2c(c1COc1ccccc1)cccc2)C1CC1", "adduct": "[M+Na]+"},
        {"smiles": "O=C(N1CCOCC1)CCCSc1nnc(n1c1ccccc1)COc1ccccc1F", "adduct": "[M+Na]+"},
        {"smiles": "COc1ccc(cc1)S(=O)(=O)N1N=C(CC1c1ccc2c(c1)nccn2)c1cccs1", "adduct": "[M+Na]+"},
        {"smiles": "CC(C(=O)NC1CCCC1)OC(=O)c1ccccc1", "adduct": "[M+Na]+"},
        {"smiles": "CCOc1ccc(cc1)NC(=O)C(C(CC)C)NC(=O)c1ccccc1", "adduct": "[M+Na]+"},
        {"smiles": "O=C1NC(C(=O)N1Cc1cccc(c1)S(=O)(=O)N1CCOCC1)(C)c1ccc2c(c1)OCO2", "adduct": "[M+Na]+"},
        {"smiles": "Brc1ccccc1c1nnc(o1)Sc1ncnc2c1ccs2", "adduct": "[M+Na]+"},
        {"smiles": "O=C(N1CCCCCCC1)CN1C(=O)NC(C1=O)(C)c1cc2c(o1)cccc2", "adduct": "[M+Na]+"},
        {"smiles": "COc1cc(NC(=O)CC2CC3CC2CC3)cc(c1)OC", "adduct": "[M+Na]+"},
        {"smiles": "CC=Cc1cc(cc(c1OC)OC)C(=O)NC1CC1", "adduct": "[M+Na]+"},
        {"smiles": "O=C(N1CCN(CC1)c1ncccn1)C=Cc1ccc2c(c1)OCCCO2", "adduct": "[M+Na]+"},
        {"smiles": "CN(CC(=O)Nc1ccccc1C(=O)C)CC(=O)Nc1ccc(cc1)C", "adduct": "[M+Na]+"},
        {"smiles": "O=C(c1cnc2c(c1)cnn2C(C)C)Nc1sc2c(n1)CCC2", "adduct": "[M+Na]+"},
        {"smiles": "O=C(N1CCN(CC1)c1ncc(cc1Cl)C(F)(F)F)COc1ccccc1C", "adduct": "[M+Na]+"},
        {"smiles": "COC(=O)C(NC(=O)C(c1ccccc1)c1ccccc1)Cc1ccc(cc1)O", "adduct": "[M+Na]+"},
        {"smiles": "CN(Cc1ccc(cc1)C(=O)N)CCOc1ccccc1", "adduct": "[M+Na]+"},
        {"smiles": "O=C(N1CCN(CC1)S(=O)(=O)c1ccc2c(c1)CCC2)C1CCC1", "adduct": "[M+Na]+"},
        {"smiles": "Cc1ccc(cc1)n1c(nn(c1=S)CN(CC(=O)N1CCCC1)C)c1ccc(cc1)Cl", "adduct": "[M+Na]+"},
        {"smiles": "O=C(COc1ccccc1)NCC(=O)NC1CCCc2c1cccc2", "adduct": "[M+Na]+"},
        {"smiles": "CCC(=O)N(c1nnc(s1)SCCOc1ccc(cc1)OC)C1CC1", "adduct": "[M+Na]+"},
        {"smiles": "CCOc1cc(NC(=O)c2ccccc2)c(cc1NC(=O)Cn1ccccc1=O)OCC", "adduct": "[M+Na]+"},
        {"smiles": "CC(C(=O)NCc1ccc2c(c1)OCO2)Sc1nnc(n1Cc1ccccc1)c1ccncc1", "adduct": "[M+Na]+"},
        {"smiles": "O=C(NC(C)(C)C)CN1CCCN(CC1)c1ncnc2c1cc(s2)c1ccccc1", "adduct": "[M+Na]+"},
        {"smiles": "O=C(Nc1c(C)n(n(c1=O)c1ccccc1)C)CSc1ccccc1C(=O)NC1=NCCS1", "adduct": "[M+Na]+"},
        {"smiles": "N#Cc1c(NC(=O)C2CCCO2)sc2c1CCCC2", "adduct": "[M+Na]+"},
        {"smiles": "COCc1c(sc2c1c(F)ccc2)C(=O)OCC(=O)c1cc(n(c1C)C(COC)C)C", "adduct": "[M+Na]+"},
        {"smiles": "O=C(C1CN(C(=O)C1)Cc1ccco1)N1CCN(CC1)C(=O)c1cccs1", "adduct": "[M+Na]+"},
        {"smiles": "COC(=O)C1(CCCCC1)NS(=O)(=O)N1CCCC(C1)C", "adduct": "[M+Na]+"},
        {"smiles": "COC(=O)c1ccc(cc1)NC(=O)COC(=O)c1ccccc1c1ccc(cc1)C(F)(F)F", "adduct": "[M+Na]+"},
        {"smiles": "O=C(Nc1sc(c(c1C(=O)N1CCOCC1)C)c1ccccc1)CSc1nncn1C", "adduct": "[M+Na]+"},
        {"smiles": "O=C(CSc1ccccc1C(=O)Nc1cccc(c1)F)NCc1ccco1", "adduct": "[M+Na]+"},
        {"smiles": "COc1ccc(cc1)OCc1nnc(n1c1ccccc1)SCc1nnc(o1)C", "adduct": "[M+Na]+"},
        {"smiles": "CCn1c(Cn2nnc(n2)c2cscc2)nc2c1ccc(c2)S(=O)(=O)N1CCOCC1", "adduct": "[M+Na]+"},
        {"smiles": "O=C(Nc1ccccc1C)CN1C(=O)NC(C1=O)(C)c1ccc(cc1)C(C)(C)C", "adduct": "[M+Na]+"},
        {"smiles": "COc1cc(ccc1OCc1ccccc1)C(=O)N(Cc1ccccc1)C", "adduct": "[M+Na]+"},
        {"smiles": "NC(=O)C1CCN(CC1)C(=O)c1cc(nn1c1ccccc1)c1ccccc1", "adduct": "[M+Na]+"},
        {"smiles": "C=CCn1c(SCC(=O)N2CCC(CC2)C)nc2c(c1=O)c(cs2)c1ccco1", "adduct": "[M+Na]+"},
        {"smiles": "CN(Cc1[nH]c(=O)c2c(n1)sc1c2CCC1)Cc1ccc(cc1)Cl", "adduct": "[M+Na]+"},
        {"smiles": "O=C(c1cccc(c1)N1CCCC1=O)N1CCN(CC1)S(=O)(=O)c1ccc2c(c1)CCC2", "adduct": "[M+Na]+"},
        {"smiles": "CCN1CCN(CC1)C1=NS(=O)(=O)c2c1cccc2", "adduct": "[M+Na]+"},
    ]

    prompts_func = {
        "simple": lambda s, a: build_prompt_simple(s, a),
        "expert knowledge": lambda s, a: build_prompt_expert(s, a),
        "examples": lambda s, a: build_prompt_examples(s, a, TRAIN_DATA, n=5),
        "combined": lambda s, a: build_prompt_combined(s, a, TRAIN_DATA, n=5),
    }

    variants = list(prompts_func.keys())

    # Initialize Excel with timestamp to avoid overwriting previous runs
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    excel_path = f"../data/results/prompt_predictions_{timestamp}.xlsx"
    os.makedirs(os.path.dirname(excel_path), exist_ok=True)
    initialize_excel(excel_path, variants)

    print("=" * 70)
    print(f"PROMPT TEST\nTest start: {datetime.now()}")

    for i, case in enumerate(test_cases, 1):
        smiles, adduct = case["smiles"], case["adduct"]
        print(f"{'=' * 70}")
        print(f"COMPOUND {i} | Adduct = {adduct}")
        print(f"SMILES: {smiles[:60]}...")
        print(f"{'=' * 70}")

        # Clear cache between predictions
        torch.cuda.empty_cache()  # just in case, even on CPU
        if hasattr(MODEL, 'reset_cache'):
            MODEL.reset_cache()

        # Create the base row for the compound (empty predictions, filled in below)
        base_row = [i, smiles, adduct] + [None] * len(variants)
        save_row_excel(excel_path, base_row)
        # Header = row 1, so compound i goes to row i+1
        row_idx = i + 1

        for j, (name, fn_prompt) in enumerate(prompts_func.items()):
            prompt = fn_prompt(smiles, adduct)
            print(f"Prompt {name}: ")
            print(f" - Execution time: {datetime.now().strftime('%H:%M:%S')}")

            examples = select_examples(TRAIN_DATA, smiles, adduct, n=5)
            result = predict_ccs(MODEL, TOKENIZER, prompt, STATS, examples)

            fallback_str = " [FALLBACK]" if result["fallback"] else ""
            print(f" - CCS = {result['predicted_ccs']:.2f} Å²{fallback_str}")
            print(f" - Reasoning: {result['reasoning'][:80]}")

            # Save result immediately to Excel
            # Columns: 1=id, 2=smiles, 3=adduct, 4..=variants
            col_idx = 4 + j
            update_cell_excel(excel_path, row_idx, col_idx, result["predicted_ccs"])

    print(f"\n{'=' * 70}")
    print(f"TEST COMPLETE!\nTest end: {datetime.now()}")
    print("=" * 70)


# For archiving results
class Logger:
    def __init__(self, filepath):
        self.terminal = sys.stdout
        os.makedirs(os.path.dirname(filepath), exist_ok=True)
        self.log = open(filepath, 'w', encoding='utf-8')

    def write(self, message):
        self.terminal.write(message)
        self.log.write(message)
        self.log.flush()

    def flush(self):
        self.terminal.flush()
        self.log.flush()

    def close(self):
        self.log.close()


if __name__ == '__main__':
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    log_path = f"../data/results/log_{timestamp}.txt"
    sys.stdout = Logger(log_path)
    print(f"Log saved to: {log_path}")

    os.environ["TRANSFORMERS_VERBOSITY"] = "error"
    if initialize_app():
        test_prompts()
    else:
        print("ERROR during initialization. Check the configuration.")

    sys.stdout.close()
    sys.stdout = sys.stdout.terminal
